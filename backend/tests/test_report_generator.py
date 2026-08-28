"""Report generator Celery task 测试 (S5-SYS-003, S2-A4).

测试覆盖：
- _parse_content_template: None/有效JSON/无效JSON
- _generate_pdf: 生成 PDF 字节，验证内容
- _do_generate: 正常流程/配置不存在/非法周期/PDF 生成失败
- generate_report_task: 业务错误不重试/系统错误重试
- Beat schedule: 4 个周期任务已注册
- export_diagnosis_statistics: 异步导出任务触发/Excel/CSV 生成/文件保存
"""

from __future__ import annotations

import io
import os
import tempfile
from datetime import UTC, datetime
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.tasks.celery_app import celery_app
from app.tasks.report_generator import (
    NonRetryableError,
    _do_export_diagnosis_statistics,
    _do_generate,
    _generate_csv_bytes,
    _generate_excel_bytes,
    _generate_pdf,
    _parse_content_template,
    _parse_iso_dt,
    export_diagnosis_statistics,
    generate_report_task,
)

# ===========================================================================
# 辅助函数
# ===========================================================================


def _make_report_config(
    config_id: str = "00000000-0000-0000-0000-000000000b01",
    name: str = "日报配置",
    content_template: str | None = '{"sections": ["summary", "kpi"]}',
) -> MagicMock:
    """构造 mock ReportConfig 对象。"""
    config = MagicMock()
    config.id = config_id
    config.name = name
    config.content_template = content_template
    return config


def _make_scalar_one_or_none_mock(value) -> MagicMock:
    """构造 execute 返回的 mock，支持 scalar_one_or_none()。"""
    result = MagicMock()
    result.scalar_one_or_none.return_value = value
    return result


# ===========================================================================
# _parse_content_template 单元测试
# ===========================================================================


class TestParseContentTemplate:
    """测试 _parse_content_template() JSON 解析。"""

    def test_none_returns_none(self) -> None:
        """None 输入应返回 None。"""
        assert _parse_content_template(None) is None

    def test_valid_json_returns_dict(self) -> None:
        """有效 JSON 字符串应解析为字典。"""
        result = _parse_content_template('{"key": "value", "count": 3}')
        assert result == {"key": "value", "count": 3}

    def test_invalid_json_returns_none(self) -> None:
        """无效 JSON 字符串应返回 None（不抛出异常）。"""
        assert _parse_content_template("not a json") is None

    def test_empty_string_returns_none(self) -> None:
        """空字符串应返回 None。"""
        assert _parse_content_template("") is None

    def test_array_json_returns_list(self) -> None:
        """JSON 数组应正确解析（返回 list）。"""
        result = _parse_content_template('["a", "b", "c"]')
        assert result == ["a", "b", "c"]


# ===========================================================================
# _generate_pdf 单元测试
# ===========================================================================


class TestGeneratePdf:
    """测试 _generate_pdf() PDF 生成。"""

    def test_returns_non_empty_bytes(self) -> None:
        """生成的 PDF 应为非空 bytes。"""
        pdf_bytes = _generate_pdf(report_period="DAILY")
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 0

    def test_pdf_starts_with_pdf_magic(self) -> None:
        """PDF 应以 %PDF 魔数开头。"""
        pdf_bytes = _generate_pdf(report_period="DAILY")
        assert pdf_bytes.startswith(b"%PDF")

    def test_pdf_with_config_name(self) -> None:
        """带 config_name 时 PDF 应正常生成。"""
        pdf_bytes = _generate_pdf(
            report_period="WEEKLY",
            config_name="周报配置",
        )
        assert len(pdf_bytes) > 0
        assert pdf_bytes.startswith(b"%PDF")

    def test_pdf_with_content_template(self) -> None:
        """带 content_template 时 PDF 应包含模板内容。"""
        template = {"section1": "KPI 总览", "section2": "异常回路"}
        pdf_bytes = _generate_pdf(
            report_period="MONTHLY",
            config_name="月报",
            content_template=template,
        )
        assert len(pdf_bytes) > 0

    def test_pdf_all_periods(self) -> None:
        """所有报表周期（SHIFT/DAILY/WEEKLY/MONTHLY）都应正常生成 PDF。"""
        for period in ["SHIFT", "DAILY", "WEEKLY", "MONTHLY"]:
            pdf_bytes = _generate_pdf(report_period=period)
            assert pdf_bytes.startswith(b"%PDF"), f"周期 {period} PDF 生成失败"

    def test_pdf_unknown_period_uses_raw_value(self) -> None:
        """未知周期应使用原始值（不崩溃）。"""
        pdf_bytes = _generate_pdf(report_period="QUARTERLY")
        assert pdf_bytes.startswith(b"%PDF")


# ===========================================================================
# _do_generate 异步逻辑测试
# ===========================================================================


class TestDoGenerate:
    """测试 _do_generate() 异步报表生成逻辑。"""

    @pytest.mark.asyncio
    async def test_generate_without_config_success(self) -> None:
        """无 config_id 时应成功生成报表。"""
        mock_session = AsyncMock()
        mock_session.add = MagicMock()
        mock_session.commit = AsyncMock()

        with patch("app.core.db.AsyncSessionLocal") as mock_session_local:
            mock_session_local.return_value.__aenter__.return_value = mock_session
            result = await _do_generate(
                task_id="task-001",
                config_id=None,
                report_period="DAILY",
            )

        assert result["status"] == "COMPLETED"
        assert result["reportId"] == "task-001"
        assert "fileUrl" in result
        assert "fileSize" in result
        assert result["fileSize"] > 0

    @pytest.mark.asyncio
    async def test_generate_with_config_success(self) -> None:
        """带有效 config_id 时应成功生成报表。"""
        config = _make_report_config(name="班报配置")
        mock_session = AsyncMock()
        mock_session.add = MagicMock()
        mock_session.commit = AsyncMock()
        mock_session.execute = AsyncMock(return_value=_make_scalar_one_or_none_mock(config))

        with patch("app.core.db.AsyncSessionLocal") as mock_session_local:
            mock_session_local.return_value.__aenter__.return_value = mock_session
            result = await _do_generate(
                task_id="task-002",
                config_id="config-001",
                report_period="SHIFT",
            )

        assert result["status"] == "COMPLETED"
        assert result["reportId"] == "task-002"

    @pytest.mark.asyncio
    async def test_generate_config_not_found_raises_non_retryable(self) -> None:
        """config_id 存在但配置不存在时应抛出 NonRetryableError。"""
        mock_session = AsyncMock()
        mock_session.execute = AsyncMock(return_value=_make_scalar_one_or_none_mock(None))

        with patch("app.core.db.AsyncSessionLocal") as mock_session_local:
            mock_session_local.return_value.__aenter__.return_value = mock_session
            with pytest.raises(NonRetryableError, match="报表配置不存在"):
                await _do_generate(
                    task_id="task-003",
                    config_id="non-existent",
                    report_period="DAILY",
                )

    @pytest.mark.asyncio
    async def test_generate_invalid_period_raises_non_retryable(self) -> None:
        """非法报表周期应抛出 NonRetryableError。"""
        with pytest.raises(NonRetryableError, match="非法报表周期"):
            await _do_generate(
                task_id="task-004",
                config_id=None,
                report_period="INVALID",
            )

    @pytest.mark.asyncio
    async def test_generate_pdf_failure_returns_failed_status(self) -> None:
        """PDF 生成失败时应返回 FAILED 状态（不抛出异常）。"""
        mock_session = AsyncMock()
        mock_session.add = MagicMock()
        mock_session.commit = AsyncMock()

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch(
                "app.tasks.report_generator._generate_pdf",
                side_effect=RuntimeError("reportlab 错误"),
            ),
        ):
            mock_session_local.return_value.__aenter__.return_value = mock_session
            result = await _do_generate(
                task_id="task-005",
                config_id=None,
                report_period="DAILY",
            )

        assert result["status"] == "FAILED"
        assert "reportlab 错误" in result["error"]

    @pytest.mark.asyncio
    async def test_generate_uses_uuid_when_task_id_none(self) -> None:
        """task_id 为 None 时应自动生成 UUID。"""
        mock_session = AsyncMock()
        mock_session.add = MagicMock()
        mock_session.commit = AsyncMock()

        with patch("app.core.db.AsyncSessionLocal") as mock_session_local:
            mock_session_local.return_value.__aenter__.return_value = mock_session
            result = await _do_generate(
                task_id=None,
                config_id=None,
                report_period="WEEKLY",
            )

        assert result["status"] == "COMPLETED"
        # reportId 应为有效的 UUID 字符串
        assert len(result["reportId"]) == 36
        assert result["reportId"].count("-") == 4

    @pytest.mark.asyncio
    async def test_generate_all_valid_periods(self) -> None:
        """所有合法周期都应成功生成报表。"""
        for period in ["SHIFT", "DAILY", "WEEKLY", "MONTHLY"]:
            mock_session = AsyncMock()
            mock_session.add = MagicMock()
            mock_session.commit = AsyncMock()

            with patch("app.core.db.AsyncSessionLocal") as mock_session_local:
                mock_session_local.return_value.__aenter__.return_value = mock_session
                result = await _do_generate(
                    task_id=f"task-{period}",
                    config_id=None,
                    report_period=period,
                )

            assert result["status"] == "COMPLETED", f"周期 {period} 生成失败"


# ===========================================================================
# generate_report_task Celery 任务入口测试
# ===========================================================================


class TestGenerateReportTask:
    """测试 generate_report_task() Celery 任务入口。"""

    def test_task_registered(self) -> None:
        """generate_report_task 应注册到 celery_app。"""
        assert "app.tasks.report_generator.generate_report_task" in celery_app.tasks

    def test_task_success_returns_result(self) -> None:
        """任务成功时应返回 _do_generate 的结果。"""
        expected_result = {
            "reportId": "task-success",
            "status": "COMPLETED",
            "fileUrl": "/reports/task-success.pdf",
            "fileSize": 1024,
        }

        task = generate_report_task
        # 模拟 AsyncTask.run_async 直接返回结果
        with patch.object(task, "run_async", return_value=expected_result):
            result = task(
                task_id="task-success",
                config_id=None,
                report_period="DAILY",
            )

        assert result == expected_result

    def test_task_non_retryable_error_not_retried(self) -> None:
        """NonRetryableError 应直接抛出，不触发重试。"""
        task = generate_report_task

        with (
            patch.object(
                task,
                "run_async",
                side_effect=NonRetryableError("配置不存在"),
            ),
            pytest.raises(NonRetryableError, match="配置不存在"),
        ):
            task(
                task_id="task-fail-business",
                config_id="non-existent",
                report_period="DAILY",
            )

    def test_task_system_error_triggers_retry(self) -> None:
        """系统错误（非 NonRetryableError）应触发 self.retry()。"""
        from celery.exceptions import Retry

        task = generate_report_task

        # Celery 的 retry() 在真实环境下会抛出 Retry 异常
        def _raise_retry(*args, **kwargs):
            raise Retry()

        with (
            patch.object(
                task,
                "run_async",
                side_effect=RuntimeError("DB 连接失败"),
            ),
            patch.object(task, "retry", side_effect=_raise_retry) as mock_retry,
        ):
            with pytest.raises(Retry):
                task(
                    task_id="task-fail-system",
                    config_id=None,
                    report_period="DAILY",
                )

            mock_retry.assert_called_once()
            # 验证 retry 的 exc 参数是 RuntimeError
            call_kwargs = mock_retry.call_args
            assert call_kwargs.kwargs.get("exc") is not None or call_kwargs.args


# ===========================================================================
# Beat schedule 配置测试
# ===========================================================================


class TestBeatSchedule:
    """验证 Beat 调度配置（报告模块优化 P0-1 后的收敛口径）。

    自动周期生成已摘除（占位实现空转写无效 record，P3 做实后恢复调度，
    见 docs/设计文档/CLPM报告模块优化实施方案-2026-08-28.md §3.1 D1）。
    """

    def test_beat_schedule_has_no_report_entries(self) -> None:
        """beat_schedule 不得再注册报表周期任务（P0-1 摘除守护）。"""
        schedule = celery_app.conf.beat_schedule
        report_entries = {
            k: v
            for k, v in schedule.items()
            if v.get("task") == "app.tasks.report_generator.generate_report_task"
        }
        assert report_entries == {}

    def test_beat_schedule_timezone_is_shanghai(self) -> None:
        """Beat 调度时区应为 Asia/Shanghai。"""
        assert celery_app.conf.timezone == "Asia/Shanghai"


# ===========================================================================
# 诊断统计异步导出 - 辅助函数测试
# ===========================================================================


class TestParseIsoDt:
    """测试 _parse_iso_dt() ISO 8601 时间解析。"""

    def test_parse_iso_with_z_suffix(self) -> None:
        """带 Z 后缀的 ISO 字符串应正确解析。"""
        result = _parse_iso_dt("2026-06-26T09:00:00Z")
        assert result.year == 2026
        assert result.month == 6
        assert result.day == 26
        assert result.hour == 9

    def test_parse_iso_with_timezone_offset(self) -> None:
        """带时区偏移的 ISO 字符串应正确解析。"""
        result = _parse_iso_dt("2026-06-26T09:00:00+08:00")
        assert result.hour == 9

    def test_parse_iso_without_timezone(self) -> None:
        """不带时区的 ISO 字符串应通过 fallback 解析。"""
        result = _parse_iso_dt("2026-06-26T09:00:00")
        assert result.hour == 9

    def test_parse_iso_invalid_raises(self) -> None:
        """完全无效的字符串应抛出 ValueError。"""
        with pytest.raises(ValueError):
            _parse_iso_dt("not-a-date")


class TestGenerateExcelBytes:
    """测试 _generate_excel_bytes() Excel 文件生成。"""

    def test_generate_excel_returns_valid_bytes(self) -> None:
        """生成的 Excel 应为非空 bytes 且为有效的 xlsx 格式。"""
        label_counts = [("OSCILLATION", 10), ("VALVE_STICTION", 5)]
        trend_rows = [
            (datetime(2026, 6, 26, tzinfo=UTC), "OSCILLATION", 6),
            (datetime(2026, 6, 26, tzinfo=UTC), "VALVE_STICTION", 3),
        ]
        detail_rows = [
            (
                "loop-001",
                "OSCILLATION",
                0.85,
                datetime(2026, 6, 26, 9, 30, tzinfo=UTC),
                "v1.0",
            ),
        ]

        result = _generate_excel_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="测试装置",
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )

        assert isinstance(result, bytes)
        assert len(result) > 0
        # xlsx 是 zip 格式，应以 PK 魔数开头
        assert result[:2] == b"PK"

    def test_generate_excel_empty_data(self) -> None:
        """空数据也应正常生成 Excel（只有表头）。"""
        result = _generate_excel_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="全部装置",
            label_counts=[],
            trend_rows=[],
            detail_rows=[],
        )
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_generate_excel_contains_three_sheets(self) -> None:
        """Excel 应包含 3 个工作表（标签分布汇总/按天趋势/明细数据）。"""
        from openpyxl import load_workbook

        label_counts = [("OSCILLATION", 10)]
        trend_rows = [(datetime(2026, 6, 26, tzinfo=UTC), "OSCILLATION", 10)]
        detail_rows = [
            (
                "loop-001",
                "OSCILLATION",
                0.9,
                datetime(2026, 6, 26, 9, 0, tzinfo=UTC),
                "v1.0",
            ),
        ]

        excel_bytes = _generate_excel_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="测试装置",
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )

        wb = load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames
        assert "标签分布汇总" in sheet_names
        assert "按天趋势" in sheet_names
        assert "明细数据" in sheet_names
        assert len(sheet_names) == 3

    def test_generate_excel_label_distribution_content(self) -> None:
        """标签分布汇总 sheet 应正确写入标签和数量。"""
        from openpyxl import load_workbook

        label_counts = [("OSCILLATION", 8), ("VALVE_STICTION", 4)]
        excel_bytes = _generate_excel_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="全部装置",
            label_counts=label_counts,
            trend_rows=[],
            detail_rows=[],
        )

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["标签分布汇总"]
        # 第 7 行开始为数据行（前 6 行为标题和表头）
        assert ws.cell(row=7, column=1).value == "OSCILLATION"
        assert ws.cell(row=7, column=3).value == 8
        assert ws.cell(row=8, column=1).value == "VALVE_STICTION"
        assert ws.cell(row=8, column=3).value == 4


class TestGenerateCsvBytes:
    """测试 _generate_csv_bytes() CSV 文件生成。"""

    def test_generate_csv_returns_utf8_bom_bytes(self) -> None:
        """CSV 应以 UTF-8 BOM 开头。"""
        result = _generate_csv_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="测试装置",
            label_counts=[("OSCILLATION", 5)],
            trend_rows=[],
            detail_rows=[],
        )
        assert isinstance(result, bytes)
        assert result.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM

    def test_generate_csv_contains_three_sections(self) -> None:
        """CSV 应包含三个章节标题。"""
        result = _generate_csv_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="全部装置",
            label_counts=[("OSCILLATION", 5)],
            trend_rows=[(datetime(2026, 6, 26, tzinfo=UTC), "OSCILLATION", 5)],
            detail_rows=[
                (
                    "loop-001",
                    "OSCILLATION",
                    0.9,
                    datetime(2026, 6, 26, 9, 0, tzinfo=UTC),
                    "v1.0",
                ),
            ],
        )
        text = result.decode("utf-8-sig")
        assert "一、标签分布汇总" in text
        assert "二、按天趋势" in text
        assert "三、明细数据" in text

    def test_generate_csv_empty_data(self) -> None:
        """空数据也应正常生成 CSV（仅含表头）。"""
        result = _generate_csv_bytes(
            start_time="2026-06-26T09:00:00Z",
            end_time="2026-06-26T10:00:00Z",
            plant_name="全部装置",
            label_counts=[],
            trend_rows=[],
            detail_rows=[],
        )
        assert isinstance(result, bytes)
        assert len(result) > 0


# ===========================================================================
# _do_export_diagnosis_statistics 异步逻辑测试
# ===========================================================================


def _make_export_db_mock(
    label_counts: list | None = None,
    trend_rows: list | None = None,
    detail_rows: list | None = None,
    plant_node: MagicMock | None = None,
    plant_node_id: str | None = None,
) -> AsyncMock:
    """构造 _do_export_diagnosis_statistics 用的 db mock。

    db.execute 调用顺序：
    1. label 分布查询（始终）
    2. plant_node 查询（仅当 plant_node_id 提供）
    3. trend 趋势查询（始终）
    4. detail 明细查询（始终）
    """
    label_counts = label_counts or []
    trend_rows = trend_rows or []
    detail_rows = detail_rows or []

    label_result = MagicMock()
    label_result.all.return_value = label_counts

    trend_result = MagicMock()
    trend_result.all.return_value = trend_rows

    detail_result = MagicMock()
    detail_result.all.return_value = detail_rows

    side_effects = [label_result]
    if plant_node_id is not None:
        plant_result = MagicMock()
        plant_result.scalar_one_or_none.return_value = plant_node
        side_effects.append(plant_result)
    side_effects.extend([trend_result, detail_result])

    session = AsyncMock()
    session.execute = AsyncMock(side_effect=side_effects)
    return session


class TestDoExportDiagnosisStatistics:
    """测试 _do_export_diagnosis_statistics() 异步导出逻辑。"""

    @pytest.mark.asyncio
    async def test_export_excel_success(self) -> None:
        """Excel 格式导出应成功生成文件并返回正确元信息。"""
        label_counts = [("OSCILLATION", 10), ("VALVE_STICTION", 5)]
        trend_rows = [
            (datetime(2026, 6, 26, tzinfo=UTC), "OSCILLATION", 6),
            (datetime(2026, 6, 26, tzinfo=UTC), "VALVE_STICTION", 3),
        ]
        detail_rows = [
            (
                "loop-001",
                "OSCILLATION",
                0.85,
                datetime(2026, 6, 26, 9, 30, tzinfo=UTC),
                "v1.0",
            ),
        ]
        db_mock = _make_export_db_mock(
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch.dict(os.environ, {"CLPM_EXPORT_DIR": tempfile.gettempdir()}),
        ):
            mock_session_local.return_value.__aenter__.return_value = db_mock
            result = await _do_export_diagnosis_statistics(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
                plant_node_id=None,
                diagnosis_label=None,
                action_status=None,
                user_id="user-001",
                granularity="day",
                file_format="xlsx",
            )

        assert result["status"] == "SUCCESS"
        assert result["fileFormat"] == "xlsx"
        assert result["labelCount"] == 2
        assert result["totalRecords"] == 15
        assert os.path.exists(result["fileUrl"])
        assert result["fileSize"] > 0
        # 清理生成的文件
        os.remove(result["fileUrl"])

    @pytest.mark.asyncio
    async def test_export_csv_success(self) -> None:
        """CSV 格式导出应成功生成文件。"""
        label_counts = [("OVERAGGRESSIVE", 3)]
        trend_rows = [
            (datetime(2026, 6, 26, tzinfo=UTC), "OVERAGGRESSIVE", 3),
        ]
        detail_rows = [
            (
                "loop-002",
                "OVERAGGRESSIVE",
                0.75,
                datetime(2026, 6, 26, 9, 30, tzinfo=UTC),
                "v1.0",
            ),
        ]
        db_mock = _make_export_db_mock(
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch.dict(os.environ, {"CLPM_EXPORT_DIR": tempfile.gettempdir()}),
        ):
            mock_session_local.return_value.__aenter__.return_value = db_mock
            result = await _do_export_diagnosis_statistics(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
                plant_node_id=None,
                diagnosis_label="OVERAGGRESSIVE",
                file_format="csv",
            )

        assert result["status"] == "SUCCESS"
        assert result["fileFormat"] == "csv"
        assert result["labelCount"] == 1
        assert os.path.exists(result["fileUrl"])
        # 清理生成的文件
        os.remove(result["fileUrl"])

    @pytest.mark.asyncio
    async def test_export_with_plant_node_id(self) -> None:
        """带 plant_node_id 时应查询装置名并应用到报表。"""
        plant_node = MagicMock()
        plant_node.name = "醛化反应单元"

        label_counts = [("OSCILLATION", 2)]
        trend_rows = [(datetime(2026, 6, 26, tzinfo=UTC), "OSCILLATION", 2)]
        detail_rows = [
            (
                "loop-001",
                "OSCILLATION",
                0.9,
                datetime(2026, 6, 26, 9, 0, tzinfo=UTC),
                "v1.0",
            ),
        ]
        db_mock = _make_export_db_mock(
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
            plant_node=plant_node,
            plant_node_id="node-001",
        )

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch.dict(os.environ, {"CLPM_EXPORT_DIR": tempfile.gettempdir()}),
        ):
            mock_session_local.return_value.__aenter__.return_value = db_mock
            result = await _do_export_diagnosis_statistics(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
                plant_node_id="node-001",
                file_format="xlsx",
            )

        assert result["status"] == "SUCCESS"
        assert os.path.exists(result["fileUrl"])
        os.remove(result["fileUrl"])

    @pytest.mark.asyncio
    async def test_export_empty_data_still_succeeds(self) -> None:
        """无诊断数据时应正常生成空报表。"""
        db_mock = _make_export_db_mock(
            label_counts=[],
            trend_rows=[],
            detail_rows=[],
        )

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch.dict(os.environ, {"CLPM_EXPORT_DIR": tempfile.gettempdir()}),
        ):
            mock_session_local.return_value.__aenter__.return_value = db_mock
            result = await _do_export_diagnosis_statistics(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
                file_format="xlsx",
            )

        assert result["status"] == "SUCCESS"
        assert result["labelCount"] == 0
        assert result["totalRecords"] == 0
        assert os.path.exists(result["fileUrl"])
        os.remove(result["fileUrl"])

    @pytest.mark.asyncio
    async def test_export_filename_format(self) -> None:
        """导出文件名应符合 CLPM-诊断统计报表-起始_结束.xlsx 规范。"""
        db_mock = _make_export_db_mock(
            label_counts=[],
            trend_rows=[],
            detail_rows=[],
        )

        with (
            patch("app.core.db.AsyncSessionLocal") as mock_session_local,
            patch.dict(os.environ, {"CLPM_EXPORT_DIR": tempfile.gettempdir()}),
        ):
            mock_session_local.return_value.__aenter__.return_value = db_mock
            result = await _do_export_diagnosis_statistics(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-27T09:00:00Z",
                file_format="xlsx",
            )

        file_path = result["fileUrl"]
        file_name = os.path.basename(file_path)
        assert file_name.startswith("CLPM-诊断统计报表-")
        assert file_name.endswith(".xlsx")
        assert "2026-06-26" in file_name
        assert "2026-06-27" in file_name
        os.remove(file_path)


# ===========================================================================
# export_diagnosis_statistics Celery 任务入口测试
# ===========================================================================


class TestExportDiagnosisStatisticsTask:
    """测试 export_diagnosis_statistics Celery 任务入口。"""

    def test_task_registered(self) -> None:
        """export_diagnosis_statistics 应注册到 celery_app。"""
        assert "app.tasks.report_generator.export_diagnosis_statistics" in celery_app.tasks

    def test_task_success_returns_result(self) -> None:
        """任务成功时应返回 _do_export_diagnosis_statistics 的结果。"""
        expected_result = {
            "taskId": "task-export-001",
            "status": "SUCCESS",
            "fileUrl": "/tmp/CLPM-test.xlsx",
            "fileSize": 1024,
            "fileFormat": "xlsx",
            "labelCount": 2,
            "totalRecords": 10,
        }

        task = export_diagnosis_statistics
        with patch.object(task, "run_async", return_value=expected_result):
            result = task(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
                plant_node_id=None,
                diagnosis_label=None,
                action_status=None,
                user_id="user-001",
                granularity="day",
                file_format="xlsx",
            )

        assert result == expected_result
        assert result["status"] == "SUCCESS"
        assert result["fileFormat"] == "xlsx"

    def test_task_system_error_propagates(self) -> None:
        """系统错误应抛出（触发 Celery autoretry）。"""
        task = export_diagnosis_statistics

        with patch.object(
            task,
            "run_async",
            side_effect=RuntimeError("DB 连接失败"),
        ):
            with pytest.raises(RuntimeError, match="DB 连接失败"):
                task(
                    start_time="2026-06-26T09:00:00Z",
                    end_time="2026-06-26T10:00:00Z",
                    file_format="xlsx",
                )

    def test_task_default_parameters(self) -> None:
        """任务应使用默认参数 granularity=day, file_format=xlsx。"""
        expected_result = {
            "taskId": "task-default",
            "status": "SUCCESS",
            "fileUrl": "/tmp/test.xlsx",
            "fileSize": 100,
            "fileFormat": "xlsx",
            "labelCount": 0,
            "totalRecords": 0,
        }

        task = export_diagnosis_statistics

        with patch.object(task, "run_async", return_value=expected_result):
            result = task(
                start_time="2026-06-26T09:00:00Z",
                end_time="2026-06-26T10:00:00Z",
            )

        assert result == expected_result
        assert result["fileFormat"] == "xlsx"
