#!/usr/bin/env python3
"""工厂节点 + 回路测试数据导入脚本。

从 Excel 文件读取工厂结构和回路清单，写入 PostgreSQL 数据库。

数据来源：
- 单元信息：docs/预研文档/单元信息_20260625092945.xlsx
- 回路清单：docs/预研文档/loopList.xlsx

工厂：诚志永清
工艺装置：MTO / 烯烃分离 / 丁二烯 / 2EH辛醇 / 公用工程

用法::

    cd backend && uv run python scripts/import_factory_data.py
    cd backend && uv run python scripts/import_factory_data.py --clean  # 清理后重新导入
    cd backend && uv run python scripts/import_factory_data.py --loop-count 100  # 指定回路数量
"""

from __future__ import annotations

import argparse
import asyncio
import random
import re
import uuid
from collections import defaultdict
from pathlib import Path

import openpyxl
from sqlalchemy import text

from app.core.db import AsyncSessionLocal

# ============================================================================
# 常量
# ============================================================================

PROJECT_ROOT = Path(__file__).parent.parent.parent
UNIT_INFO_XLSX = PROJECT_ROOT / "docs" / "预研文档" / "单元信息_20260625092945.xlsx"
LOOP_LIST_XLSX = PROJECT_ROOT / "docs" / "预研文档" / "loopList.xlsx"

FACTORY_NAME = "诚志永清"
FACTORY_ID = "00000000-0000-0000-0000-000000000001"

AREA_NAME_MAP = {
    "UT": "公用工程",
    "BD": "丁二烯装置",
    "MTO": "MTO装置",
    "OPU": "烯烃分离装置",
    "2EH": "2EH辛醇装置",
}

# 各装置的工艺单元定义：{单元编号: (单元名称, 设备位号数字前缀列表)}
AREA_UNIT_DEFS = {
    "MTO": [
        ("MTO-10", "反应再生单元", ["10"]),
        ("MTO-20", "急冷分离单元", ["20"]),
        ("MTO-30", "烯烃压缩单元", ["30"]),
    ],
    "OPU": [
        ("OPU-30", "预处理单元", ["30"]),
        ("OPU-40", "脱甲烷精馏单元", ["40"]),
        ("OPU-50", "乙烯丙烯精馏单元", ["50"]),
    ],
    "BD": [
        ("BD-11", "加氢反应单元", ["11"]),
        ("BD-12", "萃取精馏单元", ["12"]),
        ("BD-13", "丁二烯精制单元", ["13"]),
        ("BD-14", "溶剂回收单元", ["14"]),
        ("BD-15", "压缩制冷单元", ["15"]),
    ],
    "2EH": [
        ("2EH-10", "醛化反应单元", ["10"]),
        ("2EH-11", "缩合单元", ["11"]),
        ("2EH-30", "加氢精馏单元", ["30"]),
        ("2EH-31", "辛醇精制单元", ["31"]),
    ],
}

LOOP_TYPE_MAP = {
    "T": "TEMPERATURE",
    "P": "PRESSURE",
    "L": "LEVEL",
    "F": "FLOW",
    "A": "ANALYSIS",
    "S": "SPEED",
}

EQUIP_PATTERN = re.compile(r"[A-Z]-?(\d{2,5})[A-Z]?")


# ============================================================================
# Excel 解析
# ============================================================================


def parse_unit_info() -> list[dict]:
    wb = openpyxl.load_workbook(UNIT_INFO_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()

    units = []
    for row in rows:
        code, name, typ, area, craft, desc = row[:6]
        if not code:
            continue
        units.append(
            {
                "code": str(code).strip(),
                "name": str(name).strip() if name else str(code).strip(),
                "area": str(area).strip() if area else "OTHER",
                "description": str(desc).strip() if desc else None,
            }
        )
    return units


def parse_loop_list() -> list[dict]:
    wb = openpyxl.load_workbook(LOOP_LIST_XLSX, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()

    loops = []
    for row in rows:
        (
            loop_no,
            loop_name,
            sp_tag,
            pv_tag,
            op_tag,
            mode_tag,
            area,
            enabled,
            kp_tag,
            ti_tag,
            td_tag,
        ) = row[:11]
        if not loop_no:
            continue
        loops.append(
            {
                "tag_name": str(loop_no).strip(),
                "description": str(loop_name).strip() if loop_name else None,
                "area": str(area).strip() if area else "OTHER",
                "enabled": str(enabled).strip() == "是" if enabled else False,
                "sp_tag": str(sp_tag).strip() if sp_tag else None,
                "pv_tag": str(pv_tag).strip() if pv_tag else None,
                "op_tag": str(op_tag).strip() if op_tag else None,
                "mode_tag": str(mode_tag).strip() if mode_tag else None,
                "kp_tag": str(kp_tag).strip() if kp_tag else None,
                "ti_tag": str(ti_tag).strip() if ti_tag else None,
                "td_tag": str(td_tag).strip() if td_tag else None,
            }
        )
    return loops


def infer_loop_type(tag_name: str) -> str:
    """从位号推断回路类型。

    支持数字开头的位号：41FIC20021_PIDA → F → FLOW
    """
    match = re.match(r"^\d*([A-Z])", tag_name)
    if match:
        letter = match.group(1)
        return LOOP_TYPE_MAP.get(letter, "OTHER")
    return "OTHER"


def assign_loop_to_unit(loop: dict, area: str) -> str:
    """根据回路名称中的设备位号，分配到具体单元。

    返回单元编号（如 MTO-20），如果匹配不到则返回该装置的第一个单元。
    """
    unit_defs = AREA_UNIT_DEFS.get(area, [])
    if not unit_defs:
        return None

    desc = loop.get("description", "") or ""
    # 从描述中提取设备编号
    matches = EQUIP_PATTERN.findall(desc)
    if matches:
        equip_num = matches[0]
        prefix = equip_num[:2]
        for unit_code, _, prefixes in unit_defs:
            if prefix in prefixes:
                return unit_code

    # 尝试从位号中提取（如 41FIC20015 → 20）
    tag = loop.get("tag_name", "")
    num_match = re.search(r"(\d{2,5})", tag)
    if num_match:
        prefix = num_match.group(1)[:2]
        for unit_code, _, prefixes in unit_defs:
            if prefix in prefixes:
                return unit_code

    return unit_defs[0][0]


def build_unit_list() -> list[dict]:
    """构建完整的单元列表。

    UT（公用工程）从 Excel 读取，其他装置使用 AREA_UNIT_DEFS 定义。
    """
    excel_units = parse_unit_info()
    all_units = []

    for unit in excel_units:
        area = unit["area"]
        code = unit["code"]

        if area == "UT":
            all_units.append(unit)
        elif code.endswith("-GY"):
            continue
        else:
            all_units.append(unit)

    for area, unit_defs in AREA_UNIT_DEFS.items():
        for unit_code, unit_name, _ in unit_defs:
            all_units.append(
                {
                    "code": unit_code,
                    "name": unit_name,
                    "area": area,
                    "description": None,
                }
            )

    return all_units


# ============================================================================
# 数据库操作
# ============================================================================


async def clean_data(session) -> None:
    await session.execute(
        text("""
        DELETE FROM kpi_snapshot_hourly
        WHERE loop_id IN (SELECT id FROM loop_ledger WHERE created_by = 'import_script')
    """)
    )
    await session.execute(
        text("""
        DELETE FROM loop_tag_mapping
        WHERE loop_id IN (SELECT id FROM loop_ledger WHERE created_by = 'import_script')
    """)
    )
    await session.execute(
        text("""
        DELETE FROM tag_registry
        WHERE id IN (
            SELECT tr.id FROM tag_registry tr
            JOIN loop_tag_mapping ltm ON tr.id = ltm.tag_id
            JOIN loop_ledger ll ON ltm.loop_id = ll.id
            WHERE ll.created_by = 'import_script'
        )
    """)
    )
    await session.execute(
        text("""
        DELETE FROM loop_ledger WHERE created_by = 'import_script'
    """)
    )
    print("  ✓ 旧测试回路数据已清理")

    factory_names = [FACTORY_NAME, "某化工厂", "[E2E_TEST]化工厂", "加氢联合车间"]
    for fname in factory_names:
        result = await session.execute(
            text("""
            SELECT id FROM plant_node WHERE type = 'FACTORY' AND name = :fname
        """),
            {"fname": fname},
        )
        if not result.scalar():
            continue

        await session.execute(
            text("""
            DELETE FROM plant_node
            WHERE type = 'UNIT'
              AND parent_id IN (
                SELECT id FROM plant_node
                WHERE type = 'UNIT'
                  AND parent_id IN (
                    SELECT id FROM plant_node WHERE type = 'FACTORY' AND name = :fname
                  )
              )
        """),
            {"fname": fname},
        )

        await session.execute(
            text("""
            DELETE FROM plant_node
            WHERE type = 'UNIT'
              AND parent_id IN (SELECT id FROM plant_node WHERE type = 'FACTORY' AND name = :fname)
        """),
            {"fname": fname},
        )

        await session.execute(
            text("""
            DELETE FROM plant_node WHERE type = 'FACTORY' AND name = :fname
        """),
            {"fname": fname},
        )

    print("  ✓ 旧工厂节点已清理")


async def import_plant_nodes(session, units: list[dict]) -> dict[str, str]:
    await session.execute(
        text("""
        INSERT INTO plant_node (id, name, type, parent_id, is_kpi_enabled, created_at, updated_at)
        VALUES (:id, :name, 'FACTORY', NULL, TRUE, NOW(), NOW())
        ON CONFLICT (id) DO UPDATE SET
            name = EXCLUDED.name,
            type = EXCLUDED.type,
            is_kpi_enabled = EXCLUDED.is_kpi_enabled,
            updated_at = NOW()
    """),
        {"id": FACTORY_ID, "name": FACTORY_NAME},
    )

    areas = {}
    for unit in units:
        area_code = unit["area"]
        if area_code not in areas:
            areas[area_code] = {
                "id": str(uuid.uuid4()),
                "code": area_code,
                "name": AREA_NAME_MAP.get(area_code, area_code),
            }

    area_id_map = {}
    for area_code, area in areas.items():
        await session.execute(
            text("""
            INSERT INTO plant_node
                (id, name, type, parent_id, is_kpi_enabled, created_at, updated_at)
            VALUES (:id, :name, 'UNIT', :parent_id, TRUE, NOW(), NOW())
            ON CONFLICT (id) DO UPDATE SET
                name = EXCLUDED.name,
                parent_id = EXCLUDED.parent_id,
                is_kpi_enabled = EXCLUDED.is_kpi_enabled,
                updated_at = NOW()
        """),
            {
                "id": area["id"],
                "name": area["name"],
                "parent_id": FACTORY_ID,
            },
        )
        area_id_map[area_code] = area["id"]

    unit_id_map = {}
    for unit in units:
        unit_id = str(uuid.uuid4())
        area_id = area_id_map.get(unit["area"])
        await session.execute(
            text("""
            INSERT INTO plant_node
                (id, name, type, parent_id, is_kpi_enabled, created_at, updated_at)
            VALUES (:id, :name, 'UNIT', :parent_id, FALSE, NOW(), NOW())
            ON CONFLICT (id) DO UPDATE SET
                name = EXCLUDED.name,
                parent_id = EXCLUDED.parent_id,
                updated_at = NOW()
        """),
            {
                "id": unit_id,
                "name": unit["name"],
                "parent_id": area_id,
            },
        )
        unit_id_map[unit["code"]] = unit_id

    await session.commit()
    area_count = len(areas)
    unit_count = len(units)
    print(f"  ✓ 工厂节点导入完成：1 工厂 + {area_count} 区域 + {unit_count} 单元")
    return unit_id_map


async def import_loops(
    session,
    loops: list[dict],
    unit_id_map: dict[str, str],
    loop_count: int = 100,
) -> list[dict]:
    area_loops = defaultdict(list)
    for loop in loops:
        area_loops[loop["area"]].append(loop)

    selected_loops = []
    areas = sorted(area_loops.keys())
    per_area = loop_count // len(areas)
    remainder = loop_count % len(areas)

    for i, area in enumerate(areas):
        area_loop_list = [loop for loop in area_loops[area] if loop["enabled"]]
        if not area_loop_list:
            area_loop_list = area_loops[area]

        count = per_area + (1 if i < remainder else 0)
        count = min(count, len(area_loop_list))

        random.seed(42 + i)
        selected = random.sample(area_loop_list, count)
        selected_loops.extend(selected)

    imported = []
    for loop_data in selected_loops:
        area = loop_data["area"]

        unit_code = assign_loop_to_unit(loop_data, area)
        unit_id = unit_id_map.get(unit_code) if unit_code else None

        if not unit_id:
            ut_codes = [c for c in unit_id_map if c.startswith("HU") or c == "Other"]
            if ut_codes:
                unit_id = unit_id_map[ut_codes[0]]

        loop_type = infer_loop_type(loop_data["tag_name"])

        loop_id = str(uuid.uuid4())
        result = await session.execute(
            text("""
            INSERT INTO loop_ledger (
                id, tag_name, description, unit_id, score_weight,
                is_active, status, loop_type, level, created_at, updated_at, created_by
            ) VALUES (
                :id, :tag_name, :desc, :unit_id, 1.0,
                :is_active, 'READY', :loop_type, 3, NOW(), NOW(), 'import_script'
            )
            ON CONFLICT (tag_name) DO UPDATE SET
                description = EXCLUDED.description,
                unit_id = EXCLUDED.unit_id,
                loop_type = EXCLUDED.loop_type,
                updated_at = NOW()
            RETURNING id
        """),
            {
                "id": loop_id,
                "tag_name": loop_data["tag_name"],
                "desc": loop_data["description"],
                "unit_id": unit_id,
                "is_active": loop_data["enabled"],
                "loop_type": loop_type,
            },
        )
        actual_loop_id = str(result.scalar())

        tag_defs = [
            ("PV", loop_data["pv_tag"]),
            ("SP", loop_data["sp_tag"]),
            ("OP", loop_data["op_tag"]),
            ("MODE", loop_data["mode_tag"]),
            ("PID_P", loop_data["kp_tag"]),
            ("PID_I", loop_data["ti_tag"]),
            ("PID_D", loop_data["td_tag"]),
        ]

        for role, tag_name in tag_defs:
            if not tag_name:
                continue

            tag_id = str(uuid.uuid4())
            is_required = role in ("PV", "SP", "OP", "MODE")

            result = await session.execute(
                text("""
                INSERT INTO tag_registry (
                    id, tag_name, tag_description, tag_type,
                    current_value, quality, last_sync_at, is_linked, measure_type
                ) VALUES (
                    :id, :tag_name, :desc, :type,
                    0.0, 'GOOD', NOW(), TRUE, :measure_type
                )
                ON CONFLICT (tag_name) DO UPDATE SET
                    tag_description = EXCLUDED.tag_description,
                    is_linked = TRUE,
                    measure_type = EXCLUDED.measure_type
                RETURNING id
            """),
                {
                    "id": tag_id,
                    "tag_name": tag_name,
                    "desc": f"{loop_data['description'] or loop_data['tag_name']} - {role}",
                    "type": role,
                    "measure_type": loop_type,
                },
            )
            actual_tag_id = str(result.scalar())

            mapping_id = str(uuid.uuid4())
            await session.execute(
                text("""
                INSERT INTO loop_tag_mapping
                    (id, loop_id, tag_id, tag_role, is_required, created_at)
                VALUES (:id, :loop_id, :tag_id, :role, :required, NOW())
                ON CONFLICT (loop_id, tag_role) DO UPDATE SET
                    tag_id = EXCLUDED.tag_id
            """),
                {
                    "id": mapping_id,
                    "loop_id": actual_loop_id,
                    "tag_id": actual_tag_id,
                    "role": role,
                    "required": is_required,
                },
            )

        imported.append(
            {
                "id": actual_loop_id,
                "tag_name": loop_data["tag_name"],
                "area": area,
                "unit_code": unit_code,
                "unit_id": unit_id,
                "loop_type": loop_type,
            }
        )

    await session.commit()
    print(f"  ✓ 测试回路导入完成：{len(imported)} 个回路")

    area_count = defaultdict(int)
    unit_count = defaultdict(int)
    for item in imported:
        area_count[item["area"]] += 1
        unit_count[item["unit_code"]] += 1

    print("    按区域分布:")
    for area in sorted(area_count.keys()):
        print(f"      - {AREA_NAME_MAP.get(area, area)}: {area_count[area]} 个")

    print("    按单元分布:")
    for unit in sorted(unit_count.keys(), key=lambda x: (x is None, str(x))):
        label = unit if unit else "未分配"
        print(f"      - {label}: {unit_count[unit]} 个")

    return imported


# ============================================================================
# 主函数
# ============================================================================


async def main(clean: bool = False, loop_count: int = 100) -> None:
    if not UNIT_INFO_XLSX.exists():
        print(f"错误: 单元信息文件不存在: {UNIT_INFO_XLSX}")
        return
    if not LOOP_LIST_XLSX.exists():
        print(f"错误: 回路清单文件不存在: {LOOP_LIST_XLSX}")
        return

    print(f"\n{'=' * 60}")
    print("1. 解析 Excel 数据")
    print(f"{'=' * 60}")
    units = build_unit_list()
    loops = parse_loop_list()
    print(f"  ✓ 单元信息: {len(units)} 个")
    print(f"  ✓ 回路清单: {len(loops)} 条")

    ut_units = [u for u in units if u["area"] == "UT"]
    mto_units = [u for u in units if u["area"] == "MTO"]
    opu_units = [u for u in units if u["area"] == "OPU"]
    bd_units = [u for u in units if u["area"] == "BD"]
    eh_units = [u for u in units if u["area"] == "2EH"]
    print(f"    - 公用工程: {len(ut_units)} 个")
    print(f"    - MTO装置: {len(mto_units)} 个")
    print(f"    - 烯烃分离: {len(opu_units)} 个")
    print(f"    - 丁二烯: {len(bd_units)} 个")
    print(f"    - 2EH辛醇: {len(eh_units)} 个")

    async with AsyncSessionLocal() as session:
        if clean:
            print(f"\n{'=' * 60}")
            print("2. 清理旧数据")
            print(f"{'=' * 60}")
            await clean_data(session)
            await session.commit()

        print(f"\n{'=' * 60}")
        print("3. 导入工厂节点")
        print(f"{'=' * 60}")
        unit_id_map = await import_plant_nodes(session, units)

        print(f"\n{'=' * 60}")
        print(f"4. 导入测试回路（{loop_count} 个）")
        print(f"{'=' * 60}")
        imported_loops = await import_loops(session, loops, unit_id_map, loop_count)

    print(f"\n{'=' * 60}")
    print("导入完成！")
    print(f"{'=' * 60}")
    print(f"工厂: {FACTORY_NAME}")
    print("区域: 5 个")
    print(f"单元: {len(units)} 个")
    print(f"测试回路: {len(imported_loops)} 个")
    print()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="工厂节点 + 回路测试数据导入")
    parser.add_argument("--clean", action="store_true", help="清理旧数据后重新导入")
    parser.add_argument("--loop-count", type=int, default=100, help="测试回路数量（默认 100）")
    args = parser.parse_args()
    asyncio.run(main(clean=args.clean, loop_count=args.loop_count))
