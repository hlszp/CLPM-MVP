"""Loop ledger service — CRUD + status derivation (IDS v3.2 §2.2.7~2.2.11).

状态推导规则：
- INACTIVE: is_active = false
- PARTIAL: is_active = true 但 PV/SP/OP/MODE 4 个必填 Tag 缺失任一
- READY: is_active = true 且 4 个必填 Tag 全部关联（PID_P/PID_I/PID_D 可选）
"""

from __future__ import annotations

import io
import json
import logging
import random
from datetime import UTC, datetime
from typing import Any
from uuid import uuid4

import openpyxl
from sqlalchemy import delete, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BizError
from app.core.redis import redis_client
from app.core.tdengine import _TAG_NAME_PATTERN
from app.models.audit import SysAuditLog
from app.models.dcs_model import DcsModel
from app.models.loop import COMPLEX_ROLE_MAIN, LoopLedger, LoopTagMapping
from app.models.plant_node import PlantNode
from app.models.tag import TagRegistry
from app.services.cache.invalidation import CacheInvalidator
from app.services.dict_item import (
    DICT_LOOP_TYPE,
    dict_items_hint,
    get_dict_items,
    normalize_by_dict,
)

logger = logging.getLogger(__name__)

# Phase 10 性能优化：回路类型/控制方式统计 Redis 短 TTL 缓存
# 60s TTL + ±10s 抖动，避免惊群；写入失败降级为直接查询（与 dashboard 缓存模式一致）
LOOP_STATS_CACHE_KEY_TEMPLATE = "loop:stats:type:{plant_node_id}"
LOOP_STATS_CACHE_TTL = 60

# 必填 Tag 角色
REQUIRED_ROLES = ("PV", "SP", "OP", "MODE")
# 全部 7 个 Tag 角色
ALL_ROLES = ("PV", "SP", "OP", "MODE", "PID_P", "PID_I", "PID_D")
# 角色映射到响应字段名
ROLE_TO_FIELD = {
    "PV": "pv",
    "SP": "sp",
    "OP": "op",
    "MODE": "mode",
    "PID_P": "pid_p",
    "PID_I": "pid_i",
    "PID_D": "pid_d",
}

# P2：tag 重关联（改关联 / Excel 覆盖式删建映射）会使 TDengine subtable 名
# 派生自新 tag 名，旧 subtable 中的历史数据立即不可达——检测到变更时在响应中
# 返回此 warning，并失效该回路的 L1 DataBlock 缓存与 subtable 解析缓存
TAG_REASSIGN_WARNING = "tag 变更将导致历史数据在新 subtable 下重新开始，旧数据不可达"


def _clear_subtable_cache(loop_id: str) -> None:
    """清除 tdengine_provider 模块级 subtable 解析缓存中该回路的条目。

    tdengine_provider._subtable_cache 暂无公开清除函数，此处直接 pop 模块级 dict
    （GIL 下 dict.pop 原子安全；该缓存为单进程 300s TTL 缓存，见 tdengine_provider
    文件头设计说明）。
    """
    from app.services.data_source import tdengine_provider

    tdengine_provider._subtable_cache.pop(loop_id, None)


async def get_loop_role_tag_names(db: AsyncSession, loop_id: str) -> dict[str, str]:
    """查询回路当前各角色关联的 tag 名（role → tag_name，未关联的角色不出现）。"""
    result = await db.execute(
        select(LoopTagMapping.tag_role, TagRegistry.tag_name)
        .join(TagRegistry, LoopTagMapping.tag_id == TagRegistry.id)
        .where(LoopTagMapping.loop_id == loop_id)
    )
    return dict(result.all())


def detect_tag_reassignment(
    old_role_tags: dict[str, str], new_role_tags: dict[str, str]
) -> list[str]:
    """对比变更前后各角色 tag 名，返回发生变化（含新增/移除）的角色列表。

    角色顺序固定为 ALL_ROLES，保证 warning 文案与测试断言稳定。
    """
    return [
        role
        for role in ALL_ROLES
        if (role in old_role_tags or role in new_role_tags)
        and old_role_tags.get(role) != new_role_tags.get(role)
    ]


async def notify_tag_reassignment(
    loop_id: str, loop_tag_name: str, changed_roles: list[str]
) -> str:
    """tag 重关联后置处理：warning + 失效 L1 缓存 + 清除 subtable 解析缓存。

    缓存失效失败不阻断主流程（缓存本身可安全降级为 miss 后重建）。

    Returns:
        面向调用方的 warning 文案（追加到响应 warnings 中）
    """
    warning = f"回路 {loop_tag_name}（角色 {'/'.join(changed_roles)}）：{TAG_REASSIGN_WARNING}"
    logger.warning(
        "tag 重关联: loop_id=%s, changed_roles=%s — %s",
        loop_id,
        changed_roles,
        TAG_REASSIGN_WARNING,
    )
    _clear_subtable_cache(loop_id)
    try:
        await CacheInvalidator(redis_client).invalidate_loop(loop_id)
    except Exception:  # noqa: BLE001
        logger.warning("L1 缓存失效失败（不影响关联变更）: loop_id=%s", loop_id, exc_info=True)
    return warning


# v6.1：回路类型 / 控制类型 中英文双向映射（Excel 导入导出用）
# 导出时英→中（用户友好），导入时中→英（容错识别）
LOOP_TYPE_TO_CN: dict[str, str] = {
    "TEMPERATURE": "温度",
    "PRESSURE": "压力",
    "LEVEL": "液位",
    "FLOW": "流量",
    "ANALYSIS": "分析",
    "SPEED": "速度",
    "OTHER": "其他",
}
LOOP_TYPE_FROM_CN: dict[str, str] = {v: k for k, v in LOOP_TYPE_TO_CN.items()}

CONTROL_TYPE_TO_CN: dict[str, str] = {
    "STABLE": "稳定型",
    "SLOW": "慢速型",
    "FAST": "快速型",
    "LOGIC": "逻辑型",
}
CONTROL_TYPE_FROM_CN: dict[str, str] = {v: k for k, v in CONTROL_TYPE_TO_CN.items()}


async def _write_audit(
    db: AsyncSession,
    operator: str,
    operation_type: str,
    target_type: str,
    target_id: str,
    before_value: str | None = None,
    after_value: str | None = None,
) -> None:
    """写入审计日志。"""
    log = SysAuditLog(
        id=str(uuid4()),
        operator=operator,
        operation_type=operation_type,
        target_type=target_type,
        target_id=target_id,
        before_value=before_value,
        after_value=after_value,
        operated_at=datetime.now(UTC).replace(tzinfo=None),
    )
    db.add(log)


async def _get_loop_tag_mappings(db: AsyncSession, loop_id: str) -> dict[str, LoopTagMapping]:
    """获取回路的所有 Tag 关联（按角色索引）。"""
    result = await db.execute(select(LoopTagMapping).where(LoopTagMapping.loop_id == loop_id))
    return {m.tag_role: m for m in result.scalars().all()}


async def derive_loop_status(
    db: AsyncSession,
    loop: LoopLedger,
    mappings: dict[str, LoopTagMapping] | None = None,
) -> str:
    """根据 is_active 和 Tag 关联状态推导回路 status。

    Args:
        db: 异步数据库会话
        loop: 回路对象
        mappings: 已查询的 Tag 关联（可选，避免重复查询）

    Returns:
        新状态：READY / PARTIAL / INACTIVE
    """
    # INACTIVE: 未激活
    if not loop.is_active:
        return "INACTIVE"

    # 查询 Tag 关联
    if mappings is None:
        mappings = await _get_loop_tag_mappings(db, str(loop.id))

    # 检查 4 个必填 Tag 是否全部关联
    for role in REQUIRED_ROLES:
        if role not in mappings:
            return "PARTIAL"
    return "READY"


async def _get_unit_name(db: AsyncSession, unit_id: str | None) -> str | None:
    """获取单元名称。"""
    if not unit_id:
        return None
    result = await db.execute(select(PlantNode).where(PlantNode.id == unit_id))
    node = result.scalar_one_or_none()
    return node.name if node else None


async def _get_descendant_node_ids(db: AsyncSession, parent_id: str) -> list[str]:
    """递归获取所有子孙节点 ID。

    Phase 10 性能优化：原 N 次 select 递归收敛为 1 次 ``WITH RECURSIVE`` CTE
    （``plant_node_tree.collect_descendant_node_ids``）。保留薄包装以维持公共 API，
    ``performance.py`` 通过此名间接复用。
    """
    from app.services.plant_node_tree import collect_descendant_node_ids

    return await collect_descendant_node_ids(db, parent_id)


def _loop_stats_cache_key(plant_node_id: str | None) -> str:
    """构建 stats 缓存 key。"""
    return LOOP_STATS_CACHE_KEY_TEMPLATE.format(plant_node_id=plant_node_id or "all")


async def _read_loop_stats_cache(cache_key: str) -> dict | None:
    """读取 stats 缓存，失败时返回 None（降级为直接查询）。"""
    try:
        cached = await redis_client.get(cache_key)
        if cached:
            return json.loads(cached)
    except Exception as exc:  # noqa: BLE001
        logger.warning("读取回路类型统计缓存失败，降级为直接查询: %s", exc)
    return None


async def _write_loop_stats_cache(cache_key: str, data: dict) -> None:
    """写入 stats 缓存，失败时不报错（降级模式）。

    使用 TTL 抖动（±10s）避免大量 key 同时过期导致惊群效应。
    """
    try:
        ttl = LOOP_STATS_CACHE_TTL + random.randint(-10, 10)
        await redis_client.setex(cache_key, ttl, json.dumps(data, default=str))
    except Exception as exc:  # noqa: BLE001
        logger.warning("写入回路类型统计缓存失败: %s", exc)


async def get_loop_type_stats(db: AsyncSession, plant_node_id: str | None = None) -> dict:
    """按回路类型统计数量（支持递归子节点）。

    Phase 10 性能优化：
    - 子孙节点遍历用 1 次 CTE 替代 N 次递归 select
    - loop_type 计数用 CTE 直接 JOIN loop_ledger + GROUP BY 一条 SQL 完成，
      不再"先递归拿 ID 再 IN 查询"两步
    - 整体结果加 60s Redis 短 TTL 缓存，减少重复计算

    Args:
        plant_node_id: 装置/单元 ID，为 None 时统计全部回路

    Returns:
        各回路类型的统计数量字典 + 控制方式统计
    """
    cache_key = _loop_stats_cache_key(plant_node_id)
    cached = await _read_loop_stats_cache(cache_key)
    if cached is not None:
        return cached

    if plant_node_id:
        # CTE 一次返回子孙节点 + 自身 → JOIN loop_ledger + GROUP BY loop_type
        # 一条 SQL 完成"递归子节点 + 按类型聚合"，省掉 IN 二次查询
        stmt = text(
            """
            WITH RECURSIVE node_tree AS (
                SELECT id FROM plant_node WHERE id = :plant_node_id
                UNION ALL
                SELECT child.id
                FROM plant_node child
                JOIN node_tree nt ON child.parent_id = nt.id
            )
            SELECT l.loop_type, COUNT(*) AS cnt
            FROM loop_ledger l
            JOIN node_tree nt ON l.unit_id = nt.id
            WHERE l.is_active = TRUE
            GROUP BY l.loop_type
            """
        )
        result = await db.execute(stmt, {"plant_node_id": plant_node_id})
        rows = result.all()
    else:
        # 无 plant_node_id：直接全表 GROUP BY，无需 CTE
        stmt = (
            select(LoopLedger.loop_type, func.count())
            .where(LoopLedger.is_active.is_(True))
            .group_by(LoopLedger.loop_type)
        )
        result = await db.execute(stmt)
        rows = result.all()

    type_stats: dict[str, int] = {
        "TEMPERATURE": 0,
        "PRESSURE": 0,
        "LEVEL": 0,
        "FLOW": 0,
        "ANALYSIS": 0,
        "SPEED": 0,
        "OTHER": 0,
    }
    for row in rows:
        loop_type = row[0]
        count = row[1]
        key = loop_type or "OTHER"
        if key in type_stats:
            type_stats[key] = int(count)
        else:
            type_stats["OTHER"] += int(count)

    # 控制方式统计依赖 Redis 实时 MODE 值，无法纯 SQL 聚合，仍走批量内存统计
    # 但只取本范围回路（用 IN 一次拉回，避免逐节点查询）
    conditions: list = []
    if plant_node_id:
        all_node_ids = await _get_descendant_node_ids(db, plant_node_id)
        all_node_ids.append(plant_node_id)
        conditions.append(LoopLedger.unit_id.in_(all_node_ids))

    mode_stats = await _get_control_mode_stats(db, conditions)

    payload = {
        "loopTypeStats": type_stats,
        "controlModeStats": mode_stats,
    }
    await _write_loop_stats_cache(cache_key, payload)
    return payload


async def _get_control_mode_stats(
    db: AsyncSession,
    conditions: list,
) -> dict[str, int]:
    """从全量回路中按 MODE 数值统计（0=手动,1=自动,2=串级,3=远程,4=先控）。"""
    import logging

    from app.models.loop import LoopTagMapping
    from app.models.tag import TagRegistry
    from app.services.monitor import get_subscriber

    logger = logging.getLogger(__name__)

    stats: dict[str, int] = {"0": 0, "1": 0, "2": 0, "3": 0, "4": 0}

    all_loops_stmt = select(LoopLedger).where(LoopLedger.is_active.is_(True), *conditions)
    all_result = await db.execute(all_loops_stmt)
    all_loops = all_result.scalars().all()

    if not all_loops:
        return stats

    loop_ids = [str(lp.id) for lp in all_loops]

    mappings_result = await db.execute(
        select(LoopTagMapping).where(
            LoopTagMapping.loop_id.in_(loop_ids),
            LoopTagMapping.tag_role == "MODE",
        )
    )
    mode_mappings: dict[str, LoopTagMapping] = {}
    for m in mappings_result.scalars().all():
        mode_mappings[str(m.loop_id)] = m

    tag_ids = [str(m.tag_id) for m in mode_mappings.values()]
    tags_result = await db.execute(select(TagRegistry).where(TagRegistry.id.in_(tag_ids)))
    tags_map: dict[str, TagRegistry] = {}
    for t in tags_result.scalars().all():
        tags_map[str(t.id)] = t

    tag_names = [t.tag_name for t in tags_map.values() if t.tag_name]

    redis_cache: dict[str, dict] = {}
    try:
        subscriber = get_subscriber()
        if subscriber and tag_names:
            cached_list = await subscriber.get_cached_values(tag_names)
            for item in cached_list:
                tc = item.get("tagCode")
                if tc:
                    redis_cache[tc] = item
    except Exception as exc:  # noqa: BLE001
        logger.warning("从 Redis 读取实时值失败，回退到数据库值: %s", exc)

    for loop in all_loops:
        loop_id = str(loop.id)
        mode_val: float | None = None

        mapping = mode_mappings.get(loop_id)
        if mapping:
            tag = tags_map.get(str(mapping.tag_id))
            if tag and tag.tag_name:
                cached = redis_cache.get(tag.tag_name)
                if cached and "value" in cached:
                    try:
                        mode_val = float(cached["value"])
                    except (ValueError, TypeError):
                        pass
                elif tag.current_value is not None:
                    try:
                        mode_val = float(tag.current_value)
                    except (ValueError, TypeError):
                        pass

        # 映射到 0-4，超出范围归为 Unknown（不计入）
        key = (
            str(int(mode_val))
            if mode_val is not None and mode_val in (0, 1, 2, 3, 4)
            else "unknown"
        )
        if key in stats:
            stats[key] += 1

    return stats


def _build_tag_mapping_status(mappings: dict[str, LoopTagMapping]) -> dict:
    """构建 7 Tag 关联状态摘要。"""
    return {
        "pv": "PV" in mappings,
        "sp": "SP" in mappings,
        "op": "OP" in mappings,
        "mode": "MODE" in mappings,
        "pid_p": "PID_P" in mappings,
        "pid_i": "PID_I" in mappings,
        "pid_d": "PID_D" in mappings,
    }


async def list_loops(
    db: AsyncSession,
    plant_node_id: str | None = None,
    control_mode: str | None = None,
    is_active: bool | None = None,
    status: str | None = None,
    keyword: str | None = None,
    loop_type: str | None = None,
    control_type: str | None = None,
    importance_level: int | None = None,
    monitor_status: bool | None = None,
    include_in_evaluation: bool | None = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    """分页查询回路列表。

    Args:
        importance_level: 按回路重要等级筛选（1/2/3）
        control_type: 按控制类型筛选（STABLE/SLOW/FAST/LOGIC）
        monitor_status: 按监控状态筛选（True=is_active=True，False=is_active=False）
        include_in_evaluation: 按参评状态筛选（True=参评/False=不参评）

    Raises:
        ValueError: is_active 与 monitor_status 同时传入但值不一致（语义冲突）
    """
    # P3 #42: is_active 与 monitor_status 都映射到 LoopLedger.is_active 字段，
    # 同时传入不同值会生成 is_active=X AND is_active=Y → 永远返回空结果。
    # 校验：两个都传时值必须一致；统一为单一条件避免重复 AND。
    if is_active is not None and monitor_status is not None:
        if is_active != monitor_status:
            raise ValueError(
                "isActive 与 monitorStatus 语义相同（均映射到 is_active 字段），"
                "同时传入时值必须一致"
            )
        # 值一致时只保留一个，避免重复 AND
        monitor_status = None

    conditions = []
    if plant_node_id:
        # 递归获取所有子孙节点 ID，包含自身
        all_node_ids = await _get_descendant_node_ids(db, plant_node_id)
        all_node_ids.append(plant_node_id)
        conditions.append(LoopLedger.unit_id.in_(all_node_ids))
    if is_active is not None:
        conditions.append(LoopLedger.is_active.is_(is_active))
    if status:
        conditions.append(func.upper(LoopLedger.status) == status.upper())
    if loop_type:
        conditions.append(func.upper(LoopLedger.loop_type) == loop_type.upper())
    if control_type:
        # P2 #24: 控制类型筛选（STABLE/SLOW/FAST/LOGIC）
        conditions.append(func.upper(LoopLedger.control_type) == control_type.upper())
    if importance_level is not None:
        conditions.append(LoopLedger.importance_level == importance_level)
    if monitor_status is not None:
        # monitor_status=True → is_active=True（在监控中）
        # monitor_status=False → is_active=False（已停用监控）
        conditions.append(LoopLedger.is_active.is_(monitor_status))
    if include_in_evaluation is not None:
        conditions.append(LoopLedger.include_in_evaluation == include_in_evaluation)
    if keyword:
        kw = f"%{keyword}%"
        conditions.append(
            or_(
                LoopLedger.tag_name.ilike(kw),
                LoopLedger.description.ilike(kw),
            )
        )

    # controlMode 过滤下沉到 SQL 层（EXISTS 子查询），避免后置过滤导致 total 与分页错乱
    if control_mode:
        mode_values = _control_mode_to_values(control_mode)
        if not mode_values:
            # 无法识别的控制模式标签，直接返回空结果
            return {"items": [], "total": 0, "page": page, "pageSize": page_size}
        mode_exists = (
            select(LoopTagMapping.tag_id)
            .join(TagRegistry, LoopTagMapping.tag_id == TagRegistry.id)
            .where(LoopTagMapping.loop_id == LoopLedger.id)
            .where(LoopTagMapping.tag_role == "MODE")
            .where(TagRegistry.current_value.in_(mode_values))
            .exists()
        )
        conditions.append(mode_exists)

    count_stmt = select(func.count()).select_from(LoopLedger)
    for cond in conditions:
        count_stmt = count_stmt.where(cond)
    total_result = await db.execute(count_stmt)
    total = total_result.scalar() or 0

    # 双键排序：created_at desc 为主序，tag_name asc 为次序保证稳定排序
    # （批量导入的回路 created_at 可能相同，PostgreSQL 对相同键值不保证返回顺序，
    # 加次级排序避免编辑后回路位置漂移）
    stmt = select(LoopLedger).order_by(LoopLedger.created_at.desc(), LoopLedger.tag_name.asc())
    for cond in conditions:
        stmt = stmt.where(cond)
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    loops = result.scalars().all()

    # 批量查询 unit_name
    unit_ids = [str(loop.unit_id) for loop in loops if loop.unit_id]
    unit_map: dict[str, str] = {}
    if unit_ids:
        unit_result = await db.execute(select(PlantNode).where(PlantNode.id.in_(unit_ids)))
        for node in unit_result.scalars().all():
            unit_map[str(node.id)] = node.name

    # 批量查询 Tag 关联状态
    loop_ids = [str(loop.id) for loop in loops]
    mappings_map: dict[str, dict[str, LoopTagMapping]] = {}
    if loop_ids:
        m_result = await db.execute(
            select(LoopTagMapping).where(LoopTagMapping.loop_id.in_(loop_ids))
        )
        for m in m_result.scalars().all():
            mappings_map.setdefault(str(m.loop_id), {})[m.tag_role] = m

    # 批量查询 controlMode（优先从 Redis 缓存读取，回退到数据库 current_value）
    mode_map: dict[str, str] = {}
    mode_tag_map: dict[str, TagRegistry] = {}  # loop_id -> MODE Tag 对象
    if loop_ids:
        mode_result = await db.execute(
            select(LoopTagMapping, TagRegistry)
            .join(TagRegistry, LoopTagMapping.tag_id == TagRegistry.id)
            .where(LoopTagMapping.loop_id.in_(loop_ids))
            .where(LoopTagMapping.tag_role == "MODE")
        )
        for mapping, tag in mode_result:
            mode_tag_map[str(mapping.loop_id)] = tag

    # 批量从 Redis 读取实时值
    redis_cache: dict[str, dict] = {}
    if mode_tag_map:
        try:
            from app.services.data_source.realtime_subscriber import get_subscriber

            subscriber = get_subscriber()
            tag_names = [tag.tag_name for tag in mode_tag_map.values() if tag.tag_name]
            if tag_names:
                cached_list = await subscriber.get_cached_values(tag_names)
                for item in cached_list:
                    tc = item.get("tagCode")
                    if tc:
                        redis_cache[tc] = item
        except Exception as exc:  # noqa: BLE001
            logger.warning("从 Redis 读取实时值失败，回退到数据库值: %s", exc)

    # 构建 mode_map（优先使用 Redis 缓存）
    # WS-C 6-3：按回路 dcs_model_id 构建 MODE 解析映射（dcs_mode_mapping 回退链）
    loop_model_map: dict[str, str | None] = {
        str(loop.id): (str(loop.dcs_model_id) if loop.dcs_model_id else None) for loop in loops
    }
    raw_to_standard_maps = await _build_raw_to_standard_maps(db, set(loop_model_map.values()))
    for loop_id, tag in mode_tag_map.items():
        raw_to_standard = raw_to_standard_maps.get(loop_model_map.get(loop_id)) or (
            raw_to_standard_maps.get(None) or {}
        )
        cached = redis_cache.get(tag.tag_name)
        if cached and "value" in cached:
            try:
                mode_val = float(cached["value"])
                mode_map[loop_id] = _mode_value_to_label(mode_val, raw_to_standard)
            except (ValueError, TypeError):
                mode_map[loop_id] = _mode_value_to_label(tag.current_value, raw_to_standard)
        else:
            mode_map[loop_id] = _mode_value_to_label(tag.current_value, raw_to_standard)

    # v6.1 批量查询 PV/OP Tag 量程与单位（设计文档 §4.1）
    # 数据来源：tag_registry.range_min/range_max/unit，通过 loop_tag_mapping JOIN
    range_map: dict[str, dict[str, object]] = {}
    if loop_ids:
        range_result = await db.execute(
            select(LoopTagMapping, TagRegistry)
            .join(TagRegistry, LoopTagMapping.tag_id == TagRegistry.id)
            .where(LoopTagMapping.loop_id.in_(loop_ids))
            .where(LoopTagMapping.tag_role.in_(["PV", "OP"]))
        )
        for mapping, tag in range_result:
            loop_idx = str(mapping.loop_id)
            if loop_idx not in range_map:
                range_map[loop_idx] = {}
            role_key = mapping.tag_role.lower()  # "pv" / "op"
            range_map[loop_idx][f"{role_key}_range"] = {
                "min": float(tag.range_min) if tag.range_min is not None else None,
                "max": float(tag.range_max) if tag.range_max is not None else None,
            }
            range_map[loop_idx][f"{role_key}_unit"] = tag.unit

    # P2 IA优化：批量查询每回路最新 fitness
    fitness_map: dict[str, Any] = {}
    if loop_ids:
        try:
            from app.services.loop_fitness import get_latest_fitness_per_loop

            fitness_map = await get_latest_fitness_per_loop(db, loop_ids)
        except Exception as exc:  # noqa: BLE001
            logger.warning("list_loops fitness 批量查询失败，已忽略: %s", exc)

    items = []
    for loop in loops:
        mappings = mappings_map.get(str(loop.id), {})
        loop_range = range_map.get(str(loop.id), {})
        fit = fitness_map.get(str(loop.id))
        if fit is not None:
            fitness_level = getattr(fit, "level", None)
            fitness_tags = getattr(fit, "tags", None) or getattr(fit, "human_readable_tags", None)
        else:
            fitness_level = None
            fitness_tags = None
        items.append(
            {
                "loopId": str(loop.id),
                "tagName": loop.tag_name,
                "description": loop.description,
                "unitId": str(loop.unit_id) if loop.unit_id else None,
                "unitName": unit_map.get(str(loop.unit_id)) if loop.unit_id else None,
                "controlMode": mode_map.get(str(loop.id)),
                "isActive": bool(loop.is_active),
                "status": loop.status,
                "loopType": loop.loop_type,
                "controlType": loop.control_type,
                "importanceLevel": loop.importance_level,
                "includeInEvaluation": loop.include_in_evaluation,
                "score": float(loop.score_weight) if loop.score_weight else None,
                "lastScoreAt": (
                    loop.last_aas_sync_at.isoformat() if loop.last_aas_sync_at else None
                ),
                "tagMappingStatus": _build_tag_mapping_status(mappings),
                # v6.1 新增：量程与限位
                "pvRange": loop_range.get("pv_range"),
                "pvUnit": loop_range.get("pv_unit"),
                "opRange": loop_range.get("op_range"),
                "opUnit": loop_range.get("op_unit"),
                "opOutputLowerLimit": (
                    float(loop.op_output_lower_limit)
                    if loop.op_output_lower_limit is not None
                    else None
                ),
                "opOutputUpperLimit": (
                    float(loop.op_output_upper_limit)
                    if loop.op_output_upper_limit is not None
                    else None
                ),
                "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
                "idealSettlingTime": (
                    float(loop.ideal_settling_time)
                    if loop.ideal_settling_time is not None
                    else None
                ),
                "complexLoopGroupId": (
                    str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
                ),
                "complexRole": loop.complex_role,
                "fitnessLevel": fitness_level,
                "fitnessTags": fitness_tags,
            }
        )

    return {
        "items": items,
        "total": total,
        "page": page,
        "pageSize": page_size,
    }


# 标准 MODE 值 → 前端 ControlMode 标签
# 0-3 与原硬编码映射一致（与 monitor.py 默认映射对齐）；4=APC（先控）归并为 Auto（非手动）
_STANDARD_MODE_TO_FRONTEND: dict[int, str] = {
    0: "Manual",
    1: "Auto",
    2: "Cascade",
    3: "Cascade",
    4: "Auto",
}


async def _build_raw_to_standard_maps(
    db: AsyncSession,
    dcs_model_ids: set[str | None],
) -> dict[str | None, dict[int, int]]:
    """按 dcs_model_id 批量构建 raw MODE → standard MODE 映射（回退链）。

    每个型号：dcs_mode_mapping 型号映射优先，本系统默认映射（dcs_model_id IS NULL）补充。
    None 键对应本系统默认映射，供无 dcs_model 回路使用。
    """
    from app.services.mode_resolver import build_raw_to_standard_map

    maps: dict[str | None, dict[int, int]] = {}
    for model_id in dcs_model_ids:
        maps[model_id] = await build_raw_to_standard_map(db, model_id)
    return maps


def _mode_value_to_label(
    value: float | None,
    raw_to_standard: dict[int, int] | None = None,
) -> str | None:
    """MODE tag 值 → 控制模式标签（配置驱动回退链）。

    回退链：回路 dcs_model 映射（``raw_to_standard``，含本系统默认映射）→
    raw 值本身（1:1）→ 标准 MODE 前端标签表。
    无 dcs_model 回路传入本系统默认映射（种子数据为 1:1），行为与原硬编码一致。
    """
    if value is None:
        return None
    raw = int(value)
    standard = raw_to_standard.get(raw, raw) if raw_to_standard else raw
    return _STANDARD_MODE_TO_FRONTEND.get(standard, "Unknown")


# 反向映射：控制模式标签 → MODE 值集合
# 与 _mode_value_to_label 保持一致（Cascade 对应 2 和 3）
_CONTROL_MODE_VALUES: dict[str, set[int]] = {
    "manual": {0},
    "auto": {1},
    "cascade": {2, 3},
}


def _control_mode_to_values(control_mode: str) -> list[int]:
    """控制模式标签 → MODE 值列表（大小写不敏感）。

    用于 SQL 层 EXISTS 子查询过滤，将前端传入的 controlMode label
    反向映射为 TagRegistry.current_value 的合法值集合。

    Args:
        control_mode: 控制模式标签（Manual/Auto/Cascade），大小写不敏感

    Returns:
        MODE 值列表（如 "Auto" → [1]）；无法识别时返回空列表
    """
    if not control_mode:
        return []
    return sorted(_CONTROL_MODE_VALUES.get(control_mode.lower(), set()))


async def _get_op_tag_range(
    db: AsyncSession, loop_id: str | None
) -> tuple[float | None, float | None]:
    """查询回路关联 OP Tag 的量程上下限。

    用于 OP 输出限位校验和饱和率算法回退。
    返回 (range_min, range_max)；若 OP Tag 未关联或量程为 NULL，返回 (None, None)。
    """
    if not loop_id:
        return None, None
    result = await db.execute(
        select(TagRegistry.range_min, TagRegistry.range_max)
        .join(LoopTagMapping, LoopTagMapping.tag_id == TagRegistry.id)
        .where(
            LoopTagMapping.loop_id == loop_id,
            LoopTagMapping.tag_role == "OP",
        )
    )
    row = result.first()
    if row is None:
        return None, None
    return float(row[0]) if row[0] is not None else None, (
        float(row[1]) if row[1] is not None else None
    )


def _validate_op_output_limits(
    lower: float | None,
    upper: float | None,
    op_range_min: float | None,
    op_range_max: float | None,
) -> None:
    """校验 OP 输出限位范围。

    校验规则（设计文档 §2.2）：
    - lower < upper
    - lower >= OP Tag range_min（若 range_min 已知）
    - upper <= OP Tag range_max（若 range_max 已知）

    Raises:
        BizError: ERR_OP_LIMIT_OUT_OF_RANGE
    """
    if lower is None and upper is None:
        return
    if lower is not None and upper is not None and lower >= upper:
        raise BizError(
            code="ERR_OP_LIMIT_OUT_OF_RANGE",
            message=f"OP 输出下限位 ({lower}) 必须小于上限位 ({upper})",
            status_code=400,
        )
    if lower is not None and op_range_min is not None and lower < op_range_min:
        raise BizError(
            code="ERR_OP_LIMIT_OUT_OF_RANGE",
            message=f"OP 输出下限位 ({lower}) 不能小于 OP Tag 量程下限 ({op_range_min})",
            status_code=400,
        )
    if upper is not None and op_range_max is not None and upper > op_range_max:
        raise BizError(
            code="ERR_OP_LIMIT_OUT_OF_RANGE",
            message=f"OP 输出上限位 ({upper}) 不能大于 OP Tag 量程上限 ({op_range_max})",
            status_code=400,
        )


async def _validate_complex_group(
    db: AsyncSession,
    complex_loop_group_id: str | None,
    complex_role: str | None,
    exclude_loop_id: str | None = None,
) -> None:
    """校验复杂回路分组字段（P4 S4）。

    校验规则：
    1. 一致性：group_id 和 role 须同时为 NULL 或同时非 NULL
    2. 角色取值：role 须为 'MAIN' / 'SUB' / None
    3. MAIN 唯一性：若 role == 'MAIN'，同 group_id 下不能已有其他 MAIN 回路
    4. group_id 格式：须为合法 UUID

    Raises:
        BizError: ERR_COMPLEX_GROUP_COHERENCE / ERR_COMPLEX_GROUP_ROLE_INVALID
                  / ERR_COMPLEX_GROUP_MAIN_EXISTS / ERR_COMPLEX_GROUP_ID_INVALID
    """
    # 1. 一致性校验
    has_group = complex_loop_group_id is not None
    has_role = complex_role is not None
    if has_group != has_role:
        raise BizError(
            code="ERR_COMPLEX_GROUP_COHERENCE",
            message="复杂回路分组 ID 与角色须同时为空或同时设置",
            status_code=400,
        )

    # 两者均为 NULL = 普通单回路，无需进一步校验
    if not has_group:
        return

    # 2. 角色取值校验
    if complex_role not in ("MAIN", "SUB"):
        raise BizError(
            code="ERR_COMPLEX_GROUP_ROLE_INVALID",
            message=f"复杂回路角色须为 MAIN 或 SUB，当前为 {complex_role}",
            status_code=400,
        )

    # 3. group_id UUID 格式校验
    try:
        import uuid

        uuid.UUID(str(complex_loop_group_id))
    except (ValueError, AttributeError):
        raise BizError(
            code="ERR_COMPLEX_GROUP_ID_INVALID",
            message=f"复杂回路分组 ID 须为合法 UUID，当前为 {complex_loop_group_id}",
            status_code=400,
        ) from None

    # 4. MAIN 唯一性校验
    if complex_role == COMPLEX_ROLE_MAIN:
        existing_main = await db.execute(
            select(LoopLedger.id).where(
                LoopLedger.complex_loop_group_id == complex_loop_group_id,
                LoopLedger.complex_role == COMPLEX_ROLE_MAIN,
            )
        )
        existing_id = existing_main.scalar_one_or_none()
        if existing_id is not None and str(existing_id) != str(exclude_loop_id or ""):
            raise BizError(
                code="ERR_COMPLEX_GROUP_MAIN_EXISTS",
                message="该分组已存在 MAIN 回路，一个分组仅允许一个 MAIN",
                status_code=400,
            )


async def create_loop(
    db: AsyncSession,
    tag_name: str,
    description: str | None,
    unit_id: str | None,
    score_weights: dict | None,
    is_active: bool,
    remark: str | None,
    operator: str,
    loop_type: str | None = None,
    control_type: str | None = None,
    importance_level: int | None = None,
    include_in_evaluation: bool | None = None,
    modeattr_tag_id: str | None = None,
    data_retention_days: int | None = None,
    op_output_lower_limit: float | None = None,
    op_output_upper_limit: float | None = None,
    dcs_model_id: str | None = None,
    ideal_settling_time: float | None = None,
    complex_loop_group_id: str | None = None,
    complex_role: str | None = None,
) -> dict:
    """创建回路。

    Raises:
        BizError: ERR_LOOP_DUPLICATE (tag_name 重复) / ERR_NODE_NOT_FOUND (unit_id 不存在)
                  / ERR_OP_LIMIT_OUT_OF_RANGE (OP 输出限位校验失败)
                  / ERR_COMPLEX_GROUP_* (复杂回路分组校验失败)
    """
    # tag_name 唯一校验
    existing = await db.execute(select(LoopLedger).where(LoopLedger.tag_name == tag_name))
    if existing.scalar_one_or_none() is not None:
        raise BizError(
            code="ERR_LOOP_DUPLICATE",
            message=f"回路位号 {tag_name} 已存在",
            status_code=400,
        )

    # 校验 unit_id 存在
    if unit_id:
        unit_result = await db.execute(select(PlantNode).where(PlantNode.id == unit_id))
        if unit_result.scalar_one_or_none() is None:
            raise BizError(
                code="ERR_NODE_NOT_FOUND",
                message="所属单元不存在",
                status_code=404,
            )

    # v6.1 校验 OP 输出限位范围（创建时 loop_id 还未生成，仅校验 lower < upper）
    _validate_op_output_limits(
        lower=op_output_lower_limit,
        upper=op_output_upper_limit,
        op_range_min=None,  # 创建时 OP Tag 未关联，无法校验范围
        op_range_max=None,
    )

    # P4 S4：校验复杂回路分组字段
    await _validate_complex_group(db, complex_loop_group_id, complex_role)

    # 新建回路默认状态：INACTIVE（未激活）或 PARTIAL（已激活但无 Tag）
    status = "PARTIAL" if is_active else "INACTIVE"
    # v5.3 对齐 FDS §5.2.3 / DDS v4.1：include_in_evaluation 默认 True（参与评估）
    if include_in_evaluation is None:
        include_in_evaluation = True
    # importance_level NOT NULL，缺省兜底为 2（对齐模型 default 与国标附表2）
    if importance_level is None:
        importance_level = 2

    loop = LoopLedger(
        id=str(uuid4()),
        tag_name=tag_name,
        description=description,
        unit_id=unit_id,
        is_active=is_active,
        status=status,
        loop_type=loop_type,
        control_type=control_type,
        importance_level=importance_level,
        include_in_evaluation=include_in_evaluation,
        modeattr_tag_id=modeattr_tag_id,
        data_retention_days=data_retention_days,
        op_output_lower_limit=op_output_lower_limit,
        op_output_upper_limit=op_output_upper_limit,
        dcs_model_id=dcs_model_id,
        ideal_settling_time=ideal_settling_time,
        complex_loop_group_id=complex_loop_group_id,
        complex_role=complex_role,
        score_weights=score_weights,
        remark=remark,
        created_by=operator,
        updated_by=operator,
    )
    db.add(loop)
    await db.flush()

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="LOOP_CREATE",
        target_type="loop_ledger",
        target_id=str(loop.id),
        after_value=json.dumps(
            {
                "tagName": tag_name,
                "description": description,
                "unitId": unit_id,
                "isActive": is_active,
                "status": status,
                "loopType": loop_type,
                "controlType": control_type,
                "importanceLevel": importance_level,
                "includeInEvaluation": include_in_evaluation,
                "modeattrTagId": modeattr_tag_id,
                "dataRetentionDays": data_retention_days,
                "opOutputLowerLimit": op_output_lower_limit,
                "opOutputUpperLimit": op_output_upper_limit,
                "dcsModelId": dcs_model_id,
                "idealSettlingTime": ideal_settling_time,
            },
            ensure_ascii=False,
        ),
    )
    await db.commit()

    # 通知实时订阅 Leader 刷新订阅集合（fire-and-forget，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="loop-create")

    return {
        "loopId": str(loop.id),
        "tagName": loop.tag_name,
        "description": loop.description,
        "unitId": str(loop.unit_id) if loop.unit_id else None,
        "status": loop.status,
        "loopType": loop.loop_type,
        "controlType": loop.control_type,
        "importanceLevel": loop.importance_level,
        "includeInEvaluation": loop.include_in_evaluation,
        "modeattrTagId": str(loop.modeattr_tag_id) if loop.modeattr_tag_id else None,
        "dataRetentionDays": loop.data_retention_days,
        "opOutputLowerLimit": (
            float(loop.op_output_lower_limit) if loop.op_output_lower_limit is not None else None
        ),
        "opOutputUpperLimit": (
            float(loop.op_output_upper_limit) if loop.op_output_upper_limit is not None else None
        ),
        "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
        "idealSettlingTime": (
            float(loop.ideal_settling_time) if loop.ideal_settling_time is not None else None
        ),
        "complexLoopGroupId": (
            str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
        ),
        "complexRole": loop.complex_role,
        "isActive": bool(loop.is_active),
        "scoreWeights": loop.score_weights,
        "remark": loop.remark,
        "createdAt": loop.created_at.isoformat() if loop.created_at else None,
        "createdBy": loop.created_by,
    }


async def get_loop_detail(db: AsyncSession, loop_id: str) -> dict:
    """获取回路详情（含 basicInfo/tagMapping/runtimeParams/aasSyncStatus）。

    Raises:
        BizError: ERR_LOOP_NOT_FOUND
    """
    result = await db.execute(select(LoopLedger).where(LoopLedger.id == loop_id))
    loop = result.scalar_one_or_none()
    if loop is None:
        raise BizError(
            code="ERR_LOOP_NOT_FOUND",
            message="回路不存在",
            status_code=404,
        )

    unit_name = await _get_unit_name(db, str(loop.unit_id) if loop.unit_id else None)

    # 查询 Tag 关联 + Tag 详情
    mappings = await _get_loop_tag_mappings(db, loop_id)
    tag_ids = [str(m.tag_id) for m in mappings.values()]
    tags_map: dict[str, TagRegistry] = {}
    if tag_ids:
        t_result = await db.execute(select(TagRegistry).where(TagRegistry.id.in_(tag_ids)))
        for t in t_result.scalars().all():
            tags_map[str(t.id)] = t

    # 构建 tagMapping 块
    tag_mapping: dict[str, dict] = {}
    for role in ALL_ROLES:
        field = ROLE_TO_FIELD[role]
        mapping = mappings.get(role)
        if mapping and str(mapping.tag_id) in tags_map:
            tag = tags_map[str(mapping.tag_id)]
            tag_mapping[field] = {
                "tagId": str(tag.id),
                "tagName": tag.tag_name,
                "required": role in REQUIRED_ROLES,
                "associated": True,
            }
        else:
            tag_mapping[field] = {
                "tagId": None,
                "tagName": None,
                "required": role in REQUIRED_ROLES,
                "associated": False,
            }

    # 批量从 Redis 读取实时值（优先于数据库 current_value）
    redis_cache: dict[str, dict] = {}
    try:
        from app.services.data_source.realtime_subscriber import get_subscriber

        subscriber = get_subscriber()
        all_tag_names = [tag.tag_name for tag in tags_map.values() if tag.tag_name]
        if all_tag_names:
            cached_list = await subscriber.get_cached_values(all_tag_names)
            for item in cached_list:
                tc = item.get("tagCode")
                if tc:
                    redis_cache[tc] = item
    except Exception as exc:  # noqa: BLE001
        logger.warning("从 Redis 读取实时值失败，回退到数据库值: %s", exc)

    # 构建 runtimeParams（优先从 Redis 缓存读取）
    # WS-C 6-3：按回路 dcs_model_id 构建 MODE 解析映射（dcs_mode_mapping 回退链）
    loop_dcs_model_id = str(loop.dcs_model_id) if loop.dcs_model_id else None
    raw_to_standard_maps = await _build_raw_to_standard_maps(db, {loop_dcs_model_id})
    raw_to_standard = raw_to_standard_maps.get(loop_dcs_model_id) or (
        raw_to_standard_maps.get(None) or {}
    )
    runtime_params: dict = {
        "controlMode": None,
        "pidP": None,
        "pidI": None,
        "pidD": None,
        "readAt": None,
    }
    for role in ALL_ROLES:
        mapping = mappings.get(role)
        if mapping and str(mapping.tag_id) in tags_map:
            tag = tags_map[str(mapping.tag_id)]
            cached = redis_cache.get(tag.tag_name)
            if cached:
                try:
                    if role == "MODE":
                        mode_val = float(cached.get("value"))
                        runtime_params["controlMode"] = _mode_value_to_label(
                            mode_val, raw_to_standard
                        )
                    elif role == "PID_P":
                        runtime_params["pidP"] = float(cached.get("value"))
                    elif role == "PID_I":
                        runtime_params["pidI"] = float(cached.get("value"))
                    elif role == "PID_D":
                        runtime_params["pidD"] = float(cached.get("value"))
                except (TypeError, ValueError):
                    if role == "MODE":
                        runtime_params["controlMode"] = _mode_value_to_label(
                            tag.current_value, raw_to_standard
                        )
                    elif role == "PID_P":
                        runtime_params["pidP"] = tag.current_value
                    elif role == "PID_I":
                        runtime_params["pidI"] = tag.current_value
                    elif role == "PID_D":
                        runtime_params["pidD"] = tag.current_value
                if cached.get("collectTime"):
                    ts = cached["collectTime"]
                    if runtime_params["readAt"] is None or ts > runtime_params["readAt"]:
                        runtime_params["readAt"] = ts
            else:
                if role == "MODE":
                    runtime_params["controlMode"] = _mode_value_to_label(
                        tag.current_value, raw_to_standard
                    )
                elif role == "PID_P":
                    runtime_params["pidP"] = tag.current_value
                elif role == "PID_I":
                    runtime_params["pidI"] = tag.current_value
                elif role == "PID_D":
                    runtime_params["pidD"] = tag.current_value
                if tag.last_sync_at:
                    ts = tag.last_sync_at.isoformat()
                    if runtime_params["readAt"] is None or ts > runtime_params["readAt"]:
                        runtime_params["readAt"] = ts

    # aasSyncStatus
    last_sync = None
    for tag in tags_map.values():
        if tag.last_sync_at:
            ts = tag.last_sync_at
            if last_sync is None or ts > last_sync:
                last_sync = ts

    # v6.1 查询 PV/OP Tag 量程与单位
    pv_range_info: dict | None = None
    pv_unit: str | None = None
    op_range_info: dict | None = None
    op_unit: str | None = None
    for role in ("PV", "OP"):
        mapping = mappings.get(role)
        if mapping and str(mapping.tag_id) in tags_map:
            tag = tags_map[str(mapping.tag_id)]
            range_info = {
                "min": float(tag.range_min) if tag.range_min is not None else None,
                "max": float(tag.range_max) if tag.range_max is not None else None,
            }
            if role == "PV":
                pv_range_info = range_info
                pv_unit = tag.unit
            elif role == "OP":
                op_range_info = range_info
                op_unit = tag.unit

    return {
        "basicInfo": {
            "loopId": str(loop.id),
            "tagName": loop.tag_name,
            "description": loop.description,
            "unitId": str(loop.unit_id) if loop.unit_id else None,
            "unitName": unit_name,
            "isActive": bool(loop.is_active),
            "status": loop.status,
            "loopType": loop.loop_type,
            "controlType": loop.control_type,
            "importanceLevel": loop.importance_level,
            "includeInEvaluation": loop.include_in_evaluation,
            "modeattrTagId": str(loop.modeattr_tag_id) if loop.modeattr_tag_id else None,
            "dataRetentionDays": loop.data_retention_days,
            "scoreWeights": loop.score_weights,
            "remark": loop.remark,
            # v6.1 新增：量程与限位
            "pvRange": pv_range_info,
            "pvUnit": pv_unit,
            "opRange": op_range_info,
            "opUnit": op_unit,
            "opOutputLowerLimit": (
                float(loop.op_output_lower_limit)
                if loop.op_output_lower_limit is not None
                else None
            ),
            "opOutputUpperLimit": (
                float(loop.op_output_upper_limit)
                if loop.op_output_upper_limit is not None
                else None
            ),
            "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
            "idealSettlingTime": (
                float(loop.ideal_settling_time) if loop.ideal_settling_time is not None else None
            ),
            "complexLoopGroupId": (
                str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
            ),
            "complexRole": loop.complex_role,
            "createdAt": loop.created_at.isoformat() if loop.created_at else None,
            "createdBy": loop.created_by,
            "updatedAt": loop.updated_at.isoformat() if loop.updated_at else None,
            "updatedBy": loop.updated_by,
        },
        "tagMapping": tag_mapping,
        "runtimeParams": runtime_params,
        "aasSyncStatus": {
            "lastSyncAt": last_sync.isoformat() if last_sync else None,
            "associatedTagCount": len(mappings),
        },
    }


async def update_loop(
    db: AsyncSession,
    loop_id: str,
    operator: str,
    description: str | None = None,
    score_weights: dict | None = None,
    is_active: bool | None = None,
    remark: str | None = None,
    loop_type: str | None = None,
    control_type: str | None = None,
    importance_level: int | None = None,
    include_in_evaluation: bool | None = None,
    modeattr_tag_id: str | None = None,
    data_retention_days: int | None = None,
    op_output_lower_limit: float | None = None,
    op_output_upper_limit: float | None = None,
    dcs_model_id: str | None = None,
    ideal_settling_time: float | None = None,
    unit_id: str | None = None,
    _op_lower_set: bool = False,
    _op_upper_set: bool = False,
    _dcs_model_id_set: bool = False,
    _ideal_settling_time_set: bool = False,
    complex_loop_group_id: str | None = None,
    complex_role: str | None = None,
    _complex_group_set: bool = False,
    _complex_role_set: bool = False,
) -> dict:
    """更新回路（描述/评分权重/启用状态/备注/回路类型/控制类型/重要等级/参评/APC位号/保留周期/OP输出限位/理想稳态时间/复杂回路分组）。

    Raises:
        BizError: ERR_LOOP_NOT_FOUND / ERR_NODE_NOT_FOUND / ERR_OP_LIMIT_OUT_OF_RANGE
                  / ERR_COMPLEX_GROUP_*
    """
    result = await db.execute(select(LoopLedger).where(LoopLedger.id == loop_id))
    loop = result.scalar_one_or_none()
    if loop is None:
        raise BizError(
            code="ERR_LOOP_NOT_FOUND",
            message="回路不存在",
            status_code=404,
        )

    # WS-C 6-2：unitId 更新时校验目标工艺单元节点存在（与 create_loop 一致处理）
    if unit_id is not None:
        unit_result = await db.execute(select(PlantNode).where(PlantNode.id == unit_id))
        if unit_result.scalar_one_or_none() is None:
            raise BizError(
                code="ERR_NODE_NOT_FOUND",
                message="所属单元不存在",
                status_code=404,
            )

    # v6.1 校验 OP 输出限位范围
    # 查询 OP Tag 量程作为校验基准
    op_range_min, op_range_max = await _get_op_tag_range(db, loop_id)
    # 计算生效的 lower/upper（更新后的值优先，否则保持原值）
    effective_lower = (
        op_output_lower_limit if op_output_lower_limit is not None else loop.op_output_lower_limit
    )
    effective_upper = (
        op_output_upper_limit if op_output_upper_limit is not None else loop.op_output_upper_limit
    )
    _validate_op_output_limits(
        lower=effective_lower,
        upper=effective_upper,
        op_range_min=op_range_min,
        op_range_max=op_range_max,
    )

    # P4 S4：校验复杂回路分组（生效值 = 更新值优先，否则保持原值）
    effective_group = (
        complex_loop_group_id
        if _complex_group_set
        else (str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None)
    )
    effective_role = complex_role if _complex_role_set else loop.complex_role
    await _validate_complex_group(db, effective_group, effective_role, exclude_loop_id=loop_id)

    before = {
        "description": loop.description,
        "unitId": str(loop.unit_id) if loop.unit_id else None,
        "scoreWeights": loop.score_weights,
        "isActive": loop.is_active,
        "remark": loop.remark,
        "loopType": loop.loop_type,
        "controlType": loop.control_type,
        "importanceLevel": loop.importance_level,
        "includeInEvaluation": loop.include_in_evaluation,
        "modeattrTagId": str(loop.modeattr_tag_id) if loop.modeattr_tag_id else None,
        "dataRetentionDays": loop.data_retention_days,
        "opOutputLowerLimit": (
            float(loop.op_output_lower_limit) if loop.op_output_lower_limit is not None else None
        ),
        "opOutputUpperLimit": (
            float(loop.op_output_upper_limit) if loop.op_output_upper_limit is not None else None
        ),
        "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
        "idealSettlingTime": (
            float(loop.ideal_settling_time) if loop.ideal_settling_time is not None else None
        ),
        "complexLoopGroupId": (
            str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
        ),
        "complexRole": loop.complex_role,
    }
    before_json = json.dumps(before, ensure_ascii=False, default=str)

    if description is not None:
        loop.description = description
    if unit_id is not None:
        loop.unit_id = unit_id
    if score_weights is not None:
        loop.score_weights = score_weights
    if is_active is not None:
        loop.is_active = is_active
    if remark is not None:
        loop.remark = remark
    if loop_type is not None:
        loop.loop_type = loop_type
    if control_type is not None:
        loop.control_type = control_type
    if importance_level is not None:
        loop.importance_level = importance_level
    if include_in_evaluation is not None:
        loop.include_in_evaluation = include_in_evaluation
    if modeattr_tag_id is not None:
        loop.modeattr_tag_id = modeattr_tag_id
    if data_retention_days is not None:
        loop.data_retention_days = data_retention_days
    # v6.1：使用 _op_lower_set / _op_upper_set 标记区分"未传递"和"传递了 NULL"
    # 允许用户通过 PUT null 清空 OP 输出限位（恢复默认值）
    if _op_lower_set:
        loop.op_output_lower_limit = op_output_lower_limit
    if _op_upper_set:
        loop.op_output_upper_limit = op_output_upper_limit
    # v6.1 DCS 型号关联：支持通过 PUT null 清空（回退到本系统默认 MODE 映射）
    if _dcs_model_id_set:
        loop.dcs_model_id = dcs_model_id
    # 理想稳态时间：支持通过 PUT null 清空（恢复按控制类型默认值）
    if _ideal_settling_time_set:
        loop.ideal_settling_time = ideal_settling_time
    # P4 S4：复杂回路分组——支持通过 PUT null 清空（解除分组）
    # _complex_group_set / _complex_role_set 标记区分"未传递"和"传递了 NULL"
    if _complex_group_set:
        loop.complex_loop_group_id = complex_loop_group_id
    if _complex_role_set:
        loop.complex_role = complex_role
    loop.updated_by = operator

    # 重新推导 status
    new_status = await derive_loop_status(db, loop)
    loop.status = new_status

    after = {
        "description": loop.description,
        "unitId": str(loop.unit_id) if loop.unit_id else None,
        "scoreWeights": loop.score_weights,
        "isActive": loop.is_active,
        "remark": loop.remark,
        "loopType": loop.loop_type,
        "controlType": loop.control_type,
        "importanceLevel": loop.importance_level,
        "includeInEvaluation": loop.include_in_evaluation,
        "modeattrTagId": str(loop.modeattr_tag_id) if loop.modeattr_tag_id else None,
        "dataRetentionDays": loop.data_retention_days,
        "opOutputLowerLimit": (
            float(loop.op_output_lower_limit) if loop.op_output_lower_limit is not None else None
        ),
        "opOutputUpperLimit": (
            float(loop.op_output_upper_limit) if loop.op_output_upper_limit is not None else None
        ),
        "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
        "idealSettlingTime": (
            float(loop.ideal_settling_time) if loop.ideal_settling_time is not None else None
        ),
        "complexLoopGroupId": (
            str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
        ),
        "complexRole": loop.complex_role,
        "status": new_status,
    }
    after_json = json.dumps(after, ensure_ascii=False, default=str)

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="LOOP_UPDATE",
        target_type="loop_ledger",
        target_id=str(loop.id),
        before_value=before_json,
        after_value=after_json,
    )
    await db.commit()
    # 刷新对象以获取 onupdate=func.now() 生成的 updated_at（async session
    # 不能同步 lazy load，否则触发 MissingGreenlet）
    await db.refresh(loop)

    # 通知实时订阅 Leader 刷新订阅集合（fire-and-forget，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="loop-update")

    return {
        "loopId": str(loop.id),
        "description": loop.description,
        "unitId": str(loop.unit_id) if loop.unit_id else None,
        "scoreWeights": loop.score_weights,
        "isActive": bool(loop.is_active),
        "remark": loop.remark,
        "loopType": loop.loop_type,
        "controlType": loop.control_type,
        "importanceLevel": loop.importance_level,
        "includeInEvaluation": loop.include_in_evaluation,
        "modeattrTagId": str(loop.modeattr_tag_id) if loop.modeattr_tag_id else None,
        "dataRetentionDays": loop.data_retention_days,
        "opOutputLowerLimit": (
            float(loop.op_output_lower_limit) if loop.op_output_lower_limit is not None else None
        ),
        "opOutputUpperLimit": (
            float(loop.op_output_upper_limit) if loop.op_output_upper_limit is not None else None
        ),
        "dcsModelId": str(loop.dcs_model_id) if loop.dcs_model_id else None,
        "idealSettlingTime": (
            float(loop.ideal_settling_time) if loop.ideal_settling_time is not None else None
        ),
        "complexLoopGroupId": (
            str(loop.complex_loop_group_id) if loop.complex_loop_group_id else None
        ),
        "complexRole": loop.complex_role,
        "updatedAt": loop.updated_at.isoformat() if loop.updated_at else None,
        "updatedBy": loop.updated_by,
    }


async def delete_loop(
    db: AsyncSession,
    loop_id: str,
    operator: str,
) -> dict:
    """硬删除回路（与前端"不可恢复"承诺对齐；批量删除保持软删）。

    级联解绑（WS-C 6-4）：删除回路前先删除 LoopTagMapping 关联记录
    （通常 7 条：PV/SP/OP/MODE/PID_P/PID_I/PID_D）。
    解除关联后不再被任何回路引用的 Tag，其 is_linked 一并清除（is_linked 由映射派生）。
    回路本体硬删除：db.delete(loop)，ON DELETE CASCADE 自动级联清理
    kpi_snapshot/action_tracker/diagnosis_result/tuning_record/loop_mode_mapping/
    kpi_custom/diagnosis_tag/diagnosis_task/loop_confidence_latest/process_model_version。

    Raises:
        BizError: ERR_LOOP_NOT_FOUND
    """
    result = await db.execute(select(LoopLedger).where(LoopLedger.id == loop_id))
    loop = result.scalar_one_or_none()
    if loop is None:
        raise BizError(
            code="ERR_LOOP_NOT_FOUND",
            message="回路不存在",
            status_code=404,
        )

    before_json = json.dumps(
        {"tagName": loop.tag_name, "is_active": loop.is_active, "status": loop.status},
        ensure_ascii=False,
    )

    # 级联解绑：删除本回路的 LoopTagMapping 关联记录
    mappings_result = await db.execute(
        select(LoopTagMapping).where(LoopTagMapping.loop_id == loop_id)
    )
    mappings = mappings_result.scalars().all()
    mapped_tag_ids = [str(m.tag_id) for m in mappings]
    if mappings:
        await db.execute(delete(LoopTagMapping).where(LoopTagMapping.loop_id == loop_id))
        # is_linked 由映射派生：仅当 Tag 不再被任何回路引用时才清除
        for tag_id in mapped_tag_ids:
            ref_count_result = await db.execute(
                select(func.count())
                .select_from(LoopTagMapping)
                .where(LoopTagMapping.tag_id == tag_id)
            )
            if (ref_count_result.scalar() or 0) > 0:
                continue
            t_result = await db.execute(select(TagRegistry).where(TagRegistry.id == tag_id))
            tag = t_result.scalar_one_or_none()
            if tag:
                tag.is_linked = False

    # 硬删除（与前端"不可恢复"承诺对齐；批量删除保持软删可恢复）
    # ON DELETE CASCADE 自动级联清理所有关联表数据
    await db.delete(loop)

    after_json = json.dumps(
        {"tagName": loop.tag_name, "deleted": True},
        ensure_ascii=False,
    )

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="LOOP_DELETE",
        target_type="loop_ledger",
        target_id=loop_id,
        before_value=before_json,
        after_value=after_json,
    )
    await db.commit()

    # 级联解绑已清除 is_linked：通知实时订阅 Leader 刷新订阅集合（免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="loop-delete")

    return {
        "loopId": loop_id,
        "deleted": True,
        "deletedAt": datetime.now(UTC).replace(tzinfo=None).isoformat(),
    }


# ---------------------------------------------------------------------------
# 复杂回路分组 (P4 S4)
# ---------------------------------------------------------------------------


async def batch_group_loops(
    db: AsyncSession,
    loop_ids: list[str],
    main_loop_id: str,
    operator: str,
) -> dict:
    """批量建立复杂回路分组（P4 S4）。

    将 2-20 个回路归为一个复杂控制回路（串级/超驰等），系统自动生成新的 group_id，
    指定其中一个为 MAIN（聚合代表），其余自动为 SUB。

    Args:
        loop_ids: 回路 ID 列表（2-20 个）
        main_loop_id: 主回路 ID（须在 loop_ids 中）
        operator: 操作人

    Raises:
        BizError: ERR_COMPLEX_GROUP_MAIN_NOT_IN_LIST / ERR_LOOP_NOT_FOUND
                  / ERR_COMPLEX_GROUP_MAIN_EXISTS（所选主回路已是其他分组 MAIN）
    """
    # 1. 主回路须在列表中
    if main_loop_id not in loop_ids:
        raise BizError(
            code="ERR_COMPLEX_GROUP_MAIN_NOT_IN_LIST",
            message="主回路 ID 须在 loopIds 列表中",
            status_code=400,
        )

    # 2. 查询全部目标回路（一次 IN 查询，去重避免重复 ID）
    unique_ids = list(dict.fromkeys(loop_ids))
    result = await db.execute(select(LoopLedger).where(LoopLedger.id.in_(unique_ids)))
    loops = result.scalars().all()
    if len(loops) != len(unique_ids):
        found_ids = {str(lp.id) for lp in loops}
        missing = [i for i in unique_ids if i not in found_ids]
        raise BizError(
            code="ERR_LOOP_NOT_FOUND",
            message=f"以下回路不存在: {missing}",
            status_code=404,
        )

    # 3. 校验主回路当前是否已是其他分组的 MAIN（避免跨分组重复 MAIN）
    main_loop = next((lp for lp in loops if str(lp.id) == main_loop_id), None)
    if main_loop is None:
        raise BizError(
            code="ERR_LOOP_NOT_FOUND",
            message=f"主回路 {main_loop_id} 不存在",
            status_code=404,
        )
    if main_loop.complex_loop_group_id and main_loop.complex_role == COMPLEX_ROLE_MAIN:
        raise BizError(
            code="ERR_COMPLEX_GROUP_MAIN_EXISTS",
            message=(
                f"回路 {main_loop.tag_name} 已是其他分组的 MAIN 回路，请先解除其原分组再行分配"
            ),
            status_code=400,
        )

    # 4. 生成新分组 ID，批量赋值
    new_group_id = str(uuid4())
    assignments: list[dict] = []
    for lp in loops:
        role = COMPLEX_ROLE_MAIN if str(lp.id) == main_loop_id else "SUB"
        lp.complex_loop_group_id = new_group_id
        lp.complex_role = role
        lp.updated_by = operator
        assignments.append({"loopId": str(lp.id), "tagName": lp.tag_name, "role": role})

    # 5. 审计日志
    await _write_audit(
        db=db,
        operator=operator,
        operation_type="LOOP_BATCH_GROUPING",
        target_type="loop_ledger",
        target_id=new_group_id,
        after_value=json.dumps(
            {
                "groupId": new_group_id,
                "mainLoopId": main_loop_id,
                "mainTagName": main_loop.tag_name,
                "memberCount": len(loops),
                "assignments": assignments,
            },
            ensure_ascii=False,
        ),
    )
    await db.commit()

    # 按 MAIN 在前、SUB 其后排序，便于前端展示
    assignments.sort(key=lambda a: 0 if a["role"] == COMPLEX_ROLE_MAIN else 1)

    return {
        "groupId": new_group_id,
        "affected": len(loops),
        "assignments": assignments,
    }


async def list_complex_groups(db: AsyncSession) -> list[dict]:
    """查询所有复杂回路分组（P4 S4）。

    仅统计 is_active=True 的回路，按 group_id 聚合，提取 MAIN 回路位号与组成员数。
    单回路（complex_loop_group_id IS NULL）不在统计范围内。

    Returns:
        分组列表，每项含 groupId / mainTagName / memberCount
    """
    stmt = text(
        """
        SELECT complex_loop_group_id,
               MAX(CASE WHEN complex_role = 'MAIN' THEN tag_name END) AS main_tag_name,
               COUNT(*) AS member_count
        FROM loop_ledger
        WHERE complex_loop_group_id IS NOT NULL
          AND is_active = TRUE
        GROUP BY complex_loop_group_id
        ORDER BY MAX(CASE WHEN complex_role = 'MAIN' THEN tag_name END)
        """
    )
    result = await db.execute(stmt)
    rows = result.all()

    return [
        {
            "groupId": str(row[0]),
            "mainTagName": row[1],
            "memberCount": int(row[2]),
        }
        for row in rows
    ]


# ---------------------------------------------------------------------------
# 批量导入导出 (S2-LOOP-009)
# ---------------------------------------------------------------------------

# Excel 列头（18 列，v6.1 扩展：追加控制类型/等级/参评/OP 限位/备注）
# 保持原 12 列顺序不变（向后兼容已有 Excel 模板），末尾追加新列
# v6.1：列头语义化（"自控回路名称"→"回路描述"，"所属区域编号"→"所属单元名称"）
EXPORT_HEADERS = [
    "自控回路编号",
    "回路描述",
    "设定值位号",
    "测量值位号",
    "输出值位号",
    "控制方式位号",
    "所属单元名称",
    "是否启用",
    "比例带",
    "积分时间",
    "微分时间",
    "回路类型",
    # v6.1 新增列
    "控制类型",
    "等级",
    "参评状态",
    "OP输出下限位",
    "OP输出上限位",
    "DCS型号",
    "备注",
]

# 导入时列索引 → Tag 角色（索引从 0 开始）
_IMPORT_ROLE_COLUMNS: dict[int, str] = {
    2: "SP",
    3: "PV",
    4: "OP",
    5: "MODE",
    8: "PID_P",
    9: "PID_I",
    10: "PID_D",
}
# 回路类型列索引（"回路类型"列）
_LOOP_TYPE_COLUMN_INDEX = 11

# v6.1 新增列索引
_CONTROL_TYPE_COLUMN_INDEX = 12
_IMPORTANCE_LEVEL_COLUMN_INDEX = 13
_INCLUDE_IN_EVALUATION_COLUMN_INDEX = 14
_OP_OUTPUT_LOWER_LIMIT_COLUMN_INDEX = 15
_OP_OUTPUT_UPPER_LIMIT_COLUMN_INDEX = 16
_DCS_MODEL_COLUMN_INDEX = 17
_REMARK_COLUMN_INDEX = 18


def _cell_str(value: object) -> str:
    """将 Excel 单元格值转为去除首尾空白的字符串，None/空返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


# 注：回路类型归一化已改走字典（normalize_by_dict，DICT_LOOP_TYPE），
# 原硬编码 _normalize_loop_type 已移除；LOOP_TYPE_TO_CN/FROM_CN 仅保留作
# 测试与统计场景的出厂默认映射。


def _normalize_control_type(raw: str) -> str | None:
    """v6.1：控制类型中英文双向识别。

    接受中文（稳定型/慢速型/快速型/逻辑型）或英文（STABLE/SLOW/FAST/LOGIC），
    统一返回英文枚举值。未知值原样返回（由上层校验）。
    """
    if not raw:
        return None
    val = raw.strip()
    # 中文 → 英文
    if val in CONTROL_TYPE_FROM_CN:
        return CONTROL_TYPE_FROM_CN[val]
    # 英文（大小写不敏感）→ 标准大写
    upper = val.upper()
    if upper in CONTROL_TYPE_TO_CN:
        return upper
    # 未知值原样返回
    return val


async def export_loops(
    db: AsyncSession,
    plant_node_id: str | None = None,
    status: str | None = None,
    keyword: str | None = None,
    control_type: str | None = None,
    importance_level: int | None = None,
    include_in_evaluation: bool | None = None,
    loop_type: str | None = None,
) -> bytes:
    """导出所有回路为 Excel 文件（.xlsx），返回文件字节。

    支持按 plantNodeId/status/keyword/controlType/importanceLevel/
    includeInEvaluation/loopType 筛选（可选）。
    """
    conditions = []
    if plant_node_id:
        conditions.append(LoopLedger.unit_id == plant_node_id)
    if status:
        conditions.append(func.upper(LoopLedger.status) == status.upper())
    if keyword:
        kw = f"%{keyword}%"
        conditions.append(
            or_(
                LoopLedger.tag_name.ilike(kw),
                LoopLedger.description.ilike(kw),
            )
        )
    # v6.1 新增筛选条件
    if control_type:
        conditions.append(LoopLedger.control_type == control_type.upper())
    if importance_level is not None:
        conditions.append(LoopLedger.importance_level == importance_level)
    if include_in_evaluation is not None:
        conditions.append(LoopLedger.include_in_evaluation == include_in_evaluation)
    if loop_type:
        conditions.append(LoopLedger.loop_type == loop_type.upper())

    stmt = select(LoopLedger).order_by(LoopLedger.tag_name)
    for cond in conditions:
        stmt = stmt.where(cond)
    result = await db.execute(stmt)
    loops = result.scalars().all()

    loop_ids = [str(loop.id) for loop in loops]
    unit_ids = [str(loop.unit_id) for loop in loops if loop.unit_id]

    # 批量查询 unit_name
    unit_map: dict[str, str] = {}
    if unit_ids:
        u_result = await db.execute(select(PlantNode).where(PlantNode.id.in_(unit_ids)))
        for node in u_result.scalars().all():
            unit_map[str(node.id)] = node.name

    # 批量查询 Tag 关联 + Tag 名称
    tag_name_map: dict[str, dict[str, str]] = {}
    if loop_ids:
        m_result = await db.execute(
            select(LoopTagMapping, TagRegistry)
            .join(TagRegistry, LoopTagMapping.tag_id == TagRegistry.id)
            .where(LoopTagMapping.loop_id.in_(loop_ids))
        )
        for mapping, tag in m_result:
            tag_name_map.setdefault(str(mapping.loop_id), {})[mapping.tag_role] = tag.tag_name

    # 批量查询 DCS 型号名称（v6.1：DCS 型号列导出）
    dcs_model_map: dict[str, str] = {}
    dcs_model_ids = [str(loop.dcs_model_id) for loop in loops if loop.dcs_model_id]
    if dcs_model_ids:
        dm_result = await db.execute(select(DcsModel).where(DcsModel.id.in_(dcs_model_ids)))
        for dm in dm_result.scalars().all():
            dcs_model_map[str(dm.id)] = dm.name

    # 构建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "回路台账"
    ws.append(EXPORT_HEADERS)

    # 回路类型 code → label（以字典为准；enabled_only=False 含禁用项，
    # 存量自定义 code 兜底显示原值；导入支持 code/label 双向识别）
    loop_type_cn = dict(await get_dict_items(db, DICT_LOOP_TYPE, enabled_only=False))

    for loop in loops:
        tags = tag_name_map.get(str(loop.id), {})
        unit_name = unit_map.get(str(loop.unit_id)) if loop.unit_id else ""
        is_active_str = "是" if loop.is_active else "否"
        # v6.1 新增字段导出
        include_in_eval_str = ""
        if loop.include_in_evaluation is not None:
            include_in_eval_str = "是" if loop.include_in_evaluation else "否"
        importance_level_str = str(loop.importance_level) if loop.importance_level else ""
        # v6.1：枚举值导出为中文（用户友好，便于 Excel 编辑）
        loop_type_str = (
            loop_type_cn.get(loop.loop_type or "", loop.loop_type or "") if loop.loop_type else ""
        )
        control_type_str = (
            CONTROL_TYPE_TO_CN.get(loop.control_type.upper(), loop.control_type)
            if loop.control_type
            else ""
        )
        # OP 限位为空时留空（导入时空值表示使用默认 = OP Tag 量程）
        op_lower_str = (
            str(loop.op_output_lower_limit) if loop.op_output_lower_limit is not None else ""
        )
        op_upper_str = (
            str(loop.op_output_upper_limit) if loop.op_output_upper_limit is not None else ""
        )
        # DCS 型号名称（空值表示使用本系统默认 MODE 映射）
        dcs_model_str = dcs_model_map.get(str(loop.dcs_model_id)) if loop.dcs_model_id else ""
        ws.append(
            [
                loop.tag_name,
                loop.description or "",
                tags.get("SP", ""),
                tags.get("PV", ""),
                tags.get("OP", ""),
                tags.get("MODE", ""),
                unit_name,
                is_active_str,
                tags.get("PID_P", ""),
                tags.get("PID_I", ""),
                tags.get("PID_D", ""),
                loop_type_str,
                # v6.1 新增
                control_type_str,
                importance_level_str,
                include_in_eval_str,
                op_lower_str,
                op_upper_str,
                dcs_model_str,
                loop.remark or "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_loops(
    db: AsyncSession,
    file_bytes: bytes,
    operator: str,
) -> dict:
    """批量导入回路（Excel .xlsx）。

    逐行处理：回路编号已存在则更新，否则新建。
    含非法字符 tag 名的行整行跳过并计入 failed/errors，不中断整体导入。
    返回 {total, inserted, updated, failed, errors[], warnings[]}；
    warnings 记录 tag 重关联（历史数据在新 subtable 下重新开始）的回路。
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(
            code="ERR_FILE_PARSE",
            message=f"Excel 文件解析失败: {exc}",
            status_code=400,
        ) from exc

    ws = wb.active
    # P2：read_only 流式逐行处理，不转 list 全量物化（防超大表打爆 API 进程内存）；
    # 文件大小（10MB）与数据行数（5000 行）上限由 API 层 upload_guard.read_excel_upload 统一拦截
    rows = ws.iter_rows(min_row=2, values_only=True)

    total = 0
    inserted = 0
    updated = 0
    failed = 0
    errors: list[dict] = []
    # P2：tag 重关联记录（loop_id, 回路位号, 变更角色），提交后统一做缓存失效 + warning
    tag_reassigned: list[tuple[str, str, list[str]]] = []

    # 缓存：plant_node name → id，tag name → id，dcs_model name → id
    plant_node_cache: dict[str, str] = {}
    tag_cache: dict[str, str] = {}
    dcs_model_cache: dict[str, str | None] = {}

    for row_idx, row in enumerate(rows, start=2):  # 第 1 行为表头
        total += 1
        tag_name = _cell_str(row[0]) if len(row) > 0 else ""

        if not tag_name:
            errors.append({"row": row_idx, "message": "回路编号不能为空"})
            failed += 1
            continue

        # 解析单元格（不涉及 DB，放在 savepoint 外）
        description = _cell_str(row[1]) if len(row) > 1 else ""
        role_tag_values: dict[str, str] = {}
        for col_idx, role in _IMPORT_ROLE_COLUMNS.items():
            role_tag_values[role] = _cell_str(row[col_idx]) if len(row) > col_idx else ""

        # P2：tag 名白名单校验（与 core/tdengine._TAG_NAME_PATTERN 一致）——
        # 含单引号等非法字符的 tag 名会在 TDengine SQL 拼接时引发语法错误甚至
        # 注入，整行跳过并记录，不中断整体导入
        invalid_tag_names = ([tag_name] if not _TAG_NAME_PATTERN.match(tag_name) else []) + [
            t for t in role_tag_values.values() if t and not _TAG_NAME_PATTERN.match(t)
        ]
        if invalid_tag_names:
            failed += 1
            errors.append(
                {
                    "row": row_idx,
                    "tagName": tag_name,
                    "message": "tag 名含非法字符（仅允许字母/数字/下划线/连字符/点号/空白，"
                    f"1~128 字符）: {', '.join(invalid_tag_names)}",
                }
            )
            logger.warning(
                "Excel 导入跳过非法 tag 名: row=%d, names=%s", row_idx, invalid_tag_names
            )
            continue

        unit_name = _cell_str(row[6]) if len(row) > 6 else ""
        is_active_str = _cell_str(row[7]) if len(row) > 7 else "是"
        is_active = is_active_str in ("是", "true", "True", "1", "YES", "yes", "Y", "y")
        # v6.1：回路类型 / 控制类型 中英文双向识别
        # 回路类型以字典为准（可配置：系统管理 → 字典管理 → 回路类型）；
        # 归一失败（如「电流」未注册）报行级错误并指引，不再原样落库触发约束
        loop_type_raw = (
            _cell_str(row[_LOOP_TYPE_COLUMN_INDEX]) if len(row) > _LOOP_TYPE_COLUMN_INDEX else ""
        )
        loop_type: str | None = None
        if loop_type_raw:
            loop_type = await normalize_by_dict(db, DICT_LOOP_TYPE, loop_type_raw)
            if loop_type is None:
                failed += 1
                errors.append(
                    {
                        "row": row_idx,
                        "tagName": tag_name,
                        "message": (
                            f"回路类型无效: {loop_type_raw}，"
                            f"支持的类型：{await dict_items_hint(db, DICT_LOOP_TYPE)}；"
                            "自定义类型请先在「系统管理 → 字典管理 → 回路类型」中添加"
                        ),
                    }
                )
                continue
        control_type_raw = (
            _cell_str(row[_CONTROL_TYPE_COLUMN_INDEX])
            if len(row) > _CONTROL_TYPE_COLUMN_INDEX
            else ""
        )
        control_type = _normalize_control_type(control_type_raw)
        importance_level_str = (
            _cell_str(row[_IMPORTANCE_LEVEL_COLUMN_INDEX])
            if len(row) > _IMPORTANCE_LEVEL_COLUMN_INDEX
            else ""
        )
        importance_level = None
        if importance_level_str:
            try:
                importance_level = int(importance_level_str)
                if importance_level not in (1, 2, 3):
                    importance_level = None
            except ValueError:
                importance_level = None
        include_in_eval_str = (
            _cell_str(row[_INCLUDE_IN_EVALUATION_COLUMN_INDEX])
            if len(row) > _INCLUDE_IN_EVALUATION_COLUMN_INDEX
            else ""
        )
        # 空值表示不修改（保持原值）；非空时按 是/否 解析
        include_in_evaluation: bool | None = None
        if include_in_eval_str:
            include_in_evaluation = include_in_eval_str in (
                "是",
                "true",
                "True",
                "1",
                "YES",
                "yes",
                "Y",
                "y",
            )
        # OP 限位：空值表示使用默认（NULL），非空时解析为 float
        op_lower_str = (
            _cell_str(row[_OP_OUTPUT_LOWER_LIMIT_COLUMN_INDEX])
            if len(row) > _OP_OUTPUT_LOWER_LIMIT_COLUMN_INDEX
            else ""
        )
        op_output_lower_limit: float | None = None
        if op_lower_str:
            try:
                op_output_lower_limit = float(op_lower_str)
            except ValueError:
                op_output_lower_limit = None
        op_upper_str = (
            _cell_str(row[_OP_OUTPUT_UPPER_LIMIT_COLUMN_INDEX])
            if len(row) > _OP_OUTPUT_UPPER_LIMIT_COLUMN_INDEX
            else ""
        )
        op_output_upper_limit: float | None = None
        if op_upper_str:
            try:
                op_output_upper_limit = float(op_upper_str)
            except ValueError:
                op_output_upper_limit = None
        # DCS 型号：空值表示使用本系统默认 MODE 映射（dcs_model_id=NULL）
        dcs_model_name = (
            _cell_str(row[_DCS_MODEL_COLUMN_INDEX]) if len(row) > _DCS_MODEL_COLUMN_INDEX else ""
        )
        dcs_model_id: str | None = None
        if dcs_model_name:
            if dcs_model_name in dcs_model_cache:
                dcs_model_id = dcs_model_cache[dcs_model_name]
            else:
                dm_result = await db.execute(
                    select(DcsModel).where(DcsModel.name == dcs_model_name)
                )
                dm = dm_result.scalar_one_or_none()
                if dm is None:
                    # 尝试按 code 匹配
                    dm_result = await db.execute(
                        select(DcsModel).where(DcsModel.code == dcs_model_name)
                    )
                    dm = dm_result.scalar_one_or_none()
                if dm is not None:
                    dcs_model_id = str(dm.id)
                else:
                    dcs_model_id = None  # 未找到型号，置空使用默认
                dcs_model_cache[dcs_model_name] = dcs_model_id
        remark = (
            _cell_str(row[_REMARK_COLUMN_INDEX]) if len(row) > _REMARK_COLUMN_INDEX else ""
        ) or None

        is_update = False
        try:
            # 使用 SAVEPOINT 保证单行失败不影响其他行
            async with db.begin_nested():
                is_update, changed_loop_id, changed_roles = await _import_one_row(
                    db=db,
                    tag_name=tag_name,
                    description=description,
                    unit_name=unit_name,
                    is_active=is_active,
                    role_tag_values=role_tag_values,
                    operator=operator,
                    plant_node_cache=plant_node_cache,
                    tag_cache=tag_cache,
                    loop_type=loop_type,
                    control_type=control_type,
                    importance_level=importance_level,
                    include_in_evaluation=include_in_evaluation,
                    op_output_lower_limit=op_output_lower_limit,
                    op_output_upper_limit=op_output_upper_limit,
                    dcs_model_id=dcs_model_id,
                    remark=remark,
                )
        except Exception as exc:  # noqa: BLE001
            failed += 1
            errors.append({"row": row_idx, "tagName": tag_name, "message": str(exc)})
            continue

        if changed_roles:
            tag_reassigned.append((changed_loop_id, tag_name, changed_roles))
        if is_update:
            updated += 1
        else:
            inserted += 1

    await db.commit()

    # P2：tag 重关联后置处理（warning + 缓存失效），提交后执行——
    # 若提前失效后事务回滚，会导致缓存被无谓清空且 warning 与库内状态不符
    warnings: list[str] = []
    for reassigned_loop_id, reassigned_loop_tag, reassigned_roles in tag_reassigned:
        warnings.append(
            await notify_tag_reassignment(reassigned_loop_id, reassigned_loop_tag, reassigned_roles)
        )

    # 导入可能改变 is_linked 订阅集合：导入结束后统一通知实时订阅 Leader 刷新一次
    # （fire-and-forget，不在逐行循环里发，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="loop-import")

    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "errors": errors,
        "warnings": warnings,
    }


async def _import_one_row(
    db: AsyncSession,
    tag_name: str,
    description: str,
    unit_name: str,
    is_active: bool,
    role_tag_values: dict[str, str],
    operator: str,
    plant_node_cache: dict[str, str],
    tag_cache: dict[str, str],
    loop_type: str | None = None,
    control_type: str | None = None,
    importance_level: int | None = None,
    include_in_evaluation: bool | None = None,
    op_output_lower_limit: float | None = None,
    op_output_upper_limit: float | None = None,
    dcs_model_id: str | None = None,
    remark: str | None = None,
) -> tuple[bool, str, list[str]]:
    """处理单行导入。

    Returns:
        (is_update, loop_id, changed_roles)：
        - is_update: True 为更新已存在回路，False 为新建
        - loop_id: 回路 ID
        - changed_roles: 更新场景下各角色 tag 名发生变化的角色列表（新建恒为 []），
          供调用方在提交后做 tag 重关联 warning + 缓存失效

    在调用方的 SAVEPOINT 内执行，异常会触发回滚至 SAVEPOINT。
    """
    # 查找/创建 plant_node
    unit_id: str | None = None
    if unit_name:
        if unit_name in plant_node_cache:
            unit_id = plant_node_cache[unit_name]
        else:
            p_result = await db.execute(select(PlantNode).where(PlantNode.name == unit_name))
            node = p_result.scalar_one_or_none()
            if node is None:
                node = PlantNode(
                    id=str(uuid4()),
                    name=unit_name,
                    type="UNIT",
                )
                db.add(node)
                await db.flush()
            unit_id = str(node.id)
            plant_node_cache[unit_name] = unit_id

    # 查找/创建回路
    result = await db.execute(select(LoopLedger).where(LoopLedger.tag_name == tag_name))
    loop = result.scalar_one_or_none()
    is_update = loop is not None

    if is_update:
        loop.description = description or loop.description
        loop.unit_id = unit_id
        loop.is_active = is_active
        if loop_type is not None:
            loop.loop_type = loop_type
        # v6.1 新增字段：仅在非空/非 None 时更新（空值表示保持原值）
        if control_type is not None:
            loop.control_type = control_type
        if importance_level is not None:
            loop.importance_level = importance_level
        if include_in_evaluation is not None:
            loop.include_in_evaluation = include_in_evaluation
        # OP 限位：显式传入 None 时表示使用默认（清除自定义值）
        # 但导入解析时 None 表示未填，这里区分：若 Excel 单元格为空则不修改
        # （op_output_lower_limit/upper_limit 在解析时 None=空单元格，已区分）
        if op_output_lower_limit is not None:
            loop.op_output_lower_limit = op_output_lower_limit
        if op_output_upper_limit is not None:
            loop.op_output_upper_limit = op_output_upper_limit
        # DCS 型号：导入时总是覆盖（空值=清空，使用默认映射）
        loop.dcs_model_id = dcs_model_id
        if remark is not None:
            loop.remark = remark
        loop.updated_by = operator
        # P2：删除旧映射前记录各角色 tag 名，用于检测 tag 重关联（历史数据孤儿化风险）
        old_role_tag_names = await get_loop_role_tag_names(db, str(loop.id))
        # 删除现有关联 Tag
        await db.execute(delete(LoopTagMapping).where(LoopTagMapping.loop_id == str(loop.id)))
    else:
        old_role_tag_names: dict[str, str] = {}
        loop = LoopLedger(
            id=str(uuid4()),
            tag_name=tag_name,
            description=description or None,
            unit_id=unit_id,
            is_active=is_active,
            status="PARTIAL",
            loop_type=loop_type,
            control_type=control_type,
            # NOT NULL 列缺省兜底：等级默认 2、参评默认 True（对齐模型 default）
            importance_level=importance_level if importance_level is not None else 2,
            include_in_evaluation=(
                include_in_evaluation if include_in_evaluation is not None else True
            ),
            op_output_lower_limit=op_output_lower_limit,
            op_output_upper_limit=op_output_upper_limit,
            dcs_model_id=dcs_model_id,
            remark=remark,
            created_by=operator,
            updated_by=operator,
        )
        db.add(loop)
        await db.flush()

    # 创建 Tag 关联
    new_mappings: dict[str, LoopTagMapping] = {}
    for role, t_name in role_tag_values.items():
        if not t_name:
            continue
        if t_name in tag_cache:
            tag_id = tag_cache[t_name]
        else:
            t_result = await db.execute(select(TagRegistry).where(TagRegistry.tag_name == t_name))
            tag = t_result.scalar_one_or_none()
            if tag is None:
                # P3 #47: 不再静默创建 TagRegistry（绕过 AAS 同步会引入"幽灵 Tag"）
                # 改为显式警告 + 标记 tag_description，提示运维人员后续需通过 AAS 同步补全元数据
                logger.warning(
                    "Excel 导入自动创建 Tag（未通过 AAS 同步）: tag_name=%s, role=%s, "
                    "operator=%s — 该 Tag 缺少量程/单位/measure_type 等元数据，"
                    "请尽快执行 AAS 同步以补全",
                    t_name,
                    role,
                    operator,
                )
                tag = TagRegistry(
                    id=str(uuid4()),
                    tag_name=t_name,
                    tag_description="[Excel 导入自动创建，未通过 AAS 同步，元数据待补全]",
                    tag_type=role,
                    last_sync_at=datetime.now(UTC).replace(tzinfo=None),
                    is_linked=True,
                )
                db.add(tag)
                await db.flush()
            else:
                tag.is_linked = True
            tag_id = str(tag.id)
            tag_cache[t_name] = tag_id

        mapping = LoopTagMapping(
            id=str(uuid4()),
            loop_id=str(loop.id),
            tag_id=tag_id,
            tag_role=role,
            is_required=role in REQUIRED_ROLES,
        )
        db.add(mapping)
        new_mappings[role] = mapping

    # 推导 status
    new_status = await derive_loop_status(db, loop, mappings=new_mappings)
    loop.status = new_status

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="LOOP_IMPORT_UPDATE" if is_update else "LOOP_IMPORT",
        target_type="loop_ledger",
        target_id=str(loop.id),
        after_value=json.dumps(
            {"tagName": tag_name, "status": new_status, "isActive": is_active},
            ensure_ascii=False,
        ),
    )

    await db.flush()

    # P2：对比变更前后各角色 tag 名（新建回路无历史，恒为空）
    new_role_tag_names = {role: t for role, t in role_tag_values.items() if t}
    changed_roles = (
        detect_tag_reassignment(old_role_tag_names, new_role_tag_names) if is_update else []
    )
    return is_update, str(loop.id), changed_roles


__all__ = [
    "ALL_ROLES",
    "REQUIRED_ROLES",
    "ROLE_TO_FIELD",
    "LOOP_TYPE_TO_CN",
    "LOOP_TYPE_FROM_CN",
    "CONTROL_TYPE_TO_CN",
    "CONTROL_TYPE_FROM_CN",
    "TAG_REASSIGN_WARNING",
    "create_loop",
    "delete_loop",
    "derive_loop_status",
    "detect_tag_reassignment",
    "export_loops",
    "get_loop_detail",
    "get_loop_role_tag_names",
    "import_loops",
    "list_loops",
    "notify_tag_reassignment",
    "update_loop",
]
