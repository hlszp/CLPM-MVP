"""Plant node service — CRUD + tree building (IDS v3.2 §2.2.1~2.2.4)."""

from __future__ import annotations

import io
import json
from datetime import UTC, datetime
from uuid import uuid4

import openpyxl
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BizError
from app.models.audit import SysAuditLog
from app.models.loop import LoopLedger
from app.models.plant_node import PlantNode

# 节点类型枚举（FACTORY → AREA → UNIT 三层结构，回路挂 UNIT 下）
VALID_NODE_TYPES = {"AREA", "FACTORY", "UNIT"}


async def _write_audit(
    db: AsyncSession,
    operator: str,
    operation_type: str,
    target_type: str,
    target_id: str,
    before_value: str | None = None,
    after_value: str | None = None,
) -> None:
    """写入审计日志（不抛异常，不影响主流程）。"""
    log = SysAuditLog(
        id=str(uuid4()),
        operator=operator,
        operation_type=operation_type,
        target_type=target_type,
        target_id=target_id,
        before_value=before_value,
        after_value=after_value,
        operated_at=datetime.now(UTC).replace(tzinfo=None),
    )
    db.add(log)


def _node_to_dict(node: PlantNode) -> dict:
    """将 PlantNode ORM 转为字典。"""
    return {
        "id": str(node.id),
        "name": node.name,
        "type": node.type,
        "parentId": str(node.parent_id) if node.parent_id else None,
        "isKpiEnabled": bool(node.is_kpi_enabled) if node.is_kpi_enabled is not None else False,
        "sortOrder": node.sort_order or 0,
    }


async def list_plant_tree(db: AsyncSession, parent_id: str | None = None) -> list[dict]:
    """获取工厂节点树形结构。

    Args:
        db: 异步数据库会话
        parent_id: 父节点 ID。None 表示从顶层开始

    Returns:
        树形结构列表（递归 children，同级按 sort_order → name 排序）
    """
    result = await db.execute(select(PlantNode))
    all_nodes = result.scalars().all()

    # 构建父子映射（同级按 sort_order → name 排序）
    children_map: dict[str | None, list[PlantNode]] = {}
    for node in sorted(all_nodes, key=lambda n: (n.sort_order or 0, n.name)):
        key = str(node.parent_id) if node.parent_id else None
        children_map.setdefault(key, []).append(node)

    def build_tree(pid: str | None) -> list[dict]:
        nodes = children_map.get(pid, [])
        tree: list[dict] = []
        for node in nodes:
            node_dict = _node_to_dict(node)
            node_dict["children"] = build_tree(str(node.id))
            tree.append(node_dict)
        return tree

    # 若指定 parent_id，则返回该节点的子树；否则从顶层开始
    root_pid = parent_id if parent_id else None
    return build_tree(root_pid)


async def create_plant_node(
    db: AsyncSession,
    name: str,
    node_type: str,
    parent_id: str | None,
    operator: str,
) -> dict:
    """创建工厂节点。

    Raises:
        BizError: ERR_VALIDATION (类型非法/FACTORY 必须为顶层) / ERR_NODE_NOT_FOUND (父节点不存在)
    """
    if node_type not in VALID_NODE_TYPES:
        raise BizError(
            code="ERR_VALIDATION",
            message=f"节点类型非法，必须为 {','.join(VALID_NODE_TYPES)}",
            status_code=400,
        )

    # FACTORY 类型必须为顶层节点
    if node_type == "FACTORY" and parent_id:
        raise BizError(
            code="ERR_VALIDATION",
            message="FACTORY 类型节点必须为顶层节点（parentId 为 null）",
            status_code=400,
        )

    # 校验父节点存在
    if parent_id:
        result = await db.execute(select(PlantNode).where(PlantNode.id == parent_id))
        parent = result.scalar_one_or_none()
        if parent is None:
            raise BizError(
                code="ERR_NODE_NOT_FOUND",
                message="父节点不存在",
                status_code=404,
            )

    # 同父重名校验（数据库唯一约束兜底，前置校验给出友好错误）
    dup_stmt = select(PlantNode).where(PlantNode.name == name)
    if parent_id:
        dup_stmt = dup_stmt.where(PlantNode.parent_id == parent_id)
    else:
        dup_stmt = dup_stmt.where(PlantNode.parent_id.is_(None))
    if (await db.execute(dup_stmt)).scalar_one_or_none() is not None:
        raise BizError(
            code="ERR_NODE_NAME_DUPLICATED",
            message=f"同级已存在同名节点「{name}」，节点名称在同一父级下不可重复",
            status_code=409,
        )

    # 创建节点
    node = PlantNode(
        id=str(uuid4()),
        name=name,
        type=node_type,
        parent_id=parent_id,
        updated_by=operator,
    )
    db.add(node)
    await db.flush()

    # 审计日志
    await _write_audit(
        db=db,
        operator=operator,
        operation_type="PLANT_NODE_CREATE",
        target_type="plant_node",
        target_id=str(node.id),
        after_value=f'{{"name":"{name}","type":"{node_type}","parentId":"{parent_id}"}}',
    )
    await db.commit()

    return {
        "id": str(node.id),
        "name": node.name,
        "type": node.type,
        "parentId": str(node.parent_id) if node.parent_id else None,
    }


async def update_plant_node(
    db: AsyncSession,
    node_id: str,
    name: str,
    operator: str,
    is_kpi_enabled: bool | None = None,
    sort_order: int | None = None,
) -> dict:
    """更新工厂节点（名称 + 排序 + 是否纳入性能评估）。

    Raises:
        BizError: ERR_NODE_NOT_FOUND (节点不存在)
    """
    result = await db.execute(select(PlantNode).where(PlantNode.id == node_id))
    node = result.scalar_one_or_none()
    if node is None:
        raise BizError(
            code="ERR_NODE_NOT_FOUND",
            message="节点不存在",
            status_code=404,
        )

    before_value = f'{{"name":"{node.name}","isKpiEnabled":{node.is_kpi_enabled}}}'

    # 改名时校验同父重名（改名场景才需要；数据库唯一约束兜底）
    if name != node.name:
        dup_stmt = select(PlantNode).where(PlantNode.name == name, PlantNode.id != node_id)
        if node.parent_id:
            dup_stmt = dup_stmt.where(PlantNode.parent_id == node.parent_id)
        else:
            dup_stmt = dup_stmt.where(PlantNode.parent_id.is_(None))
        if (await db.execute(dup_stmt)).scalar_one_or_none() is not None:
            raise BizError(
                code="ERR_NODE_NAME_DUPLICATED",
                message=f"同级已存在同名节点「{name}」，节点名称在同一父级下不可重复",
                status_code=409,
            )

    node.name = name
    if is_kpi_enabled is not None:
        node.is_kpi_enabled = is_kpi_enabled
    if sort_order is not None:
        node.sort_order = sort_order
    node.updated_by = operator
    after_value = f'{{"name":"{name}","isKpiEnabled":{node.is_kpi_enabled}}}'

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="PLANT_NODE_UPDATE",
        target_type="plant_node",
        target_id=str(node.id),
        before_value=before_value,
        after_value=after_value,
    )
    await db.commit()

    return _node_to_dict(node)


async def delete_plant_node(
    db: AsyncSession,
    node_id: str,
    operator: str,
) -> dict:
    """删除工厂节点。

    校验：节点存在子节点 → ERR_NODE_HAS_CHILDREN；节点关联回路 → ERR_NODE_HAS_LOOPS。

    Raises:
        BizError: ERR_NODE_NOT_FOUND / ERR_NODE_HAS_CHILDREN / ERR_NODE_HAS_LOOPS
    """
    result = await db.execute(select(PlantNode).where(PlantNode.id == node_id))
    node = result.scalar_one_or_none()
    if node is None:
        raise BizError(
            code="ERR_NODE_NOT_FOUND",
            message="节点不存在",
            status_code=404,
        )

    # 校验子节点
    children_count_result = await db.execute(
        select(func.count()).select_from(PlantNode).where(PlantNode.parent_id == node_id)
    )
    children_count = children_count_result.scalar() or 0
    if children_count > 0:
        raise BizError(
            code="ERR_NODE_HAS_CHILDREN",
            message="该节点存在子节点，无法删除",
            status_code=400,
        )

    # 校验关联回路（loop_ledger.unit_id）
    loops_count_result = await db.execute(
        select(func.count()).select_from(LoopLedger).where(LoopLedger.unit_id == node_id)
    )
    loops_count = loops_count_result.scalar() or 0
    if loops_count > 0:
        raise BizError(
            code="ERR_NODE_HAS_LOOPS",
            message="该节点存在关联回路，无法删除",
            status_code=400,
        )

    before_value = f'{{"name":"{node.name}","type":"{node.type}"}}'
    await db.execute(delete(PlantNode).where(PlantNode.id == node_id))

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="PLANT_NODE_DELETE",
        target_type="plant_node",
        target_id=str(node.id),
        before_value=before_value,
    )
    await db.commit()

    return {"success": True}


# ---------------------------------------------------------------------------
# 批量导入导出
# ---------------------------------------------------------------------------

# Excel 列头（5 列）
EXPORT_HEADERS = ["节点名称", "节点类型", "父节点名称", "是否参评", "层级路径"]


def _cell_str(value: object) -> str:
    """将 Excel 单元格值转为去除首尾空白的字符串，None/空返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


async def export_plant_nodes(db: AsyncSession) -> bytes:
    """导出所有工厂节点为 Excel 文件（.xlsx），返回文件字节。

    列结构（5 列）：节点名称 / 节点类型 / 父节点名称 / 是否参评 / 层级路径。
    节点按层级顺序输出（父节点在子节点之前），便于导入。
    """
    result = await db.execute(select(PlantNode))
    all_nodes = result.scalars().all()

    # id → node 映射
    node_map: dict[str, PlantNode] = {str(n.id): n for n in all_nodes}

    # 构建父子映射，用于层级排序
    children_map: dict[str | None, list[PlantNode]] = {}
    for node in all_nodes:
        key = str(node.parent_id) if node.parent_id else None
        children_map.setdefault(key, []).append(node)

    # 递归构建层级路径（带 memoization）
    path_cache: dict[str, str] = {}

    def build_path(node_id: str) -> str:
        if node_id in path_cache:
            return path_cache[node_id]
        node = node_map.get(node_id)
        if node is None:
            return ""
        if node.parent_id is None:
            path = node.name
        else:
            parent_path = build_path(str(node.parent_id))
            path = f"{parent_path}/{node.name}" if parent_path else node.name
        path_cache[node_id] = path
        return path

    # DFS 遍历，确保父节点在子节点之前
    ordered_nodes: list[PlantNode] = []

    def traverse(pid: str | None) -> None:
        for node in children_map.get(pid, []):
            ordered_nodes.append(node)
            traverse(str(node.id))

    traverse(None)

    # 构建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工厂层级"
    ws.append(EXPORT_HEADERS)

    for node in ordered_nodes:
        parent_name = ""
        if node.parent_id:
            parent = node_map.get(str(node.parent_id))
            if parent:
                parent_name = parent.name
        kpi_str = "是" if node.is_kpi_enabled else "否"
        path = build_path(str(node.id))
        ws.append([node.name, node.type, parent_name, kpi_str, path])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_plant_nodes(
    db: AsyncSession,
    file_bytes: bytes,
    operator: str,
) -> dict:
    """批量导入工厂节点（Excel .xlsx）。

    逐行处理：按 name + parent 查找节点，存在则更新（类型/参评），不存在则新建。
    Excel 列结构（5 列）：节点名称 / 节点类型 / 父节点名称 / 是否参评 / 层级路径。
    「层级路径」仅展示用，导入时忽略——层级关系由「父节点名称」推导，路径由系统自动生成。
    「是否参评」空值时不修改现有值；「是/否」显式设置。
    返回 {total, inserted, updated, failed, errors[]}。
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(
            code="ERR_FILE_PARSE",
            message=f"Excel 文件解析失败: {exc}",
            status_code=400,
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    total = 0
    inserted = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    # 缓存：parent_name → parent_id，避免重复查询
    parent_cache: dict[str, str] = {}

    for row_idx, row in enumerate(rows, start=2):  # 第 1 行为表头
        total += 1
        name = _cell_str(row[0]) if len(row) > 0 else ""

        if not name:
            errors.append({"row": row_idx, "message": "节点名称不能为空"})
            failed += 1
            continue

        node_type = _cell_str(row[1]) if len(row) > 1 else ""
        parent_name = _cell_str(row[2]) if len(row) > 2 else ""
        kpi_raw = _cell_str(row[3]) if len(row) > 3 else ""
        # 第 5 列（层级路径）仅用于展示，导入时忽略——路径由系统按父节点自动生成
        # 参评列解析：空=不修改；「是/否/true/false/1/0」显式设置
        is_kpi_enabled: bool | None = None
        if kpi_raw:
            if kpi_raw in ("是", "true", "True", "TRUE", "1"):
                is_kpi_enabled = True
            elif kpi_raw in ("否", "false", "False", "FALSE", "0"):
                is_kpi_enabled = False
            else:
                errors.append(
                    {
                        "row": row_idx,
                        "name": name,
                        "message": f"是否参评列值非法（{kpi_raw}），仅支持 是/否",
                    }
                )
                failed += 1
                continue

        # 节点类型校验
        if node_type not in VALID_NODE_TYPES:
            errors.append(
                {
                    "row": row_idx,
                    "name": name,
                    "message": f"节点类型非法，必须为 {','.join(VALID_NODE_TYPES)}",
                }
            )
            failed += 1
            continue

        try:
            # 使用 SAVEPOINT 保证单行失败不影响其他行
            async with db.begin_nested():
                is_update = await _import_one_node(
                    db=db,
                    name=name,
                    node_type=node_type,
                    parent_name=parent_name,
                    operator=operator,
                    parent_cache=parent_cache,
                    is_kpi_enabled=is_kpi_enabled,
                )
        except Exception as exc:  # noqa: BLE001
            failed += 1
            errors.append({"row": row_idx, "name": name, "message": str(exc)})
            continue

        if is_update:
            updated += 1
        else:
            inserted += 1

    await db.commit()

    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "errors": errors,
    }


async def _import_one_node(
    db: AsyncSession,
    name: str,
    node_type: str,
    parent_name: str,
    operator: str,
    parent_cache: dict[str, str],
    is_kpi_enabled: bool | None = None,
) -> bool:
    """处理单行导入，返回是否为更新（True）或新建（False）。

    在调用方的 SAVEPOINT 内执行，异常会触发回滚至 SAVEPOINT。
    is_kpi_enabled=None 时不修改现有值（新建则默认 False）。
    """
    # 查找父节点
    parent_id: str | None = None
    if parent_name:
        if parent_name in parent_cache:
            parent_id = parent_cache[parent_name]
        else:
            p_result = await db.execute(select(PlantNode).where(PlantNode.name == parent_name))
            parent = p_result.scalars().first()
            if parent is None:
                raise ValueError(f"父节点 '{parent_name}' 不存在")
            parent_id = str(parent.id)
            parent_cache[parent_name] = parent_id

    # FACTORY 类型必须为顶层节点
    if node_type == "FACTORY" and parent_id:
        raise ValueError("FACTORY 类型节点必须为顶层节点（无父节点）")

    # 按 name + parent 查找节点是否已存在
    if parent_id:
        result = await db.execute(
            select(PlantNode).where(
                PlantNode.name == name,
                PlantNode.parent_id == parent_id,
            )
        )
    else:
        result = await db.execute(
            select(PlantNode).where(
                PlantNode.name == name,
                PlantNode.parent_id.is_(None),
            )
        )
    node = result.scalars().first()
    is_update = node is not None

    if is_update:
        before_value = json.dumps(
            {
                "name": node.name,
                "type": node.type,
                "parentId": str(node.parent_id) if node.parent_id else None,
                "isKpiEnabled": node.is_kpi_enabled,
            },
            ensure_ascii=False,
        )
        node.type = node_type
        if is_kpi_enabled is not None:
            node.is_kpi_enabled = is_kpi_enabled
        node.updated_by = f"import:{operator}"
        after_value = json.dumps(
            {
                "name": name,
                "type": node_type,
                "parentId": parent_id,
                "isKpiEnabled": node.is_kpi_enabled,
            },
            ensure_ascii=False,
        )
        await _write_audit(
            db=db,
            operator=operator,
            operation_type="PLANT_NODE_IMPORT_UPDATE",
            target_type="plant_node",
            target_id=str(node.id),
            before_value=before_value,
            after_value=after_value,
        )
    else:
        node = PlantNode(
            id=str(uuid4()),
            name=name,
            type=node_type,
            parent_id=parent_id,
            is_kpi_enabled=is_kpi_enabled if is_kpi_enabled is not None else False,
            updated_by=f"import:{operator}",
        )
        db.add(node)
        await db.flush()
        # 缓存新建节点，供后续行作为父节点查找
        parent_cache[name] = str(node.id)
        await _write_audit(
            db=db,
            operator=operator,
            operation_type="PLANT_NODE_IMPORT",
            target_type="plant_node",
            target_id=str(node.id),
            after_value=json.dumps(
                {
                    "name": name,
                    "type": node_type,
                    "parentId": parent_id,
                    "isKpiEnabled": node.is_kpi_enabled,
                },
                ensure_ascii=False,
            ),
        )

    await db.flush()
    return is_update


__all__ = [
    "create_plant_node",
    "delete_plant_node",
    "export_plant_nodes",
    "import_plant_nodes",
    "list_plant_tree",
    "update_plant_node",
]
