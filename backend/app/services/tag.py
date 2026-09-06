"""Tag registry service — 测点清单 CRUD + 导入导出.

测点与工厂节点的关联通过 LoopTagMapping → LoopLedger.unit_id 间接关联，
TagRegistry 本身没有 unit_id 字段。
"""

from __future__ import annotations

import io
import json
from datetime import UTC, datetime
from typing import Any
from uuid import uuid4

import openpyxl
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BizError
from app.core.numeric import parse_finite_float
from app.models.audit import SysAuditLog
from app.models.loop import LoopLedger, LoopTagMapping
from app.models.plant_node import PlantNode
from app.models.tag import TagRegistry
from app.services.dict_item import (
    DICT_MEASURE_TYPE,
    DICT_TAG_TYPE,
    dict_items_hint,
    get_dict_items,
    normalize_by_dict,
)

# 测点类型枚举（出厂默认；运行时以 sys_dict_item 字典为准，见 dict_item service）
MEASURE_TYPES = ("TEMPERATURE", "PRESSURE", "LEVEL", "FLOW", "ANALYSIS", "SPEED", "OTHER")
# 参数类型枚举
TAG_TYPES = ("PV", "SP", "OP", "MODE", "PID_P", "PID_I", "PID_D", "OTHER")

# Excel 导出列头（10 列）
EXPORT_HEADERS = [
    "位号",
    "名称",
    "测点类型",
    "量程下限",
    "量程上限",
    "单位",
    "参数类型",
    "所属单元",
    "原始ID",
    "是否启用",
]


async def _write_audit(
    db: AsyncSession,
    operator: str,
    operation_type: str,
    target_type: str,
    target_id: str | None,
    before_value: str | None = None,
    after_value: str | None = None,
) -> None:
    """写入审计日志。

    target_id 为 None 或空字符串时写入 NULL（批量操作无单一目标）。
    """
    # 空字符串 → None（PostgreSQL UUID 类型不接受空字符串）
    if target_id is not None and not str(target_id).strip():
        target_id = None
    log = SysAuditLog(
        id=str(uuid4()),
        operator=operator,
        operation_type=operation_type,
        target_type=target_type,
        target_id=target_id,
        before_value=before_value,
        after_value=after_value,
        operated_at=datetime.now(UTC).replace(tzinfo=None),
    )
    db.add(log)


async def _get_descendant_node_ids(db: AsyncSession, parent_id: str) -> list[str]:
    """递归获取所有子孙节点 ID。"""
    result = await db.execute(select(PlantNode.id).where(PlantNode.parent_id == parent_id))
    child_ids = [str(row[0]) for row in result]
    all_ids = list(child_ids)
    for child_id in child_ids:
        all_ids.extend(await _get_descendant_node_ids(db, child_id))
    return all_ids


def _cell_str(value: object) -> str:
    """将 Excel 单元格值转为去除首尾空白的字符串，None/空返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


def normalize_measure_type(value: str) -> str | None:
    """测点类型归一化（同步版本，仅兜底场景使用）。

    运行时校验请用 ``await normalize_by_dict(db, DICT_MEASURE_TYPE, value)``
    （以 sys_dict_item 字典为准，支持用户自定义类型）。
    本函数仅按出厂默认枚举 + 中文别名归一，用于字典服务不可用的兜底。
    """
    v = value.strip()
    if not v:
        return None
    upper = v.upper()
    if upper in MEASURE_TYPES:
        return upper
    return None


def _cell_float(value: object) -> float | None:
    """将 Excel 单元格值转为 float，空值或无法解析返回 None。"""
    s = _cell_str(value)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _build_tag_dict(
    tag: TagRegistry,
    loop_info: dict | None = None,
    realtime_cache: dict | None = None,
) -> dict:
    """构建测点响应字典。

    Args:
        tag: 数据库 TagRegistry 记录
        loop_info: 关联回路信息
        realtime_cache: Redis 实时缓存（{tagCode: {value, quality, collectTime}}），
            优先于数据库 current_value
    """
    cached = realtime_cache or {}
    rt = cached.get(tag.tag_name)

    if rt:
        raw_val = rt.get("value")
        # R06（数据链路整改）：共享数值契约（app/core/numeric.py）——
        # 新值无效（"-1.#QNAN0"/"nan"/"Infinity"/"1e999"/空值等）时返回
        # currentValue=null + stale=true，不得把 DB 旧值与新 quality/collectTime
        # 拼接成"最新有效读数"；quality 仍按本条消息更新（数值有效性与质量独立）。
        parsed_val = parse_finite_float(raw_val)
        if parsed_val is not None:
            current_value: float | int | str | None = parsed_val
            stale = False
        else:
            current_value = None
            stale = True
        raw_quality = rt.get("quality")
        if isinstance(raw_quality, int | float):
            quality = "GOOD" if int(raw_quality) in (1, 2, 3, 192) else "BAD"
        elif isinstance(raw_quality, str) and raw_quality.upper() in ("GOOD", "BAD", "UNCERTAIN"):
            quality = raw_quality.upper()
        else:
            quality = tag.quality
        last_sync_at = rt.get("collectTime") or (
            tag.last_sync_at.isoformat() if tag.last_sync_at else None
        )
    else:
        current_value = tag.current_value
        quality = tag.quality
        stale = False
        last_sync_at = tag.last_sync_at.isoformat() if tag.last_sync_at else None

    return {
        "id": str(tag.id),
        "tagName": tag.tag_name,
        "tagDescription": tag.tag_description,
        "tagType": tag.tag_type,
        "currentValue": current_value,
        "quality": quality,
        # R06 增量字段：实时新值无效时置 true（旧客户端可忽略）
        "stale": stale,
        "lastSyncAt": last_sync_at,
        "isLinked": bool(tag.is_linked) if tag.is_linked is not None else False,
        "rangeMin": tag.range_min,
        "rangeMax": tag.range_max,
        "unit": tag.unit,
        "measureType": tag.measure_type,
        "tdengineTagId": tag.tdengine_tag_id,
        "loop": loop_info,
        # WS-C 7-8：所属单元名称从关联回路派生（_get_tags_loop_info_map 已带出；
        # 一个 tag 可能被多个回路映射，取第一个映射回路的单元）
        "unitName": (loop_info or {}).get("unitName"),
        # 数据健康度（方案 C 轻量版）：实时质量码 + 同步新鲜度 + 所属回路 PV 完整度
        # 不在列表页对每个 tag 实时查 TDengine（开销过大），改为复用回路级每日巡检快照
        "dataHealth": {
            # 实时质量码（GOOD/BAD/UNCERTAIN，来自 Redis 缓存或 DB）
            "quality": quality,
            # 同步新鲜度（最近一次落库时间，naive ISO 串）
            "lastSyncAt": last_sync_at,
            # 所属回路 PV 完整度（来自每日 02:00 巡检快照，无关联回路则 None）
            "loopPvCompleteness": (loop_info or {}).get("pvCompleteness"),
            "loopIntegrityStatus": (loop_info or {}).get("integrityStatus"),
            "lastIntegrityCheck": (loop_info or {}).get("lastIntegrityCheck"),
        },
    }


async def _get_tags_loop_info_map(db: AsyncSession, tag_ids: list[str]) -> dict[str, dict]:
    """批量获取测点关联的回路信息（含所属单元名称）。

    通过 loop_tag_mapping → loop_ledger → plant_node 间接关联。
    一个测点可能关联多个回路，取第一个。
    """
    if not tag_ids:
        return {}
    result = await db.execute(
        select(LoopTagMapping, LoopLedger, PlantNode)
        .join(LoopLedger, LoopTagMapping.loop_id == LoopLedger.id)
        .outerjoin(PlantNode, LoopLedger.unit_id == PlantNode.id)
        .where(LoopTagMapping.tag_id.in_(tag_ids))
    )
    loop_map: dict[str, dict] = {}
    for mapping, loop, plant_node in result:
        tag_id = str(mapping.tag_id)
        if tag_id not in loop_map:
            loop_map[tag_id] = {
                "loopId": str(loop.id),
                "loopTagName": loop.tag_name,
                "loopDescription": loop.description,
                "unitName": plant_node.name if plant_node else None,
            }
    return loop_map


async def list_tags(
    db: AsyncSession,
    keyword: str | None = None,
    measure_type: str | None = None,
    tag_type: str | None = None,
    plant_node_id: str | None = None,
    is_linked: bool | None = None,
    page: int = 1,
    page_size: int = 20,
) -> dict:
    """分页查询测点列表。

    plantNodeId 通过 JOIN loop_tag_mapping + loop_ledger 间接筛选。
    """
    conditions = []

    if keyword:
        kw = f"%{keyword}%"
        conditions.append(TagRegistry.tag_name.ilike(kw))
    if measure_type:
        conditions.append(func.upper(TagRegistry.measure_type) == measure_type.upper())
    if tag_type:
        conditions.append(func.upper(TagRegistry.tag_type) == tag_type.upper())
    if is_linked is not None:
        conditions.append(TagRegistry.is_linked.is_(is_linked))

    # plantNodeId：通过 loop_tag_mapping → loop_ledger.unit_id 层级查询
    if plant_node_id:
        descendant_ids = await _get_descendant_node_ids(db, plant_node_id)
        all_node_ids = descendant_ids + [plant_node_id]
        tag_id_subquery = (
            select(LoopTagMapping.tag_id)
            .join(LoopLedger, LoopTagMapping.loop_id == LoopLedger.id)
            .where(LoopLedger.unit_id.in_(all_node_ids))
            .distinct()
        )
        conditions.append(TagRegistry.id.in_(tag_id_subquery))

    # 计数
    count_stmt = select(func.count()).select_from(TagRegistry)
    for cond in conditions:
        count_stmt = count_stmt.where(cond)
    total_result = await db.execute(count_stmt)
    total = total_result.scalar() or 0

    # 查询
    stmt = select(TagRegistry).order_by(TagRegistry.tag_name)
    for cond in conditions:
        stmt = stmt.where(cond)
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(stmt)
    tags = result.scalars().all()

    # 批量查询回路信息
    tag_ids = [str(t.id) for t in tags]
    loop_map = await _get_tags_loop_info_map(db, tag_ids)

    # 批量查询关联回路的最新数据完整性巡检快照（用于测点页展示回路级 PV 完整度）
    # 一个 tag 关联一个回路，取该回路最新一条 check_date 的快照
    linked_loop_ids = {info["loopId"] for info in loop_map.values() if info.get("loopId")}
    integrity_map: dict[str, Any] = {}
    if linked_loop_ids:
        from app.models.metric import LoopIntegritySnapshot

        i_stmt = (
            select(LoopIntegritySnapshot)
            .where(LoopIntegritySnapshot.loop_id.in_(list(linked_loop_ids)))
            .distinct(LoopIntegritySnapshot.loop_id)
            .order_by(LoopIntegritySnapshot.loop_id, LoopIntegritySnapshot.check_date.desc())
        )
        i_result = await db.execute(i_stmt)
        for snap in i_result.scalars().all():
            integrity_map[str(snap.loop_id)] = snap

    # 将回路级完整度快照合并进 loop_info，供 _build_tag_dict 读取
    for info in loop_map.values():
        lid = info.get("loopId")
        snap = integrity_map.get(lid) if lid else None
        if snap:
            info["pvCompleteness"] = snap.pv_completeness
            info["integrityStatus"] = snap.status
            info["lastIntegrityCheck"] = snap.check_date.isoformat() if snap.check_date else None

    # 批量从 Redis 读取实时值（优先于数据库 current_value）
    realtime_cache: dict[str, dict] = {}
    if tags:
        try:
            from app.services.data_source.realtime_subscriber import get_subscriber

            subscriber = get_subscriber()
            tag_names = [t.tag_name for t in tags]
            cached_list = await subscriber.get_cached_values(tag_names)
            for item in cached_list:
                tc = item.get("tagCode")
                if tc:
                    realtime_cache[tc] = item
        except Exception:
            import logging

            logging.getLogger(__name__).warning(
                "从 Redis 读取实时值失败，回退到数据库值", exc_info=True
            )

    items = [_build_tag_dict(tag, loop_map.get(str(tag.id)), realtime_cache) for tag in tags]

    return {
        "items": items,
        "total": total,
        "page": page,
        "pageSize": page_size,
    }


async def get_tag_detail(db: AsyncSession, tag_id: str) -> dict:
    """获取测点详情（含关联回路信息）。

    Raises:
        BizError: ERR_TAG_NOT_FOUND
    """
    result = await db.execute(select(TagRegistry).where(TagRegistry.id == tag_id))
    tag = result.scalar_one_or_none()
    if tag is None:
        raise BizError(
            code="ERR_TAG_NOT_FOUND",
            message="测点不存在",
            status_code=404,
        )

    loop_map = await _get_tags_loop_info_map(db, [tag_id])

    # 从 Redis 读取实时值（优先于数据库 current_value）
    realtime_cache: dict[str, dict] = {}
    try:
        from app.services.data_source.realtime_subscriber import get_subscriber

        subscriber = get_subscriber()
        cached_list = await subscriber.get_cached_values([tag.tag_name])
        for item in cached_list:
            tc = item.get("tagCode")
            if tc:
                realtime_cache[tc] = item
    except Exception:
        import logging

        logging.getLogger(__name__).warning(
            "从 Redis 读取实时值失败，回退到数据库值", exc_info=True
        )

    return _build_tag_dict(tag, loop_map.get(tag_id), realtime_cache)


async def create_tag(
    db: AsyncSession,
    operator: str,
    tag_name: str,
    tag_description: str | None = None,
    range_min: float | None = None,
    range_max: float | None = None,
    unit: str | None = None,
    measure_type: str | None = None,
    tag_type: str | None = None,
    tdengine_tag_id: str | None = None,
) -> dict:
    """新建测点。

    校验：位号唯一（重复返回 ERR_TAG_ALREADY_EXISTS）、枚举合法。
    is_linked 恒为 False（仅由回路映射派生，与导入路径口径一致）。

    Raises:
        BizError: ERR_TAG_ALREADY_EXISTS / ERR_MEASURE_TYPE_INVALID / ERR_TAG_TYPE_INVALID
    """
    # 位号唯一性校验
    dup = await db.execute(select(TagRegistry).where(TagRegistry.tag_name == tag_name))
    if dup.scalar_one_or_none() is not None:
        raise BizError(
            code="ERR_TAG_ALREADY_EXISTS",
            message=f"位号已存在: {tag_name}",
            status_code=400,
        )

    # 枚举校验（以字典为准，支持自定义类型；中文别名自动归一为 code）
    if measure_type is not None:
        normalized = await normalize_by_dict(db, DICT_MEASURE_TYPE, measure_type)
        if normalized is None:
            raise BizError(
                code="ERR_MEASURE_TYPE_INVALID",
                message=(
                    f"测点类型无效: {measure_type}，"
                    f"支持的类型：{await dict_items_hint(db, DICT_MEASURE_TYPE)}"
                ),
                status_code=400,
            )
        measure_type = normalized
    if tag_type is not None:
        normalized_tt = await normalize_by_dict(db, DICT_TAG_TYPE, tag_type)
        if normalized_tt is None:
            raise BizError(
                code="ERR_TAG_TYPE_INVALID",
                message=(
                    f"参数类型无效: {tag_type}，"
                    f"支持的类型：{await dict_items_hint(db, DICT_TAG_TYPE)}"
                ),
                status_code=400,
            )
        tag_type = normalized_tt

    tag = TagRegistry(
        id=str(uuid4()),
        tag_name=tag_name,
        tag_description=tag_description or None,
        # tag_type 已归一为字典 code；缺省 OTHER
        tag_type=tag_type or "OTHER",
        # measure_type 已由 normalize_by_dict 归一为字典 code（保留原始大小写）
        measure_type=measure_type,
        range_min=range_min,
        range_max=range_max,
        unit=unit or None,
        tdengine_tag_id=tdengine_tag_id or None,
        is_linked=False,  # 仅由回路映射派生
        last_sync_at=datetime.now(UTC).replace(tzinfo=None),
    )
    db.add(tag)

    after_json = json.dumps(
        {
            "tagName": tag.tag_name,
            "tagDescription": tag.tag_description,
            "tagType": tag.tag_type,
            "measureType": tag.measure_type,
            "rangeMin": tag.range_min,
            "rangeMax": tag.range_max,
            "unit": tag.unit,
            "tdengineTagId": tag.tdengine_tag_id,
        },
        ensure_ascii=False,
        default=str,
    )

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="TAG_CREATE",
        target_type="tag_registry",
        target_id=str(tag.id),
        after_value=after_json,
    )
    await db.commit()

    # 通知实时订阅 Leader 刷新订阅集合（fire-and-forget，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="tag-create")

    return _build_tag_dict(tag)


async def update_tag(
    db: AsyncSession,
    tag_id: str,
    operator: str,
    tag_description: str | None = None,
    range_min: float | None = None,
    range_max: float | None = None,
    unit: str | None = None,
    measure_type: str | None = None,
    tag_type: str | None = None,
    tdengine_tag_id: str | None = None,
) -> dict:
    """更新测点（描述/量程/单位/测点类型/参数类型/TDengine tag ID）。

    Raises:
        BizError: ERR_TAG_NOT_FOUND / ERR_TAG_TYPE_INVALID
    """
    result = await db.execute(select(TagRegistry).where(TagRegistry.id == tag_id))
    tag = result.scalar_one_or_none()
    if tag is None:
        raise BizError(
            code="ERR_TAG_NOT_FOUND",
            message="测点不存在",
            status_code=404,
        )

    # WS-C 7-7：tag_type 合法值校验（以字典为准，支持中文别名与自定义类型）
    if tag_type is not None:
        normalized_tt = await normalize_by_dict(db, DICT_TAG_TYPE, tag_type)
        if normalized_tt is None:
            raise BizError(
                code="ERR_TAG_TYPE_INVALID",
                message=(
                    f"参数类型无效: {tag_type}，"
                    f"支持的类型：{await dict_items_hint(db, DICT_TAG_TYPE)}"
                ),
                status_code=400,
            )
        tag_type = normalized_tt

    # measure_type 归一化校验（以字典为准，支持中文别名与自定义类型）
    if measure_type is not None:
        normalized = await normalize_by_dict(db, DICT_MEASURE_TYPE, measure_type)
        if normalized is None:
            raise BizError(
                code="ERR_MEASURE_TYPE_INVALID",
                message=(
                    f"测点类型无效: {measure_type}，"
                    f"支持的类型：{await dict_items_hint(db, DICT_MEASURE_TYPE)}"
                ),
                status_code=400,
            )
        measure_type = normalized

    before = {
        "tagDescription": tag.tag_description,
        "rangeMin": tag.range_min,
        "rangeMax": tag.range_max,
        "unit": tag.unit,
        "measureType": tag.measure_type,
        "tagType": tag.tag_type,
        "tdengineTagId": tag.tdengine_tag_id,
    }
    before_json = json.dumps(before, ensure_ascii=False, default=str)

    if tag_description is not None:
        tag.tag_description = tag_description
    if range_min is not None:
        tag.range_min = range_min
    if range_max is not None:
        tag.range_max = range_max
    if unit is not None:
        tag.unit = unit
    if measure_type is not None:
        tag.measure_type = measure_type
    if tag_type is not None:
        # 已归一为字典 code（保留原始大小写）
        tag.tag_type = tag_type
    if tdengine_tag_id is not None:
        tag.tdengine_tag_id = tdengine_tag_id

    after = {
        "tagDescription": tag.tag_description,
        "rangeMin": tag.range_min,
        "rangeMax": tag.range_max,
        "unit": tag.unit,
        "measureType": tag.measure_type,
        "tagType": tag.tag_type,
        "tdengineTagId": tag.tdengine_tag_id,
    }
    after_json = json.dumps(after, ensure_ascii=False, default=str)

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="TAG_UPDATE",
        target_type="tag_registry",
        target_id=str(tag.id),
        before_value=before_json,
        after_value=after_json,
    )
    await db.commit()

    # 通知实时订阅 Leader 刷新订阅集合（fire-and-forget，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="tag-update")

    loop_map = await _get_tags_loop_info_map(db, [str(tag.id)])
    return _build_tag_dict(tag, loop_map.get(str(tag.id)))


async def delete_tag(db: AsyncSession, tag_id: str, operator: str) -> dict:
    """删除测点。

    校验：已关联的测点不能删除（返回 ERR_TAG_LINKED）。

    Raises:
        BizError: ERR_TAG_NOT_FOUND / ERR_TAG_LINKED
    """
    result = await db.execute(select(TagRegistry).where(TagRegistry.id == tag_id))
    tag = result.scalar_one_or_none()
    if tag is None:
        raise BizError(
            code="ERR_TAG_NOT_FOUND",
            message="测点不存在",
            status_code=404,
        )

    # 校验是否已关联回路
    link_count_result = await db.execute(
        select(func.count()).select_from(LoopTagMapping).where(LoopTagMapping.tag_id == tag_id)
    )
    link_count = link_count_result.scalar() or 0
    if link_count > 0:
        raise BizError(
            code="ERR_TAG_LINKED",
            message=f"测点已关联 {link_count} 个回路，无法删除",
            status_code=400,
        )

    before_json = json.dumps({"tagName": tag.tag_name, "tagType": tag.tag_type}, ensure_ascii=False)

    await db.execute(delete(TagRegistry).where(TagRegistry.id == tag_id))

    await _write_audit(
        db=db,
        operator=operator,
        operation_type="TAG_DELETE",
        target_type="tag_registry",
        target_id=tag_id,
        before_value=before_json,
    )
    await db.commit()

    # 通知实时订阅 Leader 刷新订阅集合（fire-and-forget，免重启生效）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="tag-delete")

    return {
        "id": tag_id,
        "deleted": True,
        "deletedAt": datetime.now(UTC).replace(tzinfo=None).isoformat(),
    }


async def batch_delete_tags(db: AsyncSession, tag_ids: list[str], operator: str) -> dict:
    """批量删除测点。

    已关联回路的测点跳过并记入 failures，不影响其他测点删除。

    Returns:
        {"deleted": int, "failed": int, "failures": [{tagId, tagName, reason}]}
    """
    deleted_count = 0
    failures: list[dict] = []

    # 批量查询这些 tag 的关联状态
    link_count_result = await db.execute(
        select(LoopTagMapping.tag_id, func.count().label("cnt"))
        .where(LoopTagMapping.tag_id.in_(tag_ids))
        .group_by(LoopTagMapping.tag_id)
    )
    linked_map = {str(row.tag_id): row.cnt for row in link_count_result.fetchall()}

    # 批量查询 tag 信息（用于审计日志和 failure 信息）
    tag_result = await db.execute(select(TagRegistry).where(TagRegistry.id.in_(tag_ids)))
    tags_map = {str(t.id): t for t in tag_result.scalars().fetchall()}

    for tag_id in tag_ids:
        tag = tags_map.get(tag_id)
        if tag is None:
            failures.append({"tagId": tag_id, "tagName": None, "reason": "测点不存在"})
            continue

        link_count = linked_map.get(tag_id, 0)
        if link_count > 0:
            failures.append(
                {
                    "tagId": tag_id,
                    "tagName": tag.tag_name,
                    "reason": f"已关联 {link_count} 个回路，无法删除",
                }
            )
            continue

        before_json = json.dumps(
            {"tagName": tag.tag_name, "tagType": tag.tag_type}, ensure_ascii=False
        )
        await db.execute(delete(TagRegistry).where(TagRegistry.id == tag_id))
        await _write_audit(
            db=db,
            operator=operator,
            operation_type="TAG_DELETE",
            target_type="tag_registry",
            target_id=tag_id,
            before_value=before_json,
        )
        deleted_count += 1

    await db.commit()

    # 批量删除只发一次：通知实时订阅 Leader 刷新订阅集合（fire-and-forget）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="tag-batch-delete")

    return {
        "deleted": deleted_count,
        "failed": len(failures),
        "failures": failures,
    }


# ---------------------------------------------------------------------------
# 批量导入导出
# ---------------------------------------------------------------------------


async def export_tags(
    db: AsyncSession,
    keyword: str | None = None,
    measure_type: str | None = None,
    tag_type: str | None = None,
    plant_node_id: str | None = None,
    is_linked: bool | None = None,
) -> bytes:
    """导出测点为 Excel 文件（.xlsx），返回文件字节。

    支持同列表查询的筛选参数。
    """
    conditions = []

    if keyword:
        kw = f"%{keyword}%"
        conditions.append(TagRegistry.tag_name.ilike(kw))
    if measure_type:
        conditions.append(func.upper(TagRegistry.measure_type) == measure_type.upper())
    if tag_type:
        conditions.append(func.upper(TagRegistry.tag_type) == tag_type.upper())
    if is_linked is not None:
        conditions.append(TagRegistry.is_linked.is_(is_linked))

    if plant_node_id:
        descendant_ids = await _get_descendant_node_ids(db, plant_node_id)
        all_node_ids = descendant_ids + [plant_node_id]
        tag_id_subquery = (
            select(LoopTagMapping.tag_id)
            .join(LoopLedger, LoopTagMapping.loop_id == LoopLedger.id)
            .where(LoopLedger.unit_id.in_(all_node_ids))
            .distinct()
        )
        conditions.append(TagRegistry.id.in_(tag_id_subquery))

    stmt = select(TagRegistry).order_by(TagRegistry.tag_name)
    for cond in conditions:
        stmt = stmt.where(cond)
    result = await db.execute(stmt)
    tags = result.scalars().all()

    tag_ids = [str(t.id) for t in tags]
    loop_map = await _get_tags_loop_info_map(db, tag_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测点清单"
    ws.append(EXPORT_HEADERS)

    # code → label（导出列与前端展示口径一致；导入支持 code/label 双向识别）
    # enabled_only=False：禁用类型的历史数据也要能显示 label
    measure_type_cn = dict(await get_dict_items(db, DICT_MEASURE_TYPE, enabled_only=False))
    tag_type_cn = dict(await get_dict_items(db, DICT_TAG_TYPE, enabled_only=False))

    for tag in tags:
        loop_info = loop_map.get(str(tag.id))
        unit_name = loop_info.get("unitName") if loop_info else ""
        is_linked_str = "是" if tag.is_linked else "否"
        ws.append(
            [
                tag.tag_name,
                tag.tag_description or "",
                measure_type_cn.get(tag.measure_type or "", tag.measure_type or ""),
                tag.range_min if tag.range_min is not None else "",
                tag.range_max if tag.range_max is not None else "",
                tag.unit or "",
                tag_type_cn.get(tag.tag_type or "", tag.tag_type or ""),
                unit_name or "",
                tag.tdengine_tag_id or "",
                is_linked_str,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_tags(
    db: AsyncSession,
    file_bytes: bytes,
    operator: str,
) -> dict:
    """批量导入测点（Excel .xlsx）。

    逐行处理：位号已存在则更新，否则新建。
    返回 {total, inserted, updated, failed, errors[]}。

    Excel 列结构（10 列）：
        0. 位号（tag_name）
        1. 名称（tag_description）
        2. 测点类型（measure_type）
        3. 量程下限（range_min）
        4. 量程上限（range_max）
        5. 单位（unit）
        6. 参数类型（tag_type）
        7. 所属单元（plant_node name，按名称查找）
        8. 原始ID（tdengine_tag_id）
        9. 是否启用（WS-C 7-10：本列忽略不导入；is_linked 仅由回路映射派生）
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(
            code="ERR_FILE_PARSE",
            message=f"Excel 文件解析失败: {exc}",
            status_code=400,
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    total = 0
    inserted = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    # 缓存：plant_node name → id（用于按名称查找）
    plant_node_cache: dict[str, str] = {}

    for row_idx, row in enumerate(rows, start=2):  # 第 1 行为表头
        total += 1
        tag_name = _cell_str(row[0]) if len(row) > 0 else ""

        if not tag_name:
            errors.append({"row": row_idx, "message": "位号不能为空"})
            failed += 1
            continue

        # 解析单元格
        tag_description = _cell_str(row[1]) if len(row) > 1 else ""
        measure_type = _cell_str(row[2]) if len(row) > 2 else ""
        range_min = _cell_float(row[3]) if len(row) > 3 else None
        range_max = _cell_float(row[4]) if len(row) > 4 else None
        unit = _cell_str(row[5]) if len(row) > 5 else ""
        tag_type = _cell_str(row[6]) if len(row) > 6 else ""
        plant_node_name = _cell_str(row[7]) if len(row) > 7 else ""
        tdengine_tag_id = _cell_str(row[8]) if len(row) > 8 else ""
        # WS-C 7-10：row[9]「是否启用」列忽略不导入（is_linked 仅由回路映射派生）

        # 校验 measure_type（以字典为准：支持中文别名与自定义类型，温度→TEMPERATURE）
        if measure_type:
            normalized_mt = await normalize_by_dict(db, DICT_MEASURE_TYPE, measure_type)
            if normalized_mt is None:
                errors.append(
                    {
                        "row": row_idx,
                        "tagName": tag_name,
                        "message": (
                            f"测点类型无效: {measure_type}，"
                            f"支持的类型：{await dict_items_hint(db, DICT_MEASURE_TYPE)}"
                        ),
                    }
                )
                failed += 1
                continue
            measure_type = normalized_mt

        # 校验 tag_type（以字典为准：支持中文别名与自定义类型，测量值→PV）
        if tag_type:
            normalized_tt = await normalize_by_dict(db, DICT_TAG_TYPE, tag_type)
            if normalized_tt is None:
                errors.append(
                    {
                        "row": row_idx,
                        "tagName": tag_name,
                        "message": (
                            f"参数类型无效: {tag_type}，"
                            f"支持的类型：{await dict_items_hint(db, DICT_TAG_TYPE)}"
                        ),
                    }
                )
                failed += 1
                continue
            tag_type = normalized_tt

        # 按名称查找所属单元（仅查找，不存储到 TagRegistry）
        if plant_node_name and plant_node_name not in plant_node_cache:
            p_result = await db.execute(select(PlantNode).where(PlantNode.name == plant_node_name))
            node = p_result.scalar_one_or_none()
            if node is not None:
                plant_node_cache[plant_node_name] = str(node.id)

        is_update = False
        try:
            async with db.begin_nested():
                is_update = await _import_one_row(
                    db=db,
                    tag_name=tag_name,
                    tag_description=tag_description,
                    # measure_type 已归一为字典 code（保留原始大小写）
                    measure_type=measure_type or None,
                    range_min=range_min,
                    range_max=range_max,
                    unit=unit or None,
                    # tag_type 已归一为字典 code（保留原始大小写）
                    tag_type=tag_type or None,
                    tdengine_tag_id=tdengine_tag_id or None,
                    operator=operator,
                )
        except Exception as exc:  # noqa: BLE001
            failed += 1
            errors.append({"row": row_idx, "tagName": tag_name, "message": str(exc)})
            continue

        if is_update:
            updated += 1
        else:
            inserted += 1

    # 写入导入汇总审计日志
    await _write_audit(
        db=db,
        operator=operator,
        operation_type="TAG_IMPORT",
        target_type="tag_registry",
        target_id=None,  # 导入操作无单一目标记录
        after_value=json.dumps(
            {
                "total": total,
                "inserted": inserted,
                "updated": updated,
                "failed": failed,
            },
            ensure_ascii=False,
        ),
    )
    await db.commit()

    # 导入结束后统一通知实时订阅 Leader 刷新一次（fire-and-forget，不逐行发）
    from app.services.data_source.realtime_subscriber import notify_subscription_changed

    await notify_subscription_changed(source="tag-import")

    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "errors": errors,
    }


async def _import_one_row(
    db: AsyncSession,
    tag_name: str,
    tag_description: str,
    measure_type: str | None,
    range_min: float | None,
    range_max: float | None,
    unit: str | None,
    tag_type: str | None,
    tdengine_tag_id: str | None,
    operator: str,
) -> bool:
    """处理单行导入，返回是否为更新（True）或新建（False）。

    在调用方的 SAVEPOINT 内执行，异常会触发回滚至 SAVEPOINT。
    注意（WS-C 7-10）：不接收 is_linked——该字段仅由回路映射派生，
    导入既不覆盖已有 tag 的 is_linked，新建 tag 也恒为 False。
    """
    result = await db.execute(select(TagRegistry).where(TagRegistry.tag_name == tag_name))
    tag = result.scalar_one_or_none()
    is_update = tag is not None

    if is_update:
        before = {
            "tagDescription": tag.tag_description,
            "measureType": tag.measure_type,
            "rangeMin": tag.range_min,
            "rangeMax": tag.range_max,
            "unit": tag.unit,
            "tagType": tag.tag_type,
            "tdengineTagId": tag.tdengine_tag_id,
            "isLinked": tag.is_linked,
        }
        before_json = json.dumps(before, ensure_ascii=False, default=str)

        if tag_description:
            tag.tag_description = tag_description
        if measure_type is not None:
            tag.measure_type = measure_type
        if range_min is not None:
            tag.range_min = range_min
        if range_max is not None:
            tag.range_max = range_max
        if unit is not None:
            tag.unit = unit
        if tag_type is not None:
            tag.tag_type = tag_type
        if tdengine_tag_id is not None:
            tag.tdengine_tag_id = tdengine_tag_id
        # WS-C 7-10：is_linked 不由导入覆盖（仅由回路映射派生）

        operation_type = "TAG_UPDATE"
    else:
        before_json = None
        tag = TagRegistry(
            id=str(uuid4()),
            tag_name=tag_name,
            tag_description=tag_description or None,
            tag_type=tag_type or "OTHER",
            measure_type=measure_type,
            range_min=range_min,
            range_max=range_max,
            unit=unit,
            tdengine_tag_id=tdengine_tag_id,
            is_linked=False,  # WS-C 7-10：新建恒为 False，待回路映射后置 True
            last_sync_at=datetime.now(UTC).replace(tzinfo=None),
        )
        db.add(tag)
        operation_type = "TAG_CREATE"

    await db.flush()

    after = {
        "tagName": tag.tag_name,
        "tagDescription": tag.tag_description,
        "tagType": tag.tag_type,
        "measureType": tag.measure_type,
        "rangeMin": tag.range_min,
        "rangeMax": tag.range_max,
        "unit": tag.unit,
        "tdengineTagId": tag.tdengine_tag_id,
        "isLinked": tag.is_linked,
    }
    after_json = json.dumps(after, ensure_ascii=False, default=str)

    await _write_audit(
        db=db,
        operator=operator,
        operation_type=operation_type,
        target_type="tag_registry",
        target_id=str(tag.id),
        before_value=before_json,
        after_value=after_json,
    )

    await db.flush()
    return is_update


__all__ = [
    "MEASURE_TYPES",
    "TAG_TYPES",
    "batch_delete_tags",
    "create_tag",
    "delete_tag",
    "export_tags",
    "get_tag_detail",
    "import_tags",
    "list_tags",
    "update_tag",
]
