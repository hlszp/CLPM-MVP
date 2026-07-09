"""DCS 配置 CRUD 服务（品牌/型号/MODE 定义/映射矩阵）.

对齐 DDS §3.1，配置驱动的 DCS 管理。

服务清单：
- DcsVendor CRUD（品牌）
- DcsModel CRUD（型号，全局唯一 code）
- ModeDefinition CRUD（标准 MODE 定义）
- DcsModeMapping CRUD + 矩阵视图查询
"""

from __future__ import annotations

import io
import logging
from datetime import UTC, datetime
from uuid import uuid4

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BizError
from app.models.dcs_mode_mapping import DcsModeMapping
from app.models.dcs_model import DcsModel
from app.models.dcs_vendor import DcsVendor
from app.models.mode_definition import ModeDefinition

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 导入导出表头（v6.1：品牌/型号 Excel 导入导出）
# ---------------------------------------------------------------------------

# 品牌导出表头（6 列）
_VENDOR_EXPORT_HEADERS = [
    "品牌代码",
    "中文名",
    "英文名",
    "描述",
    "排序",
    "启用状态",
]

# 型号导出表头（6 列）
_MODEL_EXPORT_HEADERS = [
    "型号代码",
    "型号名称",
    "品牌代码",
    "描述",
    "排序",
    "启用状态",
]


def _cell_str(value: object) -> str:
    """将 Excel 单元格值转为去除首尾空白的字符串，None/空返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(raw: str) -> bool:
    """解析启用状态：是/true/1/yes/y → True，其余 → False。"""
    return raw.lower() in ("是", "true", "1", "yes", "y")


# ---------------------------------------------------------------------------
# 序列化辅助
# ---------------------------------------------------------------------------


def _vendor_to_dict(v: DcsVendor) -> dict:
    return {
        "id": str(v.id),
        "code": v.code,
        "name": v.name,
        "name_en": v.name_en,
        "description": v.description,
        "sort_order": v.sort_order,
        "is_active": bool(v.is_active),
        "created_at": v.created_at.isoformat() if v.created_at else None,
        "updated_at": v.updated_at.isoformat() if v.updated_at else None,
    }


def _model_to_dict(m: DcsModel, vendor: DcsVendor | None = None) -> dict:
    return {
        "id": str(m.id),
        "vendor_id": str(m.vendor_id),
        "vendor_code": vendor.code if vendor else None,
        "vendor_name": vendor.name if vendor else None,
        "code": m.code,
        "name": m.name,
        "description": m.description,
        "sort_order": m.sort_order,
        "is_active": bool(m.is_active),
        "created_at": m.created_at.isoformat() if m.created_at else None,
        "updated_at": m.updated_at.isoformat() if m.updated_at else None,
    }


def _mode_def_to_dict(d: ModeDefinition) -> dict:
    return {
        "id": str(d.id),
        "standard_mode": d.standard_mode,
        "label_zh": d.label_zh,
        "label_en": d.label_en,
        "is_auto": bool(d.is_auto),
        "color": d.color,
        "sort_order": d.sort_order,
        "description": d.description,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _mapping_to_dict(m: DcsModeMapping, model: DcsModel | None = None) -> dict:
    return {
        "id": str(m.id),
        "dcs_model_id": str(m.dcs_model_id) if m.dcs_model_id else None,
        "model_code": model.code if model else None,
        "model_name": model.name if model else None,
        "standard_mode": m.standard_mode,
        "raw_mode_value": m.raw_mode_value,
        "description": m.description,
        "updated_at": m.updated_at.isoformat() if m.updated_at else None,
    }


# ---------------------------------------------------------------------------
# DcsVendor CRUD
# ---------------------------------------------------------------------------


async def list_vendors(db: AsyncSession) -> list[dict]:
    """获取全部品牌（按 sort_order 排序）。"""
    result = await db.execute(
        select(DcsVendor).order_by(DcsVendor.sort_order.asc(), DcsVendor.code.asc())
    )
    return [_vendor_to_dict(v) for v in result.scalars().all()]


async def create_vendor(
    db: AsyncSession,
    *,
    code: str,
    name: str,
    name_en: str | None = None,
    description: str | None = None,
    sort_order: int = 0,
    operator: str = "system",
) -> dict:
    """创建品牌。code 唯一，重复时抛 BizError。"""
    existing = await db.execute(select(DcsVendor).where(DcsVendor.code == code))
    if existing.scalar_one_or_none():
        raise BizError(
            code="ERR_DCS_VENDOR_CODE_DUPLICATE",
            message=f"品牌代码已存在: {code}",
            status_code=409,
        )
    vendor = DcsVendor(
        id=str(uuid4()),
        code=code,
        name=name,
        name_en=name_en,
        description=description,
        sort_order=sort_order,
        is_active=True,
    )
    db.add(vendor)
    await db.commit()
    await db.refresh(vendor)
    logger.info("[DCS 品牌] 新增 %s(%s) by %s", code, name, operator)
    return _vendor_to_dict(vendor)


async def update_vendor(
    db: AsyncSession,
    vendor_id: str,
    *,
    name: str | None = None,
    name_en: str | None = None,
    description: str | None = None,
    sort_order: int | None = None,
    is_active: bool | None = None,
    operator: str = "system",
) -> dict:
    """更新品牌（code 不可改）。"""
    result = await db.execute(select(DcsVendor).where(DcsVendor.id == vendor_id))
    vendor = result.scalar_one_or_none()
    if not vendor:
        raise BizError(
            code="ERR_DCS_VENDOR_NOT_FOUND",
            message=f"品牌不存在: {vendor_id}",
            status_code=404,
        )
    if name is not None:
        vendor.name = name
    if name_en is not None:
        vendor.name_en = name_en
    if description is not None:
        vendor.description = description
    if sort_order is not None:
        vendor.sort_order = sort_order
    if is_active is not None:
        vendor.is_active = is_active
    vendor.updated_at = datetime.now(UTC).replace(tzinfo=None)
    await db.commit()
    await db.refresh(vendor)
    logger.info("[DCS 品牌] 更新 %s by %s", vendor.code, operator)
    return _vendor_to_dict(vendor)


async def delete_vendor(db: AsyncSession, vendor_id: str, operator: str = "system") -> None:
    """删除品牌（若有关联型号则禁止删除）。"""
    result = await db.execute(select(DcsVendor).where(DcsVendor.id == vendor_id))
    vendor = result.scalar_one_or_none()
    if not vendor:
        raise BizError(
            code="ERR_DCS_VENDOR_NOT_FOUND",
            message=f"品牌不存在: {vendor_id}",
            status_code=404,
        )
    # 检查是否有关联型号
    model_count = await db.execute(
        select(DcsModel).where(DcsModel.vendor_id == vendor_id).limit(1)
    )
    if model_count.scalar_one_or_none():
        raise BizError(
            code="ERR_DCS_VENDOR_HAS_MODELS",
            message=f"品牌 {vendor.code} 下仍有型号，无法删除",
            status_code=409,
        )
    await db.execute(delete(DcsVendor).where(DcsVendor.id == vendor_id))
    await db.commit()
    logger.info("[DCS 品牌] 删除 %s by %s", vendor.code, operator)


# ---------------------------------------------------------------------------
# DcsModel CRUD
# ---------------------------------------------------------------------------


async def list_models(db: AsyncSession, vendor_id: str | None = None) -> list[dict]:
    """获取全部型号（可按品牌筛选，按 sort_order 排序）。"""
    stmt = select(DcsModel, DcsVendor).join(DcsVendor, DcsModel.vendor_id == DcsVendor.id)
    if vendor_id:
        stmt = stmt.where(DcsModel.vendor_id == vendor_id)
    stmt = stmt.order_by(DcsModel.sort_order.asc(), DcsModel.code.asc())
    result = await db.execute(stmt)
    return [_model_to_dict(m, v) for m, v in result.all()]


async def create_model(
    db: AsyncSession,
    *,
    vendor_id: str,
    code: str,
    name: str,
    description: str | None = None,
    sort_order: int = 0,
    operator: str = "system",
) -> dict:
    """创建型号。code 全局唯一，重复时抛 BizError。"""
    # 校验品牌存在
    vendor_result = await db.execute(select(DcsVendor).where(DcsVendor.id == vendor_id))
    vendor = vendor_result.scalar_one_or_none()
    if not vendor:
        raise BizError(
            code="ERR_DCS_VENDOR_NOT_FOUND",
            message=f"品牌不存在: {vendor_id}",
            status_code=404,
        )
    # 校验 code 全局唯一
    existing = await db.execute(select(DcsModel).where(DcsModel.code == code))
    if existing.scalar_one_or_none():
        raise BizError(
            code="ERR_DCS_MODEL_CODE_DUPLICATE",
            message=f"型号代码已存在: {code}",
            status_code=409,
        )
    model = DcsModel(
        id=str(uuid4()),
        vendor_id=vendor_id,
        code=code,
        name=name,
        description=description,
        sort_order=sort_order,
        is_active=True,
    )
    db.add(model)
    await db.commit()
    await db.refresh(model)
    logger.info("[DCS 型号] 新增 %s(%s) by %s", code, name, operator)
    return _model_to_dict(model, vendor)


async def update_model(
    db: AsyncSession,
    model_id: str,
    *,
    name: str | None = None,
    description: str | None = None,
    sort_order: int | None = None,
    is_active: bool | None = None,
    operator: str = "system",
) -> dict:
    """更新型号（code/vendor_id 不可改）。"""
    result = await db.execute(
        select(DcsModel, DcsVendor)
        .join(DcsVendor, DcsModel.vendor_id == DcsVendor.id)
        .where(DcsModel.id == model_id)
    )
    row = result.first()
    if not row:
        raise BizError(
            code="ERR_DCS_MODEL_NOT_FOUND",
            message=f"型号不存在: {model_id}",
            status_code=404,
        )
    model, vendor = row
    if name is not None:
        model.name = name
    if description is not None:
        model.description = description
    if sort_order is not None:
        model.sort_order = sort_order
    if is_active is not None:
        model.is_active = is_active
    model.updated_at = datetime.now(UTC).replace(tzinfo=None)
    await db.commit()
    await db.refresh(model)
    logger.info("[DCS 型号] 更新 %s by %s", model.code, operator)
    return _model_to_dict(model, vendor)


async def delete_model(db: AsyncSession, model_id: str, operator: str = "system") -> None:
    """删除型号（级联删除映射，loop_ledger.dcs_model_id SET NULL）。"""
    result = await db.execute(select(DcsModel).where(DcsModel.id == model_id))
    model = result.scalar_one_or_none()
    if not model:
        raise BizError(
            code="ERR_DCS_MODEL_NOT_FOUND",
            message=f"型号不存在: {model_id}",
            status_code=404,
        )
    await db.execute(delete(DcsModel).where(DcsModel.id == model_id))
    await db.commit()
    logger.info("[DCS 型号] 删除 %s by %s", model.code, operator)


# ---------------------------------------------------------------------------
# ModeDefinition CRUD
# ---------------------------------------------------------------------------


async def list_mode_definitions(db: AsyncSession) -> list[dict]:
    """获取全部标准 MODE 定义（按 sort_order 排序）。"""
    result = await db.execute(
        select(ModeDefinition).order_by(ModeDefinition.sort_order.asc())
    )
    return [_mode_def_to_dict(d) for d in result.scalars().all()]


async def update_mode_definition(
    db: AsyncSession,
    standard_mode: int,
    *,
    label_zh: str | None = None,
    label_en: str | None = None,
    is_auto: bool | None = None,
    color: str | None = None,
    description: str | None = None,
    operator: str = "system",
) -> dict:
    """更新标准 MODE 定义（standard_mode 不可改，is_auto 影响自控率计算）。"""
    result = await db.execute(
        select(ModeDefinition).where(ModeDefinition.standard_mode == standard_mode)
    )
    d = result.scalar_one_or_none()
    if not d:
        raise BizError(
            code="ERR_MODE_DEFINITION_NOT_FOUND",
            message=f"标准 MODE 定义不存在: {standard_mode}",
            status_code=404,
        )
    if label_zh is not None:
        d.label_zh = label_zh
    if label_en is not None:
        d.label_en = label_en
    if is_auto is not None:
        d.is_auto = is_auto
    if color is not None:
        d.color = color
    if description is not None:
        d.description = description
    d.updated_at = datetime.now(UTC).replace(tzinfo=None)
    await db.commit()
    await db.refresh(d)
    logger.info(
        "[MODE 定义] 更新 standard_mode=%d is_auto=%s by %s",
        standard_mode,
        d.is_auto,
        operator,
    )
    return _mode_def_to_dict(d)


# ---------------------------------------------------------------------------
# DcsModeMapping CRUD + 矩阵视图
# ---------------------------------------------------------------------------


async def list_mode_mappings(
    db: AsyncSession,
    dcs_model_id: str | None = None,
) -> list[dict]:
    """获取 MODE 映射列表（可按型号筛选）。"""
    stmt = select(DcsModeMapping, DcsModel).outerjoin(
        DcsModel, DcsModeMapping.dcs_model_id == DcsModel.id
    )
    if dcs_model_id is not None:
        stmt = stmt.where(
            (DcsModeMapping.dcs_model_id == dcs_model_id)
            | (DcsModeMapping.dcs_model_id.is_(None))
        )
    stmt = stmt.order_by(
        DcsModeMapping.dcs_model_id.asc().nullsfirst(),
        DcsModeMapping.standard_mode.asc(),
    )
    result = await db.execute(stmt)
    return [_mapping_to_dict(m, model) for m, model in result.all()]


async def upsert_mode_mapping(
    db: AsyncSession,
    *,
    dcs_model_id: str | None,
    standard_mode: int,
    raw_mode_value: int,
    description: str | None = None,
    operator: str = "system",
) -> dict:
    """创建或更新 MODE 映射（按 dcs_model_id + standard_mode 幂等）."""
    # 查找现有记录
    stmt = select(DcsModeMapping).where(
        DcsModeMapping.standard_mode == standard_mode
    )
    if dcs_model_id:
        stmt = stmt.where(DcsModeMapping.dcs_model_id == dcs_model_id)
    else:
        stmt = stmt.where(DcsModeMapping.dcs_model_id.is_(None))
    result = await db.execute(stmt)
    mapping = result.scalar_one_or_none()

    if mapping:
        # 更新
        mapping.raw_mode_value = raw_mode_value
        if description is not None:
            mapping.description = description
        mapping.updated_at = datetime.now(UTC).replace(tzinfo=None)
    else:
        # 新建
        mapping = DcsModeMapping(
            id=str(uuid4()),
            dcs_model_id=dcs_model_id,
            standard_mode=standard_mode,
            raw_mode_value=raw_mode_value,
            description=description,
        )
        db.add(mapping)
    await db.commit()
    await db.refresh(mapping)

    # 查关联型号
    model = None
    if mapping.dcs_model_id:
        model_result = await db.execute(select(DcsModel).where(DcsModel.id == mapping.dcs_model_id))
        model = model_result.scalar_one_or_none()

    logger.info(
        "[MODE 映射] upsert model_id=%s standard=%d raw=%d by %s",
        dcs_model_id,
        standard_mode,
        raw_mode_value,
        operator,
    )
    return _mapping_to_dict(mapping, model)


async def delete_mode_mapping(db: AsyncSession, mapping_id: str, operator: str = "system") -> None:
    """删除 MODE 映射。"""
    result = await db.execute(select(DcsModeMapping).where(DcsModeMapping.id == mapping_id))
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise BizError(
            code="ERR_MODE_MAPPING_NOT_FOUND",
            message=f"MODE 映射不存在: {mapping_id}",
            status_code=404,
        )
    await db.execute(delete(DcsModeMapping).where(DcsModeMapping.id == mapping_id))
    await db.commit()
    logger.info("[MODE 映射] 删除 id=%s by %s", mapping_id, operator)


async def get_mode_matrix(db: AsyncSession) -> dict:
    """获取 MODE 映射矩阵视图（行=标准 MODE，列=各型号）.

    返回结构：
    - columns: 列头列表（第一列为本系统默认，后续为各型号）
    - rows: 行列表，每行包含 standard_mode/label 和 columns 数据

    满足用户需求"表头和第一、第二行为本系统默认，后续为各品牌型号"。
    """
    # 查询所有型号（按 sort_order 排序）
    models_result = await db.execute(
        select(DcsModel).order_by(DcsModel.sort_order.asc(), DcsModel.code.asc())
    )
    models = models_result.scalars().all()

    # 查询所有 MODE 定义（按 sort_order 排序）
    defs_result = await db.execute(
        select(ModeDefinition).order_by(ModeDefinition.sort_order.asc())
    )
    defs = defs_result.scalars().all()

    # 查询所有映射
    mappings_result = await db.execute(select(DcsModeMapping))
    mappings = mappings_result.scalars().all()

    # 构建 {(dcs_model_id, standard_mode): raw_mode_value} 映射
    # dcs_model_id 为 None 的 key 用 "default" 表示
    mapping_map: dict[tuple[str | None, int], int] = {}
    for m in mappings:
        key = (str(m.dcs_model_id) if m.dcs_model_id else None, m.standard_mode)
        mapping_map[key] = m.raw_mode_value

    # 批量查询所有品牌（避免 N+1 查询）
    vendor_ids = {m.vendor_id for m in models if m.vendor_id}
    vendor_map: dict[str, str] = {}
    if vendor_ids:
        v_result = await db.execute(
            select(DcsVendor.id, DcsVendor.name).where(DcsVendor.id.in_(vendor_ids))
        )
        for vid, vname in v_result.all():
            vendor_map[str(vid)] = vname

    # 构建列头（第一列为本系统默认）
    columns = [
        {
            "model_id": None,
            "model_code": "default",
            "model_name": "本系统默认",
            "vendor_id": None,
            "vendor_name": None,
        }
    ]
    for model in models:
        v_name = vendor_map.get(str(model.vendor_id)) if model.vendor_id else None
        columns.append(
            {
                "model_id": str(model.id),
                "model_code": model.code,
                "model_name": model.name,
                "vendor_id": str(model.vendor_id) if model.vendor_id else None,
                "vendor_name": v_name,
            }
        )

    # 构建行（每行一个 standard_mode）
    rows = []
    for d in defs:
        row_columns = []
        # 第一列：本系统默认
        default_raw = mapping_map.get((None, d.standard_mode))
        row_columns.append(
            {
                "model_id": None,
                "model_code": "default",
                "model_name": "本系统默认",
                "vendor_id": None,
                "vendor_name": None,
                "raw_mode_value": default_raw,
            }
        )
        # 后续列：各型号
        for model in models:
            raw = mapping_map.get((str(model.id), d.standard_mode))
            v_name = vendor_map.get(str(model.vendor_id)) if model.vendor_id else None
            row_columns.append(
                {
                    "model_id": str(model.id),
                    "model_code": model.code,
                    "model_name": model.name,
                    "vendor_id": str(model.vendor_id) if model.vendor_id else None,
                    "vendor_name": v_name,
                    "raw_mode_value": raw,
                }
            )
        rows.append(
            {
                "standard_mode": d.standard_mode,
                "label_zh": d.label_zh,
                "label_en": d.label_en,
                "is_auto": bool(d.is_auto),
                "color": d.color,
                "columns": row_columns,
            }
        )

    return {"rows": rows, "columns": columns}


# ---------------------------------------------------------------------------
# 品牌导入导出（v6.1）
# ---------------------------------------------------------------------------


async def export_vendors(db: AsyncSession) -> bytes:
    """导出全部品牌为 Excel 文件（.xlsx），返回文件字节。

    按品牌代码排序导出，包含 6 列：品牌代码/中文名/英文名/描述/排序/启用状态。
    """
    result = await db.execute(
        select(DcsVendor).order_by(DcsVendor.sort_order.asc(), DcsVendor.code.asc())
    )
    vendors = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCS品牌"
    ws.append(_VENDOR_EXPORT_HEADERS)
    for v in vendors:
        ws.append(
            [
                v.code,
                v.name,
                v.name_en or "",
                v.description or "",
                v.sort_order,
                "是" if v.is_active else "否",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_vendors(
    db: AsyncSession,
    file_bytes: bytes,
    operator: str = "system",
) -> dict:
    """批量导入品牌（Excel .xlsx）。

    按 code 去重：已存在则更新，否则新建。
    返回 {total, inserted, updated, failed, errors[]}。
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(
            code="ERR_FILE_PARSE",
            message=f"Excel 文件解析失败: {exc}",
            status_code=400,
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    total = 0
    inserted = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(rows, start=2):
        total += 1
        code = _cell_str(row[0]) if len(row) > 0 else ""
        if not code:
            errors.append({"row": row_idx, "message": "品牌代码不能为空"})
            failed += 1
            continue

        name = _cell_str(row[1]) if len(row) > 1 else ""
        if not name:
            errors.append({"row": row_idx, "code": code, "message": "中文名不能为空"})
            failed += 1
            continue

        name_en = _cell_str(row[2]) if len(row) > 2 else ""
        description = _cell_str(row[3]) if len(row) > 3 else ""
        sort_order_str = _cell_str(row[4]) if len(row) > 4 else ""
        try:
            sort_order = int(sort_order_str) if sort_order_str else 0
        except ValueError:
            sort_order = 0
        is_active_str = _cell_str(row[5]) if len(row) > 5 else "是"
        is_active = _parse_bool(is_active_str) if is_active_str else True

        try:
            async with db.begin_nested():
                existing = await db.execute(select(DcsVendor).where(DcsVendor.code == code))
                vendor = existing.scalar_one_or_none()
                if vendor:
                    vendor.name = name
                    vendor.name_en = name_en or None
                    vendor.description = description or None
                    vendor.sort_order = sort_order
                    vendor.is_active = is_active
                    vendor.updated_at = datetime.now(UTC).replace(tzinfo=None)
                    updated += 1
                else:
                    new_vendor = DcsVendor(
                        id=str(uuid4()),
                        code=code,
                        name=name,
                        name_en=name_en or None,
                        description=description or None,
                        sort_order=sort_order,
                        is_active=is_active,
                    )
                    db.add(new_vendor)
                    inserted += 1
        except Exception as exc:  # noqa: BLE001
            failed += 1
            errors.append({"row": row_idx, "code": code, "message": str(exc)})
            continue

    await db.commit()
    logger.info(
        "[DCS 品牌] 导入完成 total=%d inserted=%d updated=%d failed=%d by %s",
        total,
        inserted,
        updated,
        failed,
        operator,
    )
    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# 型号导入导出（v6.1）
# ---------------------------------------------------------------------------


async def export_models(db: AsyncSession) -> bytes:
    """导出全部型号为 Excel 文件（.xlsx），返回文件字节。

    按型号代码排序导出，包含 6 列：型号代码/型号名称/品牌代码/描述/排序/启用状态。
    """
    result = await db.execute(
        select(DcsModel, DcsVendor)
        .join(DcsVendor, DcsModel.vendor_id == DcsVendor.id)
        .order_by(DcsModel.sort_order.asc(), DcsModel.code.asc())
    )
    rows_data = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCS型号"
    ws.append(_MODEL_EXPORT_HEADERS)
    for model, vendor in rows_data:
        ws.append(
            [
                model.code,
                model.name,
                vendor.code if vendor else "",
                model.description or "",
                model.sort_order,
                "是" if model.is_active else "否",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_models(
    db: AsyncSession,
    file_bytes: bytes,
    operator: str = "system",
) -> dict:
    """批量导入型号（Excel .xlsx）。

    按 code 去重：已存在则更新，否则新建。
    通过品牌代码查找 vendor_id（必须存在，否则该行失败）。
    返回 {total, inserted, updated, failed, errors[]}。
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BizError(
            code="ERR_FILE_PARSE",
            message=f"Excel 文件解析失败: {exc}",
            status_code=400,
        ) from exc

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    total = 0
    inserted = 0
    updated = 0
    failed = 0
    errors: list[dict] = []

    # 品牌代码缓存：vendor_code → vendor_id
    vendor_cache: dict[str, str] = {}

    for row_idx, row in enumerate(rows, start=2):
        total += 1
        code = _cell_str(row[0]) if len(row) > 0 else ""
        if not code:
            errors.append({"row": row_idx, "message": "型号代码不能为空"})
            failed += 1
            continue

        name = _cell_str(row[1]) if len(row) > 1 else ""
        if not name:
            errors.append({"row": row_idx, "code": code, "message": "型号名称不能为空"})
            failed += 1
            continue

        vendor_code = _cell_str(row[2]) if len(row) > 2 else ""
        if not vendor_code:
            errors.append({"row": row_idx, "code": code, "message": "品牌代码不能为空"})
            failed += 1
            continue

        # 查找品牌 ID（带缓存）
        if vendor_code in vendor_cache:
            vendor_id = vendor_cache[vendor_code]
        else:
            v_result = await db.execute(select(DcsVendor).where(DcsVendor.code == vendor_code))
            vendor = v_result.scalar_one_or_none()
            if not vendor:
                errors.append(
                    {
                        "row": row_idx,
                        "code": code,
                        "message": f"品牌代码不存在: {vendor_code}",
                    }
                )
                failed += 1
                continue
            vendor_id = str(vendor.id)
            vendor_cache[vendor_code] = vendor_id

        description = _cell_str(row[3]) if len(row) > 3 else ""
        sort_order_str = _cell_str(row[4]) if len(row) > 4 else ""
        try:
            sort_order = int(sort_order_str) if sort_order_str else 0
        except ValueError:
            sort_order = 0
        is_active_str = _cell_str(row[5]) if len(row) > 5 else "是"
        is_active = _parse_bool(is_active_str) if is_active_str else True

        try:
            async with db.begin_nested():
                existing = await db.execute(select(DcsModel).where(DcsModel.code == code))
                model = existing.scalar_one_or_none()
                if model:
                    model.name = name
                    model.description = description or None
                    model.sort_order = sort_order
                    model.is_active = is_active
                    model.updated_at = datetime.now(UTC).replace(tzinfo=None)
                    updated += 1
                else:
                    new_model = DcsModel(
                        id=str(uuid4()),
                        vendor_id=vendor_id,
                        code=code,
                        name=name,
                        description=description or None,
                        sort_order=sort_order,
                        is_active=is_active,
                    )
                    db.add(new_model)
                    inserted += 1
        except Exception as exc:  # noqa: BLE001
            failed += 1
            errors.append({"row": row_idx, "code": code, "message": str(exc)})
            continue

    await db.commit()
    logger.info(
        "[DCS 型号] 导入完成 total=%d inserted=%d updated=%d failed=%d by %s",
        total,
        inserted,
        updated,
        failed,
        operator,
    )
    return {
        "total": total,
        "inserted": inserted,
        "updated": updated,
        "failed": failed,
        "errors": errors,
    }


__all__ = [
    # Vendor
    "create_vendor",
    "delete_vendor",
    "export_vendors",
    "import_vendors",
    "list_vendors",
    "update_vendor",
    # Model
    "create_model",
    "delete_model",
    "export_models",
    "import_models",
    "list_models",
    "update_model",
    # ModeDefinition
    "list_mode_definitions",
    "update_mode_definition",
    # ModeMapping
    "delete_mode_mapping",
    "get_mode_matrix",
    "list_mode_mappings",
    "upsert_mode_mapping",
]
