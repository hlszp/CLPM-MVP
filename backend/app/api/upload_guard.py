"""Excel 上传统一防护：文件大小上限 + 数据行数上限。

背景（P2）：各导入端点 ``await file.read()`` 全量进内存、服务层
``list(ws.iter_rows())`` 全量物化，超大 xlsx 可打爆 API 进程内存。
本模块在 API 层统一拦截：

- 分块读取，超过 ``MAX_UPLOAD_BYTES``（10MB）立即拒绝（422），
  避免超大文件全量进内存；
- 以 ``read_only=True`` 流式遍历工作表统计数据行数（O(1) 额外内存），
  超过 ``MAX_IMPORT_ROWS``（5000 行，不含表头）拒绝（422）。

5 处 Excel 导入端点（loops/tags/plant_nodes/dcs vendors/dcs models）共用。
"""

from __future__ import annotations

import io

import openpyxl
from fastapi import UploadFile

from app.core.exceptions import BizError

# 上传文件大小上限：10MB
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
# 导入数据行数上限（不含表头）
MAX_IMPORT_ROWS = 5000

_READ_CHUNK_SIZE = 1024 * 1024  # 1MB


async def read_excel_upload(
    file: UploadFile,
    *,
    max_bytes: int = MAX_UPLOAD_BYTES,
    max_rows: int = MAX_IMPORT_ROWS,
) -> bytes:
    """读取并校验 Excel 上传文件，返回文件字节。

    Raises:
        BizError: ``ERR_FILE_TOO_LARGE``（422）文件超过 ``max_bytes``；
            ``ERR_TOO_MANY_ROWS``（422）数据行数超过 ``max_rows``。
    """
    chunks: list[bytes] = []
    total = 0
    while chunk := await file.read(_READ_CHUNK_SIZE):
        total += len(chunk)
        if total > max_bytes:
            raise BizError(
                code="ERR_FILE_TOO_LARGE",
                message=f"文件大小超过上限 {max_bytes // (1024 * 1024)}MB，请拆分后分批导入",
                status_code=422,
            )
        chunks.append(chunk)
    file_bytes = b"".join(chunks)
    check_excel_row_count(file_bytes, max_rows=max_rows)
    return file_bytes


def check_excel_row_count(file_bytes: bytes, *, max_rows: int = MAX_IMPORT_ROWS) -> None:
    """流式统计数据行数（不含表头），超限抛 ``ERR_TOO_MANY_ROWS``（422）。

    解析失败不在此报错，交由业务层统一返回 ``ERR_FILE_PARSE``。
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return
    try:
        ws = wb.active
        if ws is None:
            return
        row_count = 0
        for _row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            if row_count > max_rows:
                raise BizError(
                    code="ERR_TOO_MANY_ROWS",
                    message=f"数据行数超过上限 {max_rows} 行，请拆分后分批导入",
                    status_code=422,
                )
    finally:
        wb.close()


__all__ = [
    "MAX_IMPORT_ROWS",
    "MAX_UPLOAD_BYTES",
    "check_excel_row_count",
    "read_excel_upload",
]
