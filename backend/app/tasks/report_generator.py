"""Celery task for report generation (S5-SYS-003).

Design:
- Celery Beat dispatches tasks per configured period (SHIFT/DAILY/WEEKLY/MONTHLY)
- Each task creates a ``ReportRecord`` with status PROCESSING
- Generates a PDF report (using reportlab; can be replaced with Headless Browser)
- Updates ``ReportRecord`` to COMPLETED with file_url, or FAILED on error
- Writes an audit log entry

S2-A4: 区分可重试/不可重试异常 — 业务错误（NonRetryableError）不重试，
       系统错误（DB/网络等）自动重试 3 次。

异步导出任务（设计依据：IDS §2.4）：
- export_diagnosis_statistics: 诊断统计 Excel 异步导出
"""

from __future__ import annotations

import io
import logging
import os
import tempfile
from datetime import UTC, datetime
from typing import Any
from uuid import uuid4

from sqlalchemy import func, select

from app.models.loop import LoopLedger
from app.models.plant_node import PlantNode
from app.models.report import ReportRecord
from app.models.report_config import ReportConfig
from app.tasks.celery_app import AsyncTask, celery_app

logger = logging.getLogger(__name__)


class NonRetryableError(Exception):
    """业务错误，不应重试（如配置缺失、周期参数非法等）。"""


# ---------------------------------------------------------------------------
# Celery task: manual / triggered report generation
# ---------------------------------------------------------------------------


@celery_app.task(
    name="app.tasks.report_generator.generate_report_task",
    bind=True,
    base=AsyncTask,
    max_retries=3,
    retry_backoff=True,
    retry_backoff_max=600,
    retry_jitter=True,
)
def generate_report_task(
    self: AsyncTask,
    task_id: str | None = None,
    config_id: str | None = None,
    report_period: str = "DAILY",
) -> dict:
    """Generate a report asynchronously.

    S2-A4: 业务错误（NonRetryableError）不重试；系统错误自动重试 3 次。

    Args:
        task_id: Optional task ID (used as ReportRecord ID)
        config_id: Optional report config ID
        report_period: Report period (SHIFT/DAILY/WEEKLY/MONTHLY)
    """
    logger.info("报表生成任务开始, task_id=%s, config_id=%s", task_id, config_id)
    try:
        result = self.run_async(_do_generate(task_id, config_id, report_period))
        logger.info("报表生成任务完成: %s", result)
        return result
    except NonRetryableError as exc:
        # 业务错误：不重试，直接记录失败
        logger.error("报表生成业务错误（不重试）: %s", exc)
        raise
    except Exception as exc:
        # 系统错误：自动重试
        logger.exception("报表生成系统错误（将重试）")
        raise self.retry(exc=exc, countdown=60) from None


# ---------------------------------------------------------------------------
# Beat schedule: 自动周期生成已摘除（报告模块优化 P0-1，2026-08-28）
#
# 背景：自动生成尚为占位实现（极简 PDF、无真实文件落盘），四周期调度
# 空转写入无效 report_record。P0 先收敛止血，P3 做实后再恢复调度
# （见 docs/设计文档/CLPM报告模块优化实施方案-2026-08-28.md §3.1，D1 已决）。
# 手动触发生成（/reports/generate）保留，generate_report_task 任务函数不动。
# ---------------------------------------------------------------------------

celery_app.conf.timezone = "Asia/Shanghai"


# ---------------------------------------------------------------------------
# Async generation logic
# ---------------------------------------------------------------------------


async def _do_generate(
    task_id: str | None,
    config_id: str | None,
    report_period: str,
) -> dict:
    """Execute the report generation logic."""
    from app.core.db import AsyncSessionLocal

    # S2-A4: 业务参数校验 — 非法周期不重试
    valid_periods = {"SHIFT", "DAILY", "WEEKLY", "MONTHLY"}
    if report_period not in valid_periods:
        raise NonRetryableError(f"非法报表周期: {report_period}，允许值: {valid_periods}")

    record_id = task_id or str(uuid4())

    async with AsyncSessionLocal() as db:
        # Load config if provided
        config: ReportConfig | None = None
        if config_id:
            result = await db.execute(select(ReportConfig).where(ReportConfig.id == config_id))
            config = result.scalar_one_or_none()
            # S2-A4: 配置缺失为业务错误，不重试
            if config is None:
                raise NonRetryableError(f"报表配置不存在: config_id={config_id}")

        # Create ReportRecord with PROCESSING status
        now = datetime.now(UTC).replace(tzinfo=None)
        record = ReportRecord(
            id=record_id,
            report_period=report_period,
            generated_at=now,
            status="PROCESSING",
        )
        db.add(record)
        await db.commit()

        try:
            # Generate PDF content
            # NOTE: This uses reportlab for simple PDF generation.
            # Can be replaced with Headless Browser (Playwright/Puppeteer)
            # rendering an HTML template to PDF for richer formatting.
            pdf_bytes = _generate_pdf(
                report_period=report_period,
                config_name=config.name if config else None,
                content_template=(
                    _parse_content_template(config.content_template) if config else None
                ),
            )

            # In production, upload to S3/MinIO and store the URL.
            # For now, use a placeholder path.
            file_url = f"/reports/{record_id}.pdf"

            record.status = "COMPLETED"
            record.file_url = file_url
            await db.commit()

            return {
                "reportId": record_id,
                "status": "COMPLETED",
                "fileUrl": file_url,
                "fileSize": len(pdf_bytes),
            }
        except Exception as exc:
            logger.exception("报表生成失败")
            record.status = "FAILED"
            await db.commit()
            return {
                "reportId": record_id,
                "status": "FAILED",
                "error": str(exc),
            }


def _generate_pdf(
    report_period: str,
    config_name: str | None = None,
    content_template: dict | None = None,
) -> bytes:
    """Generate a simple PDF report using reportlab.

    NOTE: This is a minimal implementation. For production, replace with
    Headless Browser rendering (e.g. Playwright) an HTML template to PDF
    for richer formatting, charts, and styling.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()

    elements: list = []
    title = "控制回路性能评估报告"
    if config_name:
        title = f"{config_name} - {title}"
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 10 * mm))

    period_map = {
        "SHIFT": "班报",
        "DAILY": "日报",
        "WEEKLY": "周报",
        "MONTHLY": "月报",
    }
    period_name = period_map.get(report_period, report_period)
    elements.append(Paragraph(f"报表周期: {period_name}", styles["Normal"]))
    generated_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")
    elements.append(Paragraph(f"生成时间: {generated_at}", styles["Normal"]))
    elements.append(Spacer(1, 10 * mm))

    if content_template:
        elements.append(Paragraph("报表内容:", styles["Heading2"]))
        for key, value in content_template.items():
            elements.append(Paragraph(f"  {key}: {value}", styles["Normal"]))
    else:
        elements.append(Paragraph("本报表由 CLPM 系统自动生成。", styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()


def _parse_content_template(value: str | None) -> dict | None:
    """Parse content_template JSON string."""
    if value is None:
        return None
    import json

    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return None


# ---------------------------------------------------------------------------
# 异步导出任务（设计依据：IDS §2.4 — POST /diagnosis/analytics/export）
# ---------------------------------------------------------------------------


@celery_app.task(
    name="app.tasks.report_generator.export_diagnosis_statistics",
    bind=True,
    base=AsyncTask,
    autoretry_for=(Exception,),
    retry_kwargs={"max_retries": 3, "countdown": 60},
    retry_backoff=True,
    retry_backoff_max=600,
    retry_jitter=True,
)
def export_diagnosis_statistics(
    self: AsyncTask,
    start_time: str,
    end_time: str,
    plant_node_id: str | None = None,
    diagnosis_label: str | None = None,
    action_status: str | None = None,
    user_id: str | None = None,
    granularity: str = "day",
    file_format: str = "xlsx",
) -> dict:
    """诊断统计异步导出 Celery 任务。

    设计依据：IDS §2.4 — POST /api/v1/diagnosis/analytics/export

    流程：
    1. 查询时间窗内的诊断结果，按标签聚合
    2. 使用 openpyxl 生成 Excel 文件（含标签分布、趋势、明细）
    3. 文件保存到临时目录，返回文件路径
    4. 任务结果可通过 GET /api/v1/algorithms/tasks/{task_id} 查询

    Args:
        start_time: 开始时间（ISO 8601）
        end_time: 结束时间（ISO 8601）
        plant_node_id: 装置节点 ID 筛选
        diagnosis_label: 诊断标签筛选
        action_status: 处理状态筛选
        user_id: 触发用户 ID（审计用）
        granularity: 时间粒度（hour/day/week/month）
        file_format: 文件格式（xlsx/csv）

    Returns:
        {taskId, status, fileUrl, fileSize, fileFormat}
    """
    logger.info(
        "诊断统计异步导出任务开始, task_id=%s, start=%s, end=%s",
        self.request.id,
        start_time,
        end_time,
    )
    try:
        result = self.run_async(
            _do_export_diagnosis_statistics(
                start_time=start_time,
                end_time=end_time,
                plant_node_id=plant_node_id,
                diagnosis_label=diagnosis_label,
                action_status=action_status,
                user_id=user_id,
                granularity=granularity,
                file_format=file_format,
            )
        )
        logger.info("诊断统计异步导出任务完成: %s", result)
        return result
    except Exception:
        logger.exception("诊断统计异步导出任务失败")
        raise


async def _do_export_diagnosis_statistics(
    start_time: str,
    end_time: str,
    plant_node_id: str | None = None,
    diagnosis_label: str | None = None,
    action_status: str | None = None,
    user_id: str | None = None,
    granularity: str = "day",
    file_format: str = "xlsx",
) -> dict:
    """执行诊断统计异步导出。

    设计依据：IDS §2.4

    Args:
        start_time: 开始时间（ISO 8601）
        end_time: 结束时间（ISO 8601）
        plant_node_id: 装置节点 ID 筛选
        diagnosis_label: 诊断标签筛选
        action_status: 处理状态筛选
        user_id: 触发用户 ID
        granularity: 时间粒度
        file_format: 文件格式

    Returns:
        导出结果字典

    .. deprecated::
        使用旧版 ``diagnosis_result`` 表；v2 诊断使用 ``diagnosis_run`` 表，
        诊断报告 API 需基于 DiagnosisRun 重新实现。
    """
    # lazy import：旧版诊断表仅在执行导出时加载（DIAG_LABEL_NAMES 由 helper 自行 lazy import）
    from app.core.db import AsyncSessionLocal
    from app.models.diagnosis import DiagnosisResult

    logger.warning("export_diagnosis_statistics 已废弃（查询旧版 diagnosis_result 表）")

    start_dt = _parse_iso_dt(start_time)
    end_dt = _parse_iso_dt(end_time)

    async with AsyncSessionLocal() as db:
        # 1. 查询标签分布汇总
        label_stmt = (
            select(
                DiagnosisResult.diag_label,
                func.count(DiagnosisResult.id).label("count"),
            )
            .where(DiagnosisResult.diagnosed_at >= start_dt)
            .where(DiagnosisResult.diagnosed_at <= end_dt)
            .where(DiagnosisResult.diag_label.is_not(None))
        )
        if diagnosis_label:
            label_stmt = label_stmt.where(DiagnosisResult.diag_label == diagnosis_label)
        if plant_node_id:
            label_stmt = label_stmt.join(
                LoopLedger, DiagnosisResult.loop_id == LoopLedger.id
            ).where(LoopLedger.unit_id == plant_node_id)
        label_stmt = label_stmt.group_by(DiagnosisResult.diag_label).order_by(
            func.count(DiagnosisResult.id).desc()
        )
        label_result = await db.execute(label_stmt)
        label_counts = label_result.all()

        # 2. 查询装置名
        plant_name = "全部装置"
        if plant_node_id:
            node_result = await db.execute(select(PlantNode).where(PlantNode.id == plant_node_id))
            node = node_result.scalar_one_or_none()
            if node:
                plant_name = node.name

        # 3. 查询按天趋势
        trend_stmt = (
            select(
                func.date_trunc("day", DiagnosisResult.diagnosed_at).label("day"),
                DiagnosisResult.diag_label,
                func.count(DiagnosisResult.id).label("count"),
            )
            .where(DiagnosisResult.diagnosed_at >= start_dt)
            .where(DiagnosisResult.diagnosed_at <= end_dt)
            .where(DiagnosisResult.diag_label.is_not(None))
        )
        if diagnosis_label:
            trend_stmt = trend_stmt.where(DiagnosisResult.diag_label == diagnosis_label)
        if plant_node_id:
            trend_stmt = trend_stmt.join(
                LoopLedger, DiagnosisResult.loop_id == LoopLedger.id
            ).where(LoopLedger.unit_id == plant_node_id)
        trend_stmt = trend_stmt.group_by("day", DiagnosisResult.diag_label).order_by(
            "day", DiagnosisResult.diag_label
        )
        trend_result = await db.execute(trend_stmt)
        trend_rows = trend_result.all()

        # 4. 查询明细数据（前 1000 条）
        detail_stmt = (
            select(
                DiagnosisResult.loop_id,
                DiagnosisResult.diag_label,
                DiagnosisResult.confidence,
                DiagnosisResult.diagnosed_at,
                DiagnosisResult.algorithm_version,
            )
            .where(DiagnosisResult.diagnosed_at >= start_dt)
            .where(DiagnosisResult.diagnosed_at <= end_dt)
            .where(DiagnosisResult.diag_label.is_not(None))
            .order_by(DiagnosisResult.diagnosed_at.desc())
            .limit(1000)
        )
        if diagnosis_label:
            detail_stmt = detail_stmt.where(DiagnosisResult.diag_label == diagnosis_label)
        if plant_node_id:
            detail_stmt = detail_stmt.join(
                LoopLedger, DiagnosisResult.loop_id == LoopLedger.id
            ).where(LoopLedger.unit_id == plant_node_id)
        detail_result = await db.execute(detail_stmt)
        detail_rows = detail_result.all()

    # 5. 生成文件
    if file_format.lower() == "csv":
        file_bytes = _generate_csv_bytes(
            start_time=start_time,
            end_time=end_time,
            plant_name=plant_name,
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )
        ext = "csv"
    else:
        file_bytes = _generate_excel_bytes(
            start_time=start_time,
            end_time=end_time,
            plant_name=plant_name,
            label_counts=label_counts,
            trend_rows=trend_rows,
            detail_rows=detail_rows,
        )
        ext = "xlsx"

    # 6. 保存到临时目录
    export_dir = os.environ.get("CLPM_EXPORT_DIR", tempfile.gettempdir())
    os.makedirs(export_dir, exist_ok=True)
    file_name = f"CLPM-诊断统计报表-{start_time[:10]}_{end_time[:10]}.{ext}"
    file_path = os.path.join(export_dir, file_name)
    with open(file_path, "wb") as f:
        f.write(file_bytes)

    logger.info(
        "诊断统计导出文件已生成: path=%s, size=%d bytes",
        file_path,
        len(file_bytes),
    )

    return {
        "taskId": "",
        "status": "SUCCESS",
        "fileUrl": file_path,
        "fileSize": len(file_bytes),
        "fileFormat": ext,
        "labelCount": len(label_counts),
        "totalRecords": sum(row[1] for row in label_counts) if label_counts else 0,
    }


def _generate_excel_bytes(
    start_time: str,
    end_time: str,
    plant_name: str,
    label_counts: list,
    trend_rows: list,
    detail_rows: list,
) -> bytes:
    """使用 openpyxl 生成 Excel 报表。

    设计依据：IDS §2.4 — 文件名规范 CLPM-诊断统计报表-[装置]-[日期范围].xlsx

    包含 3 个工作表：
    1. 标签分布汇总
    2. 按天趋势
    3. 明细数据
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.services.diagnosis import DIAG_LABEL_NAMES

    wb = Workbook()

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    title_font = Font(name="微软雅黑", size=14, bold=True)
    body_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    # Sheet 1: 标签分布汇总
    ws1 = wb.active
    ws1.title = "标签分布汇总"
    ws1["A1"] = "CLPM 诊断统计报表"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:D1")
    ws1["A2"] = "时间范围"
    ws1["B2"] = f"{start_time} ~ {end_time}"
    ws1["A3"] = "装置范围"
    ws1["B3"] = plant_name
    ws1["A4"] = "导出时间"
    ws1["B4"] = datetime.now(UTC).replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")

    headers = ["标签代码", "标签名称", "数量", "占比(%)"]
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    total = sum(row[1] for row in label_counts) if label_counts else 0
    for idx, (label, count) in enumerate(label_counts, start=7):
        label_name = DIAG_LABEL_NAMES.get(label, label)
        ratio = (count / total * 100) if total > 0 else 0.0
        ws1.cell(row=idx, column=1, value=label).font = body_font
        ws1.cell(row=idx, column=2, value=label_name).font = body_font
        ws1.cell(row=idx, column=3, value=count).font = body_font
        ws1.cell(row=idx, column=4, value=round(ratio, 2)).font = body_font

    # 合计行
    total_row = len(label_counts) + 7
    ws1.cell(row=total_row, column=1, value="合计").font = body_font
    ws1.cell(row=total_row, column=3, value=total).font = body_font
    ws1.cell(row=total_row, column=4, value=100.00).font = body_font

    # 列宽
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 2: 按天趋势
    ws2 = wb.create_sheet("按天趋势")
    trend_headers = ["日期", "标签代码", "标签名称", "数量"]
    for col, header in enumerate(trend_headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    for idx, (day, label, count) in enumerate(trend_rows, start=2):
        day_str = day.strftime("%Y-%m-%d") if hasattr(day, "strftime") else str(day)
        label_name = DIAG_LABEL_NAMES.get(label, label)
        ws2.cell(row=idx, column=1, value=day_str).font = body_font
        ws2.cell(row=idx, column=2, value=label).font = body_font
        ws2.cell(row=idx, column=3, value=label_name).font = body_font
        ws2.cell(row=idx, column=4, value=count).font = body_font
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # Sheet 3: 明细数据
    ws3 = wb.create_sheet("明细数据")
    detail_headers = ["回路ID", "标签代码", "标签名称", "置信度", "诊断时间", "算法版本"]
    for col, header in enumerate(detail_headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    for idx, row in enumerate(detail_rows, start=2):
        loop_id, label, confidence, diagnosed_at, algo_ver = row
        label_name = DIAG_LABEL_NAMES.get(label, label)
        conf_val = float(confidence) if confidence is not None else 0.0
        diag_time = (
            diagnosed_at.strftime("%Y-%m-%d %H:%M:%S")
            if hasattr(diagnosed_at, "strftime")
            else str(diagnosed_at)
        )
        ws3.cell(row=idx, column=1, value=str(loop_id)).font = body_font
        ws3.cell(row=idx, column=2, value=label).font = body_font
        ws3.cell(row=idx, column=3, value=label_name).font = body_font
        ws3.cell(row=idx, column=4, value=conf_val).font = body_font
        ws3.cell(row=idx, column=5, value=diag_time).font = body_font
        ws3.cell(row=idx, column=6, value=str(algo_ver or "")).font = body_font
    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # 写入内存缓冲区
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _generate_csv_bytes(
    start_time: str,
    end_time: str,
    plant_name: str,
    label_counts: list,
    trend_rows: list,
    detail_rows: list,
) -> bytes:
    """生成 CSV 格式报表（UTF-8 with BOM）。"""
    import csv

    from app.services.diagnosis import DIAG_LABEL_NAMES

    buffer = io.StringIO()
    buffer.write("\ufeff")
    writer = csv.writer(buffer)

    writer.writerow(["CLPM 诊断统计报表"])
    writer.writerow(["时间范围", f"{start_time} ~ {end_time}"])
    writer.writerow(["装置范围", plant_name])
    writer.writerow(["导出时间", datetime.now(UTC).replace(tzinfo=None).isoformat()])
    writer.writerow([])

    writer.writerow(["一、标签分布汇总"])
    writer.writerow(["标签代码", "标签名称", "数量", "占比(%)"])
    total = sum(row[1] for row in label_counts) if label_counts else 0
    for label, count in label_counts:
        label_name = DIAG_LABEL_NAMES.get(label, label)
        ratio = (count / total * 100) if total > 0 else 0.0
        writer.writerow([label, label_name, count, f"{ratio:.2f}"])
    writer.writerow([])

    writer.writerow(["二、按天趋势"])
    writer.writerow(["日期", "标签代码", "标签名称", "数量"])
    for day, label, count in trend_rows:
        day_str = day.strftime("%Y-%m-%d") if hasattr(day, "strftime") else str(day)
        label_name = DIAG_LABEL_NAMES.get(label, label)
        writer.writerow([day_str, label, label_name, count])
    writer.writerow([])

    writer.writerow(["三、明细数据"])
    writer.writerow(["回路ID", "标签代码", "标签名称", "置信度", "诊断时间", "算法版本"])
    for row in detail_rows:
        loop_id, label, confidence, diagnosed_at, algo_ver = row
        label_name = DIAG_LABEL_NAMES.get(label, label)
        conf_val = float(confidence) if confidence is not None else 0.0
        diag_time = (
            diagnosed_at.strftime("%Y-%m-%d %H:%M:%S")
            if hasattr(diagnosed_at, "strftime")
            else str(diagnosed_at)
        )
        writer.writerow([str(loop_id), label, label_name, conf_val, diag_time, str(algo_ver or "")])

    csv_str = buffer.getvalue()
    buffer.close()
    return csv_str.encode("utf-8")


def _parse_iso_dt(s: str) -> datetime:
    """解析 ISO 8601 时间字符串为 naive datetime。

    兼容带 Z 后缀、带时区偏移、不带时区三种格式，统一剥离 tzinfo。
    PostgreSQL 列为 TIMESTAMP WITHOUT TIME ZONE，asyncpg 不允许
    tz-aware datetime 传入 naive 列。
    """
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        dt = datetime.fromisoformat(s)
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    return dt


__all__ = [
    "AsyncTask",
    "NonRetryableError",
    "export_diagnosis_statistics",
    "generate_diagnosis_pdf_task",
    "generate_report_task",
]

# ---------------------------------------------------------------------------
# V62-P3-33：诊断建议书 PDF 异步导出（修复大回路 PDF 超时网关 504）
# ---------------------------------------------------------------------------

# Redis TaskTracker 任务 Hash 的 key（与 task_tracker._task_key 保持一致）
_TASK_KEY_PREFIX = "task:"
# 产物路径过期时间：7 天（导出文件只作为短期下载缓存）
_REPORT_ARTIFACT_TTL_SEC = 7 * 24 * 3600


@celery_app.task(
    name="app.tasks.report_generator.generate_diagnosis_pdf_task",
    bind=True,
    base=AsyncTask,
    autoretry_for=(Exception,),
    retry_kwargs={"max_retries": 2, "countdown": 30},
    retry_backoff=True,
    retry_backoff_max=300,
    retry_jitter=True,
)
def generate_diagnosis_pdf_task(
    self: AsyncTask,
    tracker_task_id: str,
    loop_id: str,
    variant: str = "tracker_export",
    tag_codes: list[str] | None = None,
) -> dict[str, Any]:
    """诊断建议书 PDF 异步生成任务.

    Args:
        tracker_task_id: TaskTracker 创建的任务 ID（REPORT 类型），进度通过它回写
        loop_id: 目标回路 ID
        variant: "diagnosis_report"（POST /diagnosis/{id}/report 对应，文件名含
            diagnosis_report）或 "tracker_export"（POST /tracker/{id}/export 对应）
        tag_codes: 诊断建议书 variant 下用户勾选的推荐方案标签代码；None 表示
            取该回路下所有推荐

    Progress（前端进度条按此对齐）:
        0.25 - 加载回路信息 + 诊断快照
        0.50 - 加载推荐方案
        0.75 - PDF 字节生成完成，正在写入导出目录
        1.00 - 落盘完成，补写 file_path/file_name/result_url 到 Redis Hash
    """
    logger.info(
        "[P3-33] PDF 异步导出任务启动 tracker_task_id=%s loop=%s variant=%s",
        tracker_task_id,
        loop_id,
        variant,
    )
    try:
        result = self.run_async(
            _do_generate_diagnosis_pdf_task(
                tracker_task_id=tracker_task_id,
                loop_id=loop_id,
                variant=variant,
                tag_codes=tag_codes,
            )
        )
        logger.info("[P3-33] PDF 异步导出任务完成: %s", result)
        return result
    except Exception:
        logger.exception("[P3-33] PDF 异步导出任务失败 tracker_task_id=%s", tracker_task_id)
        # 尽力而为写 FAILED + error_message，前端轮询不会卡死
        self.run_async(_maybe_mark_failed(tracker_task_id=tracker_task_id))
        raise


async def _maybe_mark_failed(tracker_task_id: str) -> None:
    """任务异常兜底：标记 TaskTracker 为 FAILED，避免前端轮询挂死."""
    try:
        from app.services.task_tracker import TaskStatus, update_status

        await update_status(
            tracker_task_id,
            TaskStatus.FAILED,
            progress=0.0,
            error_message="任务执行异常，详见服务端日志",
            finished_at=datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds"),
        )
    except Exception:  # pragma: no cover - 兜底层再失败就只能记日志
        logger.exception("mark PDF task FAILED 失败 task_id=%s", tracker_task_id)


async def _do_generate_diagnosis_pdf_task(
    *,
    tracker_task_id: str,
    loop_id: str,
    variant: str,
    tag_codes: list[str] | None,
) -> dict[str, Any]:
    """真正执行 PDF 生成逻辑（async，通过 AsyncTask.run_async 桥接）."""
    from app.core.db import AsyncSessionLocal
    from app.schemas.task import TaskStatus
    from app.services.diagnosis import get_diagnosis_detail
    from app.services.diagnosis_recommendation import (
        get_recommendations,
        get_recommendations_for_loop,
    )
    from app.services.diagnosis_report import generate_diagnosis_report
    from app.services.task_tracker import update_status
    from app.services.tracker import export_tracker_pdf  # noqa: F401 - 语义完整性留痕

    started_at = datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds")

    # --- 25%：查回路诊断快照 ------------------------------------------------
    await update_status(
        tracker_task_id,
        TaskStatus.RUNNING,
        progress=0.25,
        started_at=started_at,
        current_stage="加载回路信息与诊断快照",
    )
    async with AsyncSessionLocal() as db:
        snapshot_data = await get_diagnosis_detail(db=db, loop_id=loop_id)

        # --- 50%：加载推荐方案 ----------------------------------------------
        await update_status(
            tracker_task_id,
            TaskStatus.RUNNING,
            progress=0.50,
            current_stage="获取整改推荐方案",
        )
        if variant == "diagnosis_report" and tag_codes:
            recommendations = get_recommendations(loop_id, tag_codes)
        else:
            recommendations = await get_recommendations_for_loop(db=db, loop_id=loop_id)

        # variant=tracker_export 需要先查 tag_name 和生成统一文件名
        from sqlalchemy import select

        from app.models.loop import LoopLedger

        loop_res = await db.execute(select(LoopLedger).where(LoopLedger.id == loop_id))
        loop = loop_res.scalar_one_or_none()
        tag_name = loop.tag_name if loop and loop.tag_name else loop_id

    if variant == "tracker_export":
        # --- 75%：生成 PDF bytes ------------------------------------------------
        await update_status(
            tracker_task_id,
            TaskStatus.RUNNING,
            progress=0.75,
            current_stage="生成诊断建议书 PDF",
        )
        pdf_bytes = generate_diagnosis_report(
            loop_id=loop_id,
            snapshot_data=snapshot_data,
            recommendations=recommendations,
        )
        date_str = datetime.now(UTC).replace(tzinfo=None).strftime("%Y-%m-%d")
        file_name = f"CLPM-诊断建议书-{tag_name}-{date_str}.pdf"
    else:  # diagnosis_report
        await update_status(
            tracker_task_id,
            TaskStatus.RUNNING,
            progress=0.75,
            current_stage="生成诊断快照报告 PDF",
        )
        pdf_bytes = generate_diagnosis_report(
            loop_id=loop_id,
            snapshot_data=snapshot_data,
            recommendations=recommendations,
        )
        file_name = f"diagnosis_report_{loop_id}.pdf"

    # --- 落盘 + 100%：补写 Redis Hash 字段 ------------------------------------
    await update_status(
        tracker_task_id,
        TaskStatus.RUNNING,
        progress=0.95,
        current_stage="写入导出文件",
    )
    export_dir = os.environ.get("CLPM_EXPORT_DIR", tempfile.gettempdir())
    os.makedirs(export_dir, exist_ok=True)
    # 文件名追加 tracker_task_id 前缀避免同回路同日导出互相覆盖
    prefixed_name = f"{tracker_task_id}_{file_name}"
    file_path = os.path.join(export_dir, prefixed_name)
    with open(file_path, "wb") as f:
        f.write(pdf_bytes)

    # file_path（内部，不返回给前端）/ file_name（下载时保存名）
    # / result_url（统一下载入口）一并写进 TaskTracker Redis Hash
    from app.services.task_tracker import redis_client as task_redis

    hash_key = f"{_TASK_KEY_PREFIX}{tracker_task_id}"
    # 让 TTL 与 tasks endpoint 保持匹配：至少存在 7 天
    await task_redis.hset(
        hash_key,
        mapping={
            "file_name": file_name,
            "file_path": file_path,
            "result_url": f"/api/v1/tasks/{tracker_task_id}/download",
        },
    )
    try:
        await task_redis.expire(hash_key, _REPORT_ARTIFACT_TTL_SEC)
    except Exception:  # pragma: no cover - expire 失败不阻断任务
        logger.warning("设置 task hash TTL 失败 task=%s", tracker_task_id)

    finished_at = datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds")
    await update_status(
        tracker_task_id,
        TaskStatus.SUCCESS,
        progress=1.0,
        finished_at=finished_at,
        current_stage="文件已生成",
    )
    return {
        "taskId": tracker_task_id,
        "status": "SUCCESS",
        "fileName": file_name,
        "resultUrl": f"/api/v1/tasks/{tracker_task_id}/download",
        "fileSize": len(pdf_bytes),
    }
